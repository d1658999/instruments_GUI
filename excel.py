import pathlib

import openpyxl

from openpyxl.chart import LineChart, Reference, ScatterChart, Series


def create_excel():

    wb = openpyxl.Workbook()
    wb.remove(wb['Sheet'])
    pwr_q_1 = wb.create_sheet('pwr_q_1')
    pwr_q_p = wb.create_sheet('pwr_q_p')
    pwr_q_f = wb.create_sheet('pwr_q_f')
    pwr_16_p = wb.create_sheet('pwr_16_p')
    pwr_16_f = wb.create_sheet('pwr_16_f')
    pwr_64_p = wb.create_sheet('pwr_64_p')
    pwr_64_f = wb.create_sheet('pwr_64_f')
    aclr_q_1 = wb.create_sheet('aclr_q_1')
    aclr_q_p = wb.create_sheet('aclr_q_p')
    aclr_q_f = wb.create_sheet('aclr_q_f')
    aclr_16_p = wb.create_sheet('aclr_16_p')
    aclr_16_f = wb.create_sheet('aclr_16_f')
    aclr_64_p = wb.create_sheet('aclr_64_p')
    aclr_64_f = wb.create_sheet('aclr_64_f')
    evm_q_1 = wb.create_sheet('evm_q_1')
    evm_q_p = wb.create_sheet('evm_q_p')
    evm_q_f = wb.create_sheet('evm_q_f')
    evm_16_p = wb.create_sheet('evm_16_p')
    evm_16_f = wb.create_sheet('evm_16_f')
    evm_64_p = wb.create_sheet('evm_64_p')
    evm_64_f = wb.create_sheet('evm_64_f')
    for sheet in wb.sheetnames:
        if 'aclr' in sheet:
            sh = wb[sheet]
            sh['A1'] = 'Band'
            sh['B1'] = 'Channel'
            sh['C1'] = 'EUTRA_-1'
            sh['D1'] = 'EUTRA_+1'
            sh['E1'] = 'UTRA_-1'
            sh['F1'] = 'UTRA_+1'
            sh['G1'] = 'UTRA_-2'
            sh['H1'] = 'UTRA_+2'

        else:
            sh = wb[sheet]
            sh['A1'] = 'Band'
            sh['B1'] = 'ch0'
            sh['C1'] = 'ch1'
            sh['D1'] = 'ch2'
    wb.save('test.xlsx')

def fill_values():
    wb = openpyxl.load_workbook('test.xlsx')
    ws = wb['pwr_q_p']

    if ws.max_row == 1:
        ws.cell(2, 1).value = 2
        ws.cell(2, 2).value = 10
        print('only title, and add the new band')
    else:
        for row in range(2, ws.max_row + 1):
            if ws.cell(row, 1).value == 2:
                ws.cell(row, 3).value = 20
                print('find the band I want ')
            elif ws.cell(row, 1).value != 2 and row == ws.max_row:
                ws.cell(row+1, 1).value = 2
                ws.cell(row+1, 3).value = 30
                print('cannot find the band and the row is final row')
            else:
                print('continue to search')
                continue



    # for i, row_cells in enumerate(ws.iter_rows()):
    #     if i == 0:
    #         continue
    #     for cell in row_cells:
    #         print(f'{cell},  {cell.value}')


    wb.save('test.xlsx')
    wb.close()






def main():
    # create_excel()
    # fill_values()
    wb = openpyxl.load_workbook('results_5MHZ_LTE.xlsx')
    ws = wb['PWR_Q_P']

    c1 = LineChart()
    c1.title = 'Power'
    c1.y_axis.title = 'ACLR(dB)'
    c1.x_axis.title = 'Band'
    c1.x_axis.tickLblPos = 'low'

    c1.height = 20
    c1.width = 40


    y_data = Reference(ws, min_col=2, min_row=1, max_col=ws.max_column, max_row=ws.max_row)
    x_data = Reference(ws, min_col=1, min_row=2, max_col=1, max_row=ws.max_row)
    c1.add_data(y_data, titles_from_data=True)
    c1.set_categories(x_data)


    ws.add_chart(c1, "F1")

    wb.save("results_5MHZ_test.xlsx")
    wb.close()



if __name__ == '__main__':
    main()