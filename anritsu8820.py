import pathlib

import pyvisa
import time
import datetime
import logging
from logging.config import fileConfig
import openpyxl
from decimal import Decimal
from openpyxl.chart import LineChart, Reference, ScatterChart

from loss_table import loss_table
import common_parameters as cm_pmt
import want_test_band as wt
from fly_mode import Flymode, get_comport_wanted

fileConfig('logging.ini')
logger = logging.getLogger()


class Anritsu8820(pyvisa.ResourceManager):
    def __init__(self):
        self.excel_path = None
        self.count = 5
        self.pwr = None
        self.aclr = None
        self.evm = None
        self.std = None
        self.mod = None
        self.bw = None
        self.band = None
        self.dl_ch = None
        self.chcoding = None
        try:
            self.build_object()
        except:
            logger.debug('Error to connect to instrument')

    def get_gpib(self):
        resources = []
        for resource in super().list_resources():
            if 'GPIB' in resource:
                resources.append(resource)
                logger.debug(resource)
        return resources

    def build_object(self):
        logger.info('start to connect')
        for gpib in self.get_gpib():  # this is to search GPIB for 8820/8821
            self.inst = super().open_resource(gpib)
            inst = self.inst.query('*IDN?').strip()
            if '8820' in inst or '8821' in inst:
                self.gpib = gpib

        self.inst = super().open_resource(self.gpib)  # to build inst object
        self.inst.timeout = 5000
        self.comport = get_comport_wanted()
        self.flymode = Flymode(self.comport)
        logger.debug(self.inst.query('*IDN?').strip())

    def flymode_circle(self):
        self.flymode.com_open()
        self.flymode.fly_on()
        time.sleep(3)
        self.flymode.fly_off()
        self.flymode.com_close()

    def query_standard(self):
        return self.inst.query("STDSEL?").strip()

    def switch_to_wcdma(self):
        """
            switch to WCDMA mode
            switch ok => return 0
            switch fail => return 1
        """
        self.inst.write('CALLSO')
        time.sleep(1)
        self.std = self.query_standard()  # WCDMA|GSM|LTE
        logger.debug("Current Function: " + self.std)
        if self.std == 'WCDMA':
            logger.info("Already WCDMA mode")
            return self.std
        else:
            self.inst.write('STDSEL WCDMA')  # switch to WCDMA
            time.sleep(1)
            self.std = self.query_standard()
            if self.std == 'WCDMA':
                logger.info("Switch to WCDMA mode OK")
                return self.std
            else:
                logger.info("Switch to WCDMA mode fail")
                return 1

    def switch_to_gsm(self):
        """
            switch to GSM mode
            switch ok => return 0
            switch fail => return 1
        """
        self.inst.write('CALLSO')
        time.sleep(1)
        self.std = self.query_standard()  # WCDMA|GSM|LTE
        logger.debug("Current Format: " + self.std)
        if self.std == 'GSM':
            logger.info("Already GSM mode")
            return self.std
        else:
            self.inst.write('STDSEL GSM')  # switch to GSM
            time.sleep(1)
            self.std = self.query_standard()
            if self.std == "GSM":
                logger.info("Switch to GSM mode OK")
                return self.std
            else:
                logger.info("Switch to GSM mode fail")
                return 1

    def switch_to_lte(self):
        """
            switch to LTE mode
            switch ok => return 0
            switch fail => return 1
        """
        self.inst.write('CALLSO')
        time.sleep(2)
        self.std = self.query_standard()  # WCDMA|GSM|LTE
        logger.info("Current Format: " + self.std)
        if self.std == 'LTE':
            logger.info("Already LTE mode")
            return self.std
        else:
            self.inst.write('STDSEL LTE')  # switch to LTE
            time.sleep(1)
            self.std = self.query_standard()
            if self.std == 'LTE':
                logger.info("Switch to LTE mode OK")
                return self.std
            else:
                logger.info("Switch to LTE mode fail")
                return 1

    def preset(self):
        """
            preset Anritsu 8820C
        """
        logger.info("Preset Anritsu 8820/8821")
        self.inst.write('*ESR?')
        self.inst.write('CALLSO')
        s = self.query_standard()  # WCDMA|GSM|LTE|CDMA2K
        if s == 'WCDMA':
            self.inst.write('CALLSO')
            time.sleep(2)
            self.preset_3gpp()
            self.set_lvl_status('OFF')
        else:
            self.inst.write("*RST")  # this command changes measurement count to "single"
        self.inst.write("*CLS")
        self.inst.write("ALLMEASITEMS_OFF")  # Set all measurements to off

    def preset_3gpp(self):
        """
            preest to 3GPP spec (for WCDMA)
        """
        self.inst.write("PRESET_3GPP")

    def preset_extarb(self):
        self.inst.write('PRESET_EXTARB')

    def set_test_parameter_normal(self):
        self.inst.write('TESTPRM NORMAL')
        self.inst.write('ULRMC_64QAM DISABLED')

    def set_init_before_calling(self, standard, dl_ch, bw=5):
        logger.info('init equipment before calling')
        s = standard
        if s == 'LTE':
            self.set_init_before_calling_lte(dl_ch, bw)
        elif s == 'WCDMA':
            self.set_init_before_calling_wcdma(dl_ch)
        elif s == 'GSM':
            pass

    def set_init_before_calling_wcdma(self, dl_ch):
        """
            preset before start to calling for WCDMA
        """
        self.set_band_cal()
        self.preset()
        self.set_integrity('WCDMA', 'ON')
        self.set_screen('OFF')
        self.set_display_remain()
        self.set_init_miscs('WCDMA')
        self.set_test_mode('OFF')
        self.set_imsi()
        self.set_authentication('WCDMA')
        self.set_all_measurement_items_off()
        self.set_path_loss('WCDMA')
        self.set_init_level('WCDMA')
        self.set_handover('WCDMA', dl_ch)
        self.inst.query('*OPC?')

    def set_init_before_calling_lte(self, dl_ch, bw):
        """
            preset before start to calling for LTE
        """
        self.preset()
        self.set_band_cal()
        self.set_screen('OFF')
        self.set_display_remain()
        self.preset_extarb()
        self.set_lvl_status('OFF')
        # if 53 >= band >= 33:
        #     self.set_fdd_tdd_mode('FDD')
        # else:
        #     self.set_fdd_tdd_mode('TDD')
        self.set_test_mode('OFF')
        self.set_integrity('LTE', 'SNOW3G')
        self.set_scenario()
        self.set_pdn_type()
        self.set_mcc_mnc()
        self.set_ant_config()
        self.set_imsi()
        self.set_authentication()
        self.set_all_measurement_items_off()
        self.set_init_miscs('LTE')
        self.set_path_loss('LTE')
        self.set_init_level('LTE')
        self.set_handover('LTE', dl_ch, bw)
        self.inst.write('ULRB_POS MIN')
        self.inst.query('*OPC?')

    def set_path_loss(self, standard):
        logger.info('Set LOSS')
        self.inst.write('DELLOSSTBL')  # delete the unknown loss table first

        loss_title = 'LOSSTBLVAL'
        freq = sorted(loss_table.keys())
        for keys in freq:
            loss = f'{loss_title} {str(keys)}MHZ, {str(loss_table[keys])}, {str(loss_table[keys])}, {str(loss_table[keys])} '
            logger.info(loss)
            self.inst.write(loss)
        s = standard  # WCDMA|GSM|LTE
        logger.debug("Current Format: " + s)
        if s == 'LTE':
            self.inst.write("EXTLOSSW COMMON")
            self.inst.query('*OPC?')
        elif s == 'WCDMA':
            self.inst.write("DLEXTLOSSW COMMON")  # Set DL external loss to COMMON
            self.inst.write("ULEXTLOSSW COMMON")  # Set UL external loss to COMMON
            self.inst.write("AUEXTLOSSW COMMON")  # Set AUX external loss to COMMON
            self.inst.query('*OPC?')
        elif s == "GSM":
            self.inst.write("EXTLOSSW COMMON")
            self.inst.query('*OPC?')

    def set_init_level(self, standard):
        """
            LTE:
                initial input_level=5 and output_level=-60
            WCDMA:
                initial input_level=5 and output_level=-75
        """
        s = standard
        if s == 'LTE':
            self.set_input_level()
            self.set_output_level()
            self.inst.query('*OPC?')
        elif s == 'WCDMA':
            self.set_input_level()
            self.set_output_level(-50)
            self.inst.query('*OPC?')
        elif s == 'GSM':
            pass

    def set_registration_calling(self, standard):
        logger.info('Start to calling')
        s = standard
        logger.debug("Current Format: " + s)
        logger.info('Start registration and calling')
        if s == 'LTE':
            self.set_registration_calling_lte()
        elif s == 'WCDMA' and self.chcoding == 'REFMEASCH':  # this is WCDMA
            self.set_registration_calling_wcdma()
        elif s == 'GSM':
            pass
        elif s == 'WCDMA' and self.chcoding == 'EDCHTEST':  # this is HSUPA
            self.set_registration_calling_hspa()
            # self.set_registration_after_calling_hspa()

        elif s == 'WCDMA' and self.chcoding == 'FIXREFCH':  # this is HSDPA
            self.set_registration_calling_hspa()
            # self.set_registration_after_calling_hspa()

    def set_registration_calling_lte(self, times=30):
        """
            ANRITSU_IDLE = 1	        #Idle state
            ANRITSU_REGIST = 3			# Under location registration
            ANRITSU_CONNECTED = 6	    # Under communication or connected
        """
        self.inst.write('CALLTHLD 1')
        self.set_lvl_status('ON')
        self.set_test_mode()
        conn_state = int(self.inst.query("CALLSTAT?").strip())
        while conn_state != cm_pmt.ANRITSU_CONNECTED:  # this is for waiting connection
            self.inst.write('CALLRFR')
            while conn_state == cm_pmt.ANRITSU_IDLE:
                self.inst.write('CALLSO')
                time.sleep(5)
                logger.info('IDLE')
                logger.info('Start to ON and OFF')
                self.flymode_circle()
                logger.info('Waiting for 10 seconds')
                time.sleep(10)
                logger.info('Start calling')
                conn_state = int(self.inst.query("CALLSTAT?").strip())
            logger.info('START CALL')
            self.inst.write('CALLSA')
            logger.info('Connected')
            time.sleep(1)

    def set_registration_calling_wcdma(self):
        """
            ANRITSU_IDLE = 1	        #Idle state
            ANRITSU_IDLE_REGIST = 2		#Idle( Regist ) Idle state (location registered)
            ANRITSU_LOOP_MODE_1 = 7	    # Under communication or connected
            ANRITSU_LOOP_MODE_1_CLOSED = 9  # it seems like waiting to loop mode between Loopback mode 1 and IDLE
        """
        self.set_lvl_status('ON')
        self.set_test_mode()
        # self.flymode_circle()
        conn_state = int(self.inst.query("CALLSTAT?").strip())

        self.count = 10
        while conn_state != cm_pmt.ANRITSU_LOOP_MODE_1:  # this is for waiting connection
            if conn_state == cm_pmt.ANRITSU_IDLE:
                self.inst.write('CALLSO')
                logger.info('IDLE')
                time.sleep(10)
                logger.info('START CALL')
                self.flymode_circle()
                time.sleep(10)
                self.inst.write('CALLSA')
                conn_state = int(self.inst.query("CALLSTAT?").strip())

            elif conn_state == cm_pmt.ANRITSU_IDLE_REGIST:
                logger.info('Status: IDLE_REGIST')
                self.inst.write('CALLSA')
                time.sleep(1)
                conn_state = int(self.inst.query("CALLSTAT?").strip())

            elif conn_state == cm_pmt.ANRITSU_REGIST:
                logger.info('Status: REGIST')
                time.sleep(1)
                conn_state = int(self.inst.query("CALLSTAT?").strip())

            elif conn_state == cm_pmt.ANRITSU_LOOP_MODE_1_CLOSE:
                if self.count < 0:
                    logger.info('END CALL and FLY ON and OFF for WCDMA')
                    self.set_init_before_calling_wcdma(self.dl_ch)
                    self.inst.write('CALLSO')
                    self.set_test_mode()
                    self.set_lvl_status('ON')
                    time.sleep(10)
                    self.flymode_circle()
                    time.sleep(10)
                    self.inst.write('CALLSA')
                    self.count = 10
                    conn_state = int(self.inst.query("CALLSTAT?").strip())
                else:
                    logger.info('Status: LOOP MODE(CLOSE)')
                    time.sleep(2)
                    conn_state = int(self.inst.query("CALLSTAT?").strip())
                    self.count -= 1

        logger.info('Loop mode 1 and connected')

    def set_registration_calling_hspa(self):
        """
            ANRITSU_IDLE = 1	        #Idle state
            ANRITSU_IDLE_REGIST = 2		#Idle( Regist ) Idle state (location registered)
            ANRITSU_LOOP_MODE_1 = 7	    # Under communication or connected
            ANRITSU_LOOP_MODE_1_CLOSED = 9  # it seems like waiting to loop mode between Loopback mode 1 and IDLE
        """
        self.set_lvl_status('ON')
        self.set_test_mode()
        self.set_init_hspa()
        # self.flymode_circle()

        conn_state = int(self.inst.query("CALLSTAT?").strip())

        self.count = 10
        while conn_state != cm_pmt.ANRITSU_LOOP_MODE_1:  # this is for waiting connection
            if conn_state == cm_pmt.ANRITSU_IDLE:
                self.inst.write('CALLSO')
                logger.info('IDLE')
                time.sleep(10)
                logger.info('START CALL')
                self.flymode_circle()
                time.sleep(10)
                self.inst.write('CALLSA')
                conn_state = int(self.inst.query("CALLSTAT?").strip())

            elif conn_state == cm_pmt.ANRITSU_IDLE_REGIST:
                logger.info('Status: IDLE_REGIST')
                self.inst.write('CALLSA')
                time.sleep(1)
                conn_state = int(self.inst.query("CALLSTAT?").strip())

            elif conn_state == cm_pmt.ANRITSU_REGIST:
                logger.info('Status: REGIST')
                time.sleep(1)
                conn_state = int(self.inst.query("CALLSTAT?").strip())

            elif conn_state == cm_pmt.ANRITSU_LOOP_MODE_1_CLOSE:
                if self.count < 0:
                    logger.info('END CALL and FLY ON and OFF for HSUPA')
                    self.set_init_before_calling_wcdma(self.dl_ch)
                    self.set_init_hspa()
                    self.inst.write('CALLSO')
                    self.set_test_mode()
                    self.set_lvl_status('ON')
                    time.sleep(10)
                    self.flymode_circle()
                    time.sleep(10)
                    self.inst.write('CALLSA')
                    conn_state = int(self.inst.query("CALLSTAT?").strip())
                    self.count = 10
                else:
                    logger.info('Status: LOOP MODE(CLOSE)')
                    time.sleep(2)
                    conn_state = int(self.inst.query("CALLSTAT?").strip())
                    self.count -= 1

        logger.info('Loop mode 1 and connected')
        time.sleep(2)

        # self.set_registration_after_calling_hspa()

    def set_registration_after_calling_hspa(self):
        if self.chcoding == 'EDCHTEST':  # this is HSUPA
            self.set_registration_after_calling_hsupa()

        elif self.chcoding == 'FIXREFCH':  # this is HSDPA
            self.set_registration_after_calling_hsdpa()

    def set_registration_after_calling_hsdpa(self):
        self.inst.write('SCRSEL FMEAS')
        self.inst.write('SET_PWRPAT HSMAXPWR')
        self.set_init_power(1)
        self.set_init_aclr('WCDMA', 1)

    def set_registration_after_calling_hsupa(self):
        self.inst.write('SCRSEL FMEAS')
        self.inst.write('SET_PWRPAT HSMAXPWR')
        self.set_output_level(-86)
        self.set_init_power(1)
        self.set_init_aclr('WCDMA', 1)
        self.inst.write('TPUTU_MEAS ON')
        self.inst.write('TPUTU_SAMPLE 15')
        self.inst.write('EHICHPAT ACK')

    def set_disconnected(self):
        self.inst.write('CALLSO')
        self.inst.write('CALLPROC OFF')
        self.inst.query('*OPC?')
        logger.info('DISCONNECTED')

    def set_handover(self, standard, dl_ch, bw=5):
        s = standard  # WCDMA|GSM|LTE
        logger.debug("Current Format: " + s)
        if s == 'LTE':
            self.set_bandwidth(bw)
            self.set_downlink_channel(s, dl_ch)
            # ul_ch = self.inst.query('ULCHAN?')
            self.inst.query('*OPC?')
            # if ul_ch != dl_ch:
            #     self.set_fdd_tdd_mode('FDD')
            # elif ul_ch == dl_ch:
            #     self.set_fdd_tdd_mode('TDD')
            # else:
            #     print('comparison between ul_ch and dl_ch seems like error!')
        elif s == 'WCDMA' or s == 'GSM':
            self.set_downlink_channel(s, dl_ch)
            self.inst.query('*OPC?')
        else:
            logger.info('Standard switch @handover function seems like error!')

    def set_fdd_tdd_mode(self, fdd_tdd):
        self.inst.write(f'FRAMETYPE {fdd_tdd}')  # FDD|TDD

    def set_test_mode(self, on_off='ON'):  # this is to set call progress ON
        self.inst.write(f"CALLPROC {on_off}")  # default = ON | OFF for signaling | non-signaling

    def set_imsi(self, imsi=cm_pmt.IMSI):
        self.inst.write(f'IMSI {imsi}')

    # def set_imei(self, IMEI):
    #     self.inst.write(IMEI)

    def set_screen(self, on_off='ON'):
        self.inst.write(f'SCREEN {on_off}')

    def set_display_remain(self):
        self.inst.write('REMDISP REMAIN')

    def set_lvl_status(self, on_off):
        self.inst.write(f'LVL {on_off}')

    def set_bandwidth(self, bw=5):
        self.inst.write(f'BANDWIDTH {str(bw)}MHZ')

    def set_scenario(self, mode='NORMAL'):
        self.inst.write(f'SCENARIO {mode}')  # default = NORMAL

    def set_pdn_type(self):
        self.inst.write('PDNTYPE AUTO')

    def set_mcc_mnc(self, mcc='001', mnc='01'):
        self.inst.write(f'MCC {mcc}')
        self.inst.write(f'MNC {mnc} ')

    def set_ant_config(self, ant='SINGLE'):
        self.inst.write(f'ANTCONFIG {ant}')  # default =  SINGLE | RX_DIVERSITY | OPEN_LOOP | CLOSED_LOOP_MULTI

    def set_band_cal(self):
        self.inst.query('BANDCAL_TEMP 2.0;*OPC?')

    def set_integrity(self, standard, status):
        s = standard
        if s == 'LTE':
            self.inst.write(f'INTEGRITY {status}')  # SNOW3G | NULL | OFF
        elif s == 'WCDMA':
            self.inst.write(f'INTEGRITY {status}')  # ON | OFF
        elif s == 'GSM':
            pass

    def set_authentication(self, standard='LTE'):
        s = standard
        if s == 'LTE':
            self.inst.write('AUTHENT ON')
            self.inst.write('AUTHENT_ALGO XOR')
            self.inst.write('AUTHENT_KEYALL 00112233,44556677,8899AABB,CCDDEEFF')
            self.inst.write('OPC_ALL 00000000,00000000,00000000,00000000')
        elif s == 'WCDMA':
            self.inst.write('AUTHENT_ALGO XOR')
            self.inst.write('AUTHENT_KEYALL 00112233,44556677,8899AABB,CCDDEEFF')
            self.inst.write('OPC_ALL 00000000,00000000,00000000,00000000')

    def set_init_miscs(self, standard):
        s = standard
        if s == 'LTE':
            self.inst.write('RRCUPDATE PAGING')
            self.inst.write('PT_TRGSRC FRAME')
            self.inst.write('MODIFPERIOD N2')
            self.inst.write('PCYCLE 32')
            self.inst.write('RRCRELEASE OFF')
            self.inst.write('FREQERRRNG NARROW')  # NORMAL | NARROW
            self.inst.write('ROBUSTCON OFF')  # ON | OFF
            self.inst.write('TESTMODE OFF')
            self.inst.write('UECAT CAT3')
            self.inst.write('SIB2_NS NS_01')

        elif s == 'WCDMA':
            self.inst.write('RFOUT MAIN')
            self.inst.write('BANDIND AUTO')
            # self.inst.write('ATTFLAG OFF')
            # self.inst.write('MEASREP OFF')
            self.inst.write('DRXCYCLNG 64')
            self.inst.write('BER_SAMPLE 10000')
            self.inst.write('CONF_MEAS ON')
            self.inst.write('RX_TIMEOUT 5')
            self.inst.write('DOMAINIDRMC CS')
            self.inst.write('REGMODE AUTO')

        elif s == 'GSM':
            pass

    def set_input_level(self, input_level=5):
        return self.inst.write(f'ILVL {str(input_level)}')

    def set_output_level(self, output_level=-60):
        return self.inst.write(f'OLVL {str(output_level)}')

    def set_tpc(self, tpc='ILPC'):
        """
        set UL target power control mode, default is ILPC(inner loop control)
        """
        self.inst.write(f'TPCPAT {tpc}')  # WCDMA default= ILPC | ALL1 | ALL0 | ALT | UCMD
        # self.inst.write(f'TPCPAT {tpc}')  # LTE default= AUTO | ALL3 |ALL1 | ALL0 | ALLM1| ALT | UCMD

    def set_uplink_channel(self, standard, ul_ch):
        """
            Use this function only in FDD test mode.
            For Anritsu8820C, it could be used in link mode
        """
        s = standard
        if s == 'LTE' or s == 'WCDMA':
            return self.inst.write(f'ULCHAN {str(ul_ch)}')

        elif s == 'GSM':
            pass

    def set_downlink_channel(self, standard, dl_ch):
        """
        Use this function only in FDD test mode
        For Anritsu8820C, it could be used in link mode
        """
        s = standard
        if s == 'LTE' or s == 'WCDMA':
            return self.inst.write(f'DLCHAN {str(dl_ch)}')
        elif s == 'GSM':
            pass

    def set_init_power(self, count=1):
        self.inst.write('PWR_MEAS ON')  # Set [Power Measurement] to [On]
        self.inst.write(f'PWR_AVG {count}')  # Set [Average Count] to [count] times

    def set_init_aclr(self, standard, count=1):
        s = standard
        if s == 'LTE':
            self.inst.write('ACLR_MEAS ON')  # Set [ACLR Measurement] to [On]
            self.inst.write(f'ACLR_AVG {count}')  # Set [ACLR Count] to [count] times
        elif s == 'WCDMA':
            self.inst.write('ADJ_MEAS ON')  # Set [ACLR Measurement] to [On]
            self.inst.write(f'ADJ_AVG {count}')  # Set [ACLR Count] to [count] times
        elif s == 'GSM':
            pass

    def set_init_sem(self, count=1):
        self.inst.write('SEM_MEAS ON')  # Set [SEM Measurement] to [On]
        self.inst.write(f'SEM_AVG {count}')  # Set [SEM Count] to [count] times

    def set_init_obw(self, count=20):
        self.inst.write('OBW_MEAS ON')  # Set [OBW Measurement] to [On]
        self.inst.write(f'OBW_AVG {count}')  # Set [OBW Count] to [count] times

    def set_init_mod(self, standard, count=1):
        """
        this for EVM usage and some other modulation
        Set the average measurement count to 16 times because the average for 16 timeslots is described in the standards
        for 6.5.2.1A PUSCH-EVM with exclusion period
        """
        s = standard
        if s == 'LTE':
            self.inst.write('MOD_MEAS ON')  # Set [MOD Measurement] to [On]
            self.inst.write(f'MOD_AVG {count}')  # Set [MOD Count] to [count] times
        elif s == 'WCDMA':
            self.inst.write('INC_ORGNOFS ON')  # set [EVM include Origin Offset] to [On]
            self.inst.write('MOD_MEAS ON')  # Set [MOD Measurement] to [On]
            self.inst.write(f'MOD_AVG {count}')  # Set [MOD Count] to [count] times
        elif s == 'GSM':
            pass

    def set_init_power_template(self, standard, count=1):
        s = standard
        if s == 'LTE':
            self.inst.write('PWRTEMP ON')  # Set [OBW Measurement] to [On]
            self.inst.write(f'PWRTEMP_AVG {count}')  # Set [OBW Count] to [count] times
        elif s == 'WCDMA':
            self.inst.write('PT_WDR ON')  # Set [OBW Measurement] to [On]
            self.inst.write(f'PT_WDR_AVG {count}')  # Set [OBW Count] to [count] times
        elif s == 'GSM':
            pass

    def set_all_measurement_items_off(self):
        self.inst.write("ALLMEASITEMS_OFF")

    def set_to_measure(self):
        """
            Anritsu8820 use 'SWP' to measure no matter what the test items are
        """
        self.inst.write('SWP')

    def set_rx_sample(self, sample=1000):
        self.inst.write(f'TPUT_SAMPLE {sample}')

    def set_init_rx(self, standard):
        s = standard
        if s == 'LTE':
            self.inst.write('TESTPRM RX_SENS')
            self.set_tpc('AUTO')
            self.set_input_level(5)
            self.set_output_level(-70)
            self.set_rx_sample(1000)
            self.inst.write('TPUT_EARLY ON')
            self.set_init_power()
            self.inst.write('MOD_MEAS OFF')

        elif s == 'WCDMA':
            self.set_input_level(5)
            self.set_output_level(-70)
            self.set_rx_sample(1000)
            self.inst.write('TPUT_EARLY OFF')
            self.set_init_power()
        elif s == 'GSM':
            pass

    def set_rb_location(self, band, bw):
        rb_num, rb_location = cm_pmt.special_uplink_config_sensitivity(band, bw)
        self.inst.write(f'ULRMC_RB {rb_num}')
        self.inst.write(f'ULRB_START {rb_location}')

    def set_init_hspa(self):
        if self.chcoding == 'EDCHTEST':  # this is HSUPA
            self.set_init_hsupa()
        elif self.chcoding == 'FIXREFCH':  # this is HSDPA
            self.set_init_hsdpa()

    def set_init_hsdpa(self):
        self.inst.write('CHCODING FIXREFCH')
        self.inst.write('DDPCHTOFS 6')
        self.inst.write('SET_PWRPAT HSMAXPWR')
        self.inst.write('REGMODE COMBINED')
        self.inst.write('DOMAINIDRMC CS')
        self.inst.write('AUTHENT_ALGO XOR')
        self.inst.write('HSHSET HSET1_QPSK')
        self.inst.write('TPUT_MEAS OFF')
        self.set_input_level(-10)
        self.inst.write('*OPC?')

    def set_init_hsupa(self):
        self.inst.write('CHCODING EDCHTEST')
        self.inst.write('DDPCHTOFS 6')
        self.inst.write('MAXULPWR 21')
        self.inst.write('TPCALGO 2')
        self.inst.write('DOMAINIDRMC CS')
        self.inst.write('AUTHENT_ALGO XOR')
        self.inst.write('HSUSET TTI10_QPSK')
        self.inst.write('*OPC?')

    def query_etfci(self):
        result = Decimal(self.inst.query('AVG_ETFCI?').strip())
        return result

    def preset_subtest1(self):
        if self.chcoding == 'EDCHTEST':  # this is HSUPA
            self.inst.write('SET_HSDELTA_CQI 8')
            self.inst.write('SET_HSSUBTEST SUBTEST1')
            self.set_tpc('ILPC')
            self.set_input_level(16)
            time.sleep(0.15)
            self.set_tpc('ALT')
            self.set_input_level(26)
            logger.debug('TPC DOWN')
            time.sleep(0.15)
            self.inst.write('TPC_CMD_DOWN')
            self.set_to_measure()
        elif self.chcoding == 'FIXREFCH':  # this is HSDPA
            self.set_input_level(24)
            self.set_output_level(-86)
            self.set_tpc('ALL1')
            self.inst.write('DTCHPAT PN9')
            self.inst.write('SET_HSDELTA_CQI 8')
            self.inst.write('SET_HSSUBTEST SUBTEST1')
            self.set_to_measure()

    def get_hsdpa_evm(self):
        """
        evm = [po, p1, p2, p3], phase_disc = [theta0, theta1]
        :return:
        """
        self.inst.write('DDPCHTOFS 6')
        # self.inst.write('CHCODING FIXREFCH')
        self.inst.write('SET_PWRPAT HSPC')
        self.inst.write('SET_HSDELTA_CQI 7')
        self.inst.write('SET_HSSUBTEST SUBTEST3')
        self.inst.write('OLVL -86.0')
        self.inst.write('DTCHPAT PN9')
        self.inst.write('SCRSEL TDMEAS')
        self.inst.write('MEASOBJ HSDPCCH_MA')
        self.inst.write('HSMA_ITEM EVMPHASE')
        self.inst.write('TDM_RRC OFF')
        self.set_input_level(35)
        self.inst.write('TPCPAT ALL1')
        time.sleep(0.3)
        self.inst.write('TPCPAT ALT')
        self.set_to_measure()
        evm_hpm = self.inst.query('POINT_EVM? ALL').strip().split(',')   # p0, p1, p2, p3
        logger.debug(evm_hpm)
        phase_disc_hpm = self.inst.query('POINT_PHASEDISC? ALL').strip().split(',')    # theta0, theta1
        logger.debug(phase_disc_hpm)

        # below is for LPM -18dBm
        # self.set_tpc('ILPC')
        # self.inst.write('HSSCCH OFF')
        # self.inst.write('CQIFEEDBACK 0')
        # self.set_input_level(-18)
        # time.sleep(1)
        # self.set_tpc('ALT')
        # self.inst.write('HSSCCH ON')
        # self.inst.write('CQIFEEDBACK 4')
        # self.set_input_level(-10)
        # self.set_to_measure()
        # evm_lpm = self.inst.query('POINT_EVM? ALL').strip().split(',')  # p0, p1, p2, p3
        # logger.debug(evm_lpm)
        # phase_disc_lpm = self.inst.query('POINT_PHASEDISC? ALL').strip().split(',')  # theta0, theta1
        # logger.debug(phase_disc_lpm)

        return evm_hpm, phase_disc_hpm

    @staticmethod
    def get_worse_phase_disc(phase_disc):
        [temp1, temp2] = phase_disc
        phase_disc = [Decimal(x) for x in phase_disc]
        phase_disc_worst = max(list(map(abs, phase_disc)))
        if abs(Decimal(temp1)) == phase_disc_worst:
            return Decimal(temp1)
        elif abs(Decimal(temp2)) == phase_disc_worst:
            return Decimal(temp2)


    def get_subtest1_power_aclr(self):
        """
        :return: power, ACLR, subtest_number for HSUPA
        :return: power, ACLR, EVM, subtest_number for HSDPA
        """
        logger.info('Start to subtest1')
        if self.chcoding == 'EDCHTEST':  # this is HSUPA
            logger.info('start to measure HSUPA')
            self.preset_subtest1()
            power = self.get_uplink_power('WCDMA')
            result = self.query_etfci()
            logger.debug(f'Now ETFCI result is: {result}')
            mstat = int(self.inst.query('mstat?').strip())
            logger.debug(f'mstat: {mstat}')
            if mstat == cm_pmt.MESUREMENT_TIMEOUT:
                logger.debug('time out and recall')
                self.inst.write('CALLSO')
                time.sleep(3)
                self.inst.write('CALLSA')
                self.preset_subtest1()
            result = cm_pmt.HSUPA_ETFCI_SUBTEST1
            logger.debug('force to the subtest1 ETFCI')
            while result == cm_pmt.HSUPA_ETFCI_SUBTEST1:
                logger.debug('TPC UP')
                self.inst.write('TPC_CMD_UP')
                time.sleep(0.15)
                self.set_to_measure()
                power = self.get_uplink_power('WCDMA')
                result = self.query_etfci()
                logger.debug(f'Now ETFCI result is: {result}')

            while result != cm_pmt.HSUPA_ETFCI_SUBTEST1:
                logger.debug('TPC DOWN')
                self.inst.write('TPC_CMD_DOWN')
                time.sleep(0.15)
                self.set_to_measure()
                power = self.get_uplink_power('WCDMA')
                result = self.query_etfci()
                logger.debug(f'Now ETFCI result is: {result}')

            logger.info(f'Subtest1 to capture the final power is {power}')

            aclr = self.get_uplink_aclr('WCDMA')

            self.set_tpc('ILPC')
            self.set_input_level(5)

            return power, aclr, 1

        elif self.chcoding == 'FIXREFCH':  # this is HSDPA
            logger.info('start to measure HSDPA')
            self.preset_subtest1()
            power = self.get_uplink_power('WCDMA')
            aclr = self.get_uplink_aclr('WCDMA')

            return power, aclr, 1

    def preset_subtest2(self):
        if self.chcoding == 'EDCHTEST':  # this is HSUPA
            self.inst.write('SET_HSDELTA_CQI 8')
            self.inst.write('SET_HSSUBTEST SUBTEST2')
            self.set_tpc('ILPC')
            self.set_input_level(14)
            time.sleep(0.15)
            self.set_tpc('ALT')
            self.set_input_level(26)
            logger.debug('TPC DOWN')
            self.inst.write('TPC_CMD_DOWN')
            time.sleep(0.15)
            self.set_to_measure()
        elif self.chcoding == 'FIXREFCH':  # this is HSDPA
            self.set_input_level(24)
            self.set_output_level(-86)
            self.set_tpc('ALL1')
            self.inst.write('DTCHPAT PN9')
            self.inst.write('SET_HSDELTA_CQI 8')
            self.inst.write('SET_HSSUBTEST SUBTEST2')
            self.set_to_measure()

    def get_subtest2_power_aclr(self):
        logger.info('Start to subtest2')
        if self.chcoding == 'EDCHTEST':  # this is HSUPA
            self.preset_subtest2()
            power = self.get_uplink_power('WCDMA')
            result = self.query_etfci()
            logger.debug(f'Now ETFCI result is: {result}')
            mstat = int(self.inst.query('mstat?').strip())
            logger.debug(f'mstat: {mstat}')
            if mstat == cm_pmt.MESUREMENT_TIMEOUT:
                self.preset_subtest2()
            result = cm_pmt.HSUPA_ETFCI_SUBTEST2
            logger.debug('force to the subtest2 ETFCI')
            while result == cm_pmt.HSUPA_ETFCI_SUBTEST2:
                logger.debug('TPC UP')
                self.inst.write('TPC_CMD_UP')
                time.sleep(0.15)
                self.set_to_measure()
                power = self.get_uplink_power('WCDMA')
                result = self.query_etfci()
                logger.debug(f'Now ETFCI result is: {result}')

            while result != cm_pmt.HSUPA_ETFCI_SUBTEST2:
                logger.debug('TPC DOWN')
                self.inst.write('TPC_CMD_DOWN')
                time.sleep(0.15)
                self.set_to_measure()
                power = self.get_uplink_power('WCDMA')
                result = self.query_etfci()
                logger.debug(f'Now ETFCI result is: {result}')

            logger.info(f'Subtest2 to capture the final power is {power}')
            aclr = self.get_uplink_aclr('WCDMA')

            self.set_tpc('ILPC')
            self.set_input_level(5)

            return power, aclr, 2

        elif self.chcoding == 'FIXREFCH':  # this is HSDPA
            logger.info('start to measure HSDPA')
            self.preset_subtest2()
            power = self.get_uplink_power('WCDMA')
            aclr = self.get_uplink_aclr('WCDMA')
            evm = self.get_uplink_evm('WCDMA')

            return power, aclr, 2

    def preset_subtest3(self):
        if self.chcoding == 'EDCHTEST':
            self.inst.write('SET_HSDELTA_CQI 8')
            self.inst.write('SET_HSSUBTEST SUBTEST3')
            self.set_tpc('ILPC')
            self.set_input_level(15)
            time.sleep(0.15)
            self.set_tpc('ALT')
            self.set_input_level(26)
            logger.debug('TPC DOWN')
            self.inst.write('TPC_CMD_DOWN')
            time.sleep(0.15)
            self.set_to_measure()
        elif self.chcoding == 'FIXREFCH':
            self.set_input_level(24)
            self.set_output_level(-86)
            self.set_tpc('ALL1')
            self.inst.write('DTCHPAT PN9')
            self.inst.write('SET_HSDELTA_CQI 8')
            self.inst.write('SET_HSSUBTEST SUBTEST3')
            self.set_to_measure()

    def get_subtest3_power_aclr(self):
        logger.info('Start to subtest3')
        if self.chcoding == 'EDCHTEST':  # this is HSUPA
            self.preset_subtest3()
            power = self.get_uplink_power('WCDMA')
            result = self.query_etfci()
            logger.debug(f'Now ETFCI result is: {result}')
            mstat = int(self.inst.query('mstat?').strip())
            logger.debug(f'mstat: {mstat}')
            if mstat == cm_pmt.MESUREMENT_TIMEOUT:
                self.preset_subtest3()
            result = cm_pmt.HSUPA_ETFCI_SUBTEST3
            logger.debug('force to the subtest3 ETFCI')
            while result == cm_pmt.HSUPA_ETFCI_SUBTEST3:
                logger.debug('TPC UP')
                self.inst.write('TPC_CMD_UP')
                time.sleep(0.15)
                self.set_to_measure()
                power = self.get_uplink_power('WCDMA')
                result = self.query_etfci()
                logger.debug(f'Now ETFCI result is: {result}')

            while result != cm_pmt.HSUPA_ETFCI_SUBTEST3:
                logger.debug('TPC DOWN')
                self.inst.write('TPC_CMD_DOWN')
                time.sleep(0.15)
                self.set_to_measure()
                power = self.get_uplink_power('WCDMA')
                result = self.query_etfci()
                logger.debug(f'Now ETFCI result is: {result}')

            logger.info(f'Subtest3 to capture the final power is {power}')
            aclr = self.get_uplink_aclr('WCDMA')

            self.set_tpc('ILPC')
            self.set_input_level(5)

            return power, aclr, 3
        elif self.chcoding == 'FIXREFCH':  # this is HSDPA
            logger.info('start to measure HSDPA')
            self.preset_subtest3()
            power = self.get_uplink_power('WCDMA')
            aclr = self.get_uplink_aclr('WCDMA')
            evm, phase_disc = self.get_hsdpa_evm()
            evm = Decimal(max(evm))
            phase_disc = Decimal(self.get_worse_phase_disc(phase_disc))

            return power, aclr, evm, 3

    def preset_subtest4(self):
        if self.chcoding == 'EDCHTEST':
            self.inst.write('SET_HSDELTA_CQI 8')
            self.inst.write('SET_HSSUBTEST SUBTEST4')
            self.set_tpc('ILPC')
            self.set_input_level(14)
            time.sleep(0.15)
            self.set_tpc('ALT')
            self.set_input_level(26)
            logger.debug('TPC DOWN')
            self.inst.write('TPC_CMD_DOWN')
            time.sleep(0.15)
            self.set_to_measure()
        elif self.chcoding == 'FIXREFCH':
            self.set_input_level(24)
            self.set_output_level(-86)
            self.set_tpc('ALL1')
            self.inst.write('DTCHPAT PN9')
            self.inst.write('SET_HSDELTA_CQI 8')
            self.inst.write('SET_HSSUBTEST SUBTEST4')
            self.set_to_measure()

    def get_subtest4_power_aclr(self):
        logger.info('Start to subtest4')
        if self.chcoding == 'EDCHTEST':  # this is HSUPA
            self.preset_subtest4()
            power = self.get_uplink_power('WCDMA')
            result = self.query_etfci()
            logger.debug(f'Now ETFCI result is: {result}')
            mstat = int(self.inst.query('mstat?').strip())
            logger.debug(f'mstat: {mstat}')
            if mstat == cm_pmt.MESUREMENT_TIMEOUT:
                self.preset_subtest4()
            result = cm_pmt.HSUPA_ETFCI_SUBTEST4
            logger.debug('force to the subtest4 ETFCI')
            while result == cm_pmt.HSUPA_ETFCI_SUBTEST4:
                logger.debug('TPC UP')
                self.inst.write('TPC_CMD_UP')
                time.sleep(0.15)
                self.set_to_measure()
                power = self.get_uplink_power('WCDMA')
                result = self.query_etfci()
                logger.debug(f'Now ETFCI result is: {result}')

            while result != cm_pmt.HSUPA_ETFCI_SUBTEST4:
                logger.debug('TPC DOWN')
                self.inst.write('TPC_CMD_DOWN')
                time.sleep(0.15)
                self.set_to_measure()
                power = self.get_uplink_power('WCDMA')
                result = self.query_etfci()
                logger.debug(f'Now ETFCI result is: {result}')

            logger.info(f'Subtest4 to capture the final power is {power}')
            aclr = self.get_uplink_aclr('WCDMA')

            self.set_tpc('ILPC')
            self.set_input_level(5)

            return power, aclr, 4
        elif self.chcoding == 'FIXREFCH':  # this is HSDPA
            logger.info('start to measure HSDPA')
            self.preset_subtest4()
            power = self.get_uplink_power('WCDMA')
            aclr = self.get_uplink_aclr('WCDMA')
            evm = self.get_uplink_evm('WCDMA')

            return power, aclr, 4

    def preset_subtest5(self):
        self.inst.write('TPUTU_MEAS OFF')
        self.inst.write('SUBTEST5_VER NEW')
        self.inst.write('SET_HSDELTA_CQI 8')
        self.inst.write('SET_HSSUBTEST SUBTEST5')
        self.set_tpc('ILPC')
        self.set_input_level(16)
        time.sleep(0.15)
        self.set_input_level(26)
        self.set_tpc('ALL1')
        time.sleep(1)
        self.set_to_measure()

    def get_subtest5_power_aclr(self):
        logger.info('Start to subtest5')
        if self.chcoding == 'EDCHTEST':  # this is HSUPA
            self.preset_subtest5()
            mstat = int(self.inst.query('mstat?').strip())
            logger.debug(f'mstat: {mstat}')
            if mstat == cm_pmt.MESUREMENT_TIMEOUT:
                self.preset_subtest5()
            logger.info('subtest5:')
            power = self.get_uplink_power('WCDMA')
            aclr = self.get_uplink_aclr('WCDMA')

            self.inst.write('TPUTU_MEAS OFF')
            self.inst.write('SET_HSSUBTEST SUBTEST1')
            self.set_tpc('ILPC')
            self.set_input_level(5)

            return power, aclr, 5
        elif self.chcoding == 'FIXREFCH':
            logger.info("HSDPA doesn't have subtest5")


    def get_subtest_power_aclr_evm_all(self):
        """
        data = [POPWER, ACLR ,EVM], and ACLR format: [low_-1, up_+1, low_-2, up_+2] for HSDPA
        data = [POPWER, ACLR], and ACLR format: [low_-1, up_+1, low_-2, up_+2] for HSUPA
        :return: data
        """
        data = {}
        subtests = [
            self.get_subtest1_power_aclr(),
            self.get_subtest2_power_aclr(),
            self.get_subtest3_power_aclr(),
            self.get_subtest4_power_aclr(),
            self.get_subtest5_power_aclr(),
        ]

        for subtest in subtests:
            if self.chcoding == 'EDCHTEST':  # this is HSUPA
                power, aclr, subtest_number = subtest
                data[subtest_number] = [power, aclr]

            elif self.chcoding == 'FIXREFCH':  # this is HSDPA
                if subtest is not None:
                    if len(subtest) == 4:
                        power, aclr, evm, subtest_number = subtest
                        data[subtest_number] = [power, aclr, evm]
                    else:
                        power, aclr, subtest_number = subtest
                        data[subtest_number] = [power, aclr]

        logger.debug(data)
        return data

    def sweep_sensitivity(self, start=-70, coarse=1, fine=0.2):
        if self.std == 'LTE':
            touch = 0  # flag if 0: reduce power by coarse, if 1: reduce power by fine
            count = 3
            while True:
                self.inst.write(f'OLVL {start}')
                logger.info(f"Search sensitivity: {self.inst.query('OLVL?').strip()}")
                self.set_to_measure()
                time.sleep(0.1)
                status = self.inst.query('TPUTPASS?').strip()
                conn_state = int(self.inst.query("CALLSTAT?").strip())

                if status == 'PASS' and touch == 0:  # by coarse
                    start -= coarse
                    touch = 0

                elif status == 'PASS' and touch == 1:  # by fine
                    start -= fine
                    status = self.inst.query('TPUTPASS?').strip()
                    while count > 1 and status == 'PASS':
                        status = self.inst.query('TPUTPASS?').strip()
                        count -= 1
                    count = 3

                elif status == 'FAIL' and touch == 0:
                    while count > 0 and status == 'FAIL':  # retest 3 time to judge if it is real sensitivity
                        self.set_to_measure()
                        time.sleep(0.1)
                        status = self.inst.query('TPUTPASS?').strip()
                        logger.info(f'{status}')
                        count -= 1
                    if count != 0:  # if it meets some sudden noise from environment
                        continue
                    else:  # real sensitivity failed
                        logger.info('Back to higher 2dB')
                        start += 2
                        count = 3
                        touch = 1

                elif status == 'FAIL' and touch == 1:  # it might be the real sensitivity
                    while True:
                        if status == 'FAIL':
                            start += fine
                            start = round(start, 1)
                            self.inst.write(f'OLVL {start}')
                            self.set_to_measure()
                            status = self.inst.query('TPUTPASS?').strip()
                            output_level = self.inst.query(f'OLVL?').strip()
                            logger.info(f'level {output_level}, {status}')


                        elif status == '*':
                            logger.info('Connection is dropped')
                            logger.info('Retest again from output level -70dBm')
                            start = -70
                            self.inst.write(f'OLVL {start}')
                            self.inst.write('CALLSO')
                            time.sleep(2)
                            self.inst.write('CALLSA')
                            time.sleep(2)
                            self.flymode_circle()
                            logger.info('waiting 10 seconds')
                            time.sleep(10)
                            conn_state = int(self.inst.query("CALLSTAT?").strip())
                            if conn_state == cm_pmt.ANRITSU_CONNECTED:
                                start -= fine
                                self.set_to_measure()
                                status = self.inst.query('TPUTPASS?').strip()
                                while status == 'PASS':
                                    self.inst.write(f'OLVL {start}')
                                    self.set_to_measure()
                                    status = self.inst.query('TPUTPASS?').strip()
                                    output_level = self.inst.query(f'OLVL?').strip()
                                    logger.info(f'reconnedted level {output_level}: {status}')
                                    start -= fine
                            else:
                                start = -70
                                logger.info('Skip this channel, and set the output level to -70dBm for notice')

                        elif status == 'PASS':
                            while count > 0:
                                self.set_to_measure()
                                output_level = self.inst.query('OLVL?').strip()
                                status = self.inst.query('TPUTPASS?').strip()
                                if status == 'FAIL':
                                    logger.info(f'{4 - count} times fail')
                                    break
                                logger.info(f"sensitivity: {output_level}, pass {4 - count} times")
                                count -= 1

                            count = 3

                            if status == 'FAIL':
                                continue
                            else:
                                break
                    sensitivity = Decimal(self.inst.query('OLVL?').strip())
                    time.sleep(0.1)
                    per = self.inst.query('TPUT? PER').strip()
                    power = Decimal(self.inst.query('POWER? AVG').strip())
                    self.inst.query('*OPC?')
                    logger.info(f'Final: POWER: {power}, SENSITIVITY: {sensitivity}, PER:{per}')
                    return [power, sensitivity, per]

                elif conn_state != cm_pmt.ANRITSU_CONNECTED:
                    count = 3
                    sub_count = 100
                    while count > 0 and conn_state != cm_pmt.ANRITSU_CONNECTED:
                        logger.info('Call drop and fly on and off')
                        self.inst.write('CALLSO')
                        time.sleep(2)
                        self.inst.write('CALLSA')
                        time.sleep(2)
                        self.flymode_circle()
                        while conn_state != cm_pmt.ANRITSU_CONNECTED and sub_count > 0:
                            time.sleep(1)
                            logger.info('Wait 1 second to connect')
                            conn_state = int(self.inst.query("CALLSTAT?").strip())
                            sub_count -= 1
                        logger.info('Reconnected')
                        sub_count = 100
                        count -= 1
                    if count == 0:
                        logger.info('Stop this circle')
                        break
        elif self.std == 'WCDMA':
            touch = 0  # flag if 0: reduce power by coarse, if 1: reduce power by fine
            count = 3
            mstate = int(self.inst.query('MSTAT?').strip())
            while True:
                self.set_output_level(start)
                logger.info(f"Search sensitivity: {self.inst.query('OLVL?').strip()}")
                time.sleep(0.1)
                conn_state = int(self.inst.query("CALLSTAT?").strip())
                self.set_to_measure()
                status = self.get_ber()
                mstate = int(self.inst.query('MSTAT?').strip())
                logger.debug(f'measuring statuse: {mstate}')

                if mstate == cm_pmt.MESUREMENT_GOOD and conn_state == cm_pmt.ANRITSU_LOOP_MODE_1 and status == 'PASS' and touch == 0:  # by coarse
                    start -= coarse
                    touch = 0



                elif mstate == cm_pmt.MESUREMENT_GOOD and conn_state == cm_pmt.ANRITSU_LOOP_MODE_1 and status == 'PASS' and touch == 1:  # by fine
                    start -= fine
                    status = self.get_ber()
                    while count > 1 and status == 'PASS':
                        status = self.get_ber()
                        count -= 1
                    count = 3


                elif mstate == cm_pmt.MESUREMENT_GOOD and conn_state == cm_pmt.ANRITSU_LOOP_MODE_1 and status == 'FAIL' and touch == 0:
                    while count > 0 and status == 'FAIL':  # retest 3 time to judge if it is real sensitivity
                        self.set_to_measure()
                        time.sleep(0.1)
                        status = self.get_ber()
                        logger.info(f'{status}')
                        count -= 1
                    if count != 0:  # if it meets some sudden noise from environment
                        continue
                    else:  # real sensitivity failed
                        logger.info('Back to higher 2dB')
                        start += 2
                        count = 3
                        touch = 1

                elif mstate == cm_pmt.MESUREMENT_GOOD and conn_state == cm_pmt.ANRITSU_LOOP_MODE_1 and status == 'FAIL' and touch == 1:  # it might be the real sensitivity
                    while mstate != cm_pmt.MESUREMENT_TIMEOUT:
                        if status == 'FAIL':
                            start += fine
                            start = round(start, 1)
                            self.inst.write(f'OLVL {start}')
                            self.set_to_measure()
                            status = self.get_ber()
                            output_level = self.inst.query(f'OLVL?').strip()
                            logger.info(f'level {output_level}, {status}')
                            mstate = int(self.inst.query('MSTAT?').strip())

                        elif conn_state != cm_pmt.ANRITSU_LOOP_MODE_1:
                            logger.debug(f'measuring statuse: {mstate}')
                            logger.info('Connection is dropped')
                            logger.info('Retest again from output level -70dBm')
                            start = -70
                            self.set_output_level(start)
                            self.set_input_level()
                            self.inst.write('CALLSO')
                            time.sleep(2)
                            self.inst.write('CALLSA')
                            time.sleep(2)
                            self.flymode_circle()
                            logger.info('waiting 10 seconds')
                            time.sleep(10)
                            conn_state = int(self.inst.query("CALLSTAT?").strip())
                            if conn_state == ANRITSU_LOOP_MODE_1:
                                self.set_input_level(30)
                                start -= fine
                                self.set_to_measure()
                                status = self.get_ber()
                                while status == 'PASS':
                                    self.inst.write(f'OLVL {start}')
                                    self.set_to_measure()
                                    status = self.get_ber()
                                    output_level = self.inst.query(f'OLVL?').strip()
                                    logger.info(f'reconnedted level {output_level}: {status}')
                                    start -= fine

                            else:
                                start = -70
                                logger.info('Skip this channel, and set the output level to -70dBm for notice')


                        elif status == 'PASS':
                            while count > 0:
                                self.set_to_measure()
                                output_level = self.inst.query('OLVL?').strip()
                                status = self.get_ber()
                                if status == 'FAIL':
                                    logger.info(f'{4 - count} times fail')
                                    break
                                logger.info(f"sensitivity: {output_level}, pass {4 - count} times")
                                count -= 1
                                mstate = int(self.inst.query('MSTAT?').strip())

                            count = 3

                            if status == 'FAIL':
                                continue
                            else:
                                break
                    sensitivity = Decimal(self.inst.query('OLVL?').strip())
                    time.sleep(0.1)
                    per = self.inst.query('BER? PER').strip()
                    power = Decimal(self.inst.query('AVG_POWER?').strip())
                    self.inst.query('*OPC?')
                    logger.info(f'Final: POWER: {power}, SENSITIVITY: {sensitivity}, PER:{per}')
                    return [power, sensitivity, per]

                elif mstate != cm_pmt.MESUREMENT_TIMEOUT and conn_state != cm_pmt.ANRITSU_LOOP_MODE_1:
                    logger.debug(f'measuring statuse: {mstate}')
                    logger.info('Call drop and fly on and off')
                    logger.info('Retest again and Reconnected')
                    touch = 0
                    start = -70
                    dl_ch = int(self.inst.query('DLCHAN?'))
                    tpc_status = self.inst.query('TPCPAT?').strip()
                    self.set_init_before_calling(self.std, dl_ch)
                    self.set_registration_calling(self.std)
                    self.set_init_rx(self.std)
                    self.set_tpc(tpc_status)
                    if tpc_status == 'ALL1':
                        self.set_input_level(30)
                    elif tpc_status == 'ILPC':
                        self.set_input_level(-10)

                    self.inst.write('BER_MEAS ON')
                    self.inst.write('PWR_MEAS ON')

                elif mstate == cm_pmt.MESUREMENT_TIMEOUT:
                    logger.info('Time out')
                    logger.info('Retest again and Reconnected')
                    touch = 0
                    start = -70
                    dl_ch = int(self.inst.query('DLCHAN?'))
                    tpc_status = self.inst.query('TPCPAT?').strip()
                    self.set_init_before_calling(self.std, dl_ch)
                    self.set_registration_calling(self.std)
                    self.set_init_rx(self.std)
                    self.set_tpc(tpc_status)
                    if tpc_status == 'ALL1':
                        self.set_input_level(30)
                    elif tpc_status == 'ILPC':
                        self.set_input_level(-10)

                    self.inst.write('BER_MEAS ON')
                    self.inst.write('PWR_MEAS ON')

    def get_ber(self):
        """
        write (BER? PER) and we can get the x % and the x will lower than 0.1
        :return: PASS or FAIL
        """
        ber = Decimal(self.inst.query('BER? PER').strip())
        if ber >= 0.1:
            return 'FAIL'
        else:
            return 'PASS'

    def get_sensitivity(self, standard, band, dl_ch, bw=None):
        self.std = s = standard
        if s == 'LTE':
            """
                sens_list = [power, sensitivity, per]
            """
            self.set_handover(s, dl_ch, bw)
            time.sleep(0.1)
            self.set_rb_location(band, bw)
            sens_list = self.sweep_sensitivity()
            return sens_list

        elif s == 'WCDMA':
            """
                sens_list = [power, sensitivity, per]
            """
            self.inst.write('BER_MEAS ON')
            self.inst.write('PWR_MEAS ON')
            self.set_handover(s, dl_ch)
            time.sleep(0.1)
            sens_list = self.sweep_sensitivity()
            return sens_list
        elif s == 'GSM':
            pass

    def get_validation(self, standard):
        s = standard  # WCDMA|GSM|LTE
        logger.debug("Current Format: " + s)
        if s == 'LTE':
            return self.get_power_aclr_evm_lte()
        elif s == 'WCDMA':
            if self.chcoding == 'REFMEASCH':  # this is WCDMA
                return self.get_power_aclr_evm_wcdma()
            elif self.chcoding == 'EDCHTEST':  # this is HSUPA
                return self.get_subtest_power_aclr_evm_all()
            elif self.chcoding == 'FIXREFCH':  # this is HSDPA
                return self.get_subtest_power_aclr_evm_all()
        elif s == 'GSM':
            pass

    def get_power_aclr_evm_lte(self):
        """
            Only measure RB@min
            The format in dictionary is {Q_1: [power, aclr, evm], Q_P: [power, aclr, evm], ...}
            and ACLR format is [EUTRA-1, EUTRA+1, UTRA-1, URTA+1, UTRA-2, URTA+2,]
        """
        want_mods = [
            'TESTPRM TX_MAXPWR_Q_1',
            'TESTPRM TX_MAXPWR_Q_P',
            'TESTPRM TX_MAXPWR_Q_F',
            'TESTPRM TX_MAXPWR_16_P',
            'TESTPRM TX_MAXPWR_16_F',
            'TESTPRM TX_MAXPWR_64_P',
            'TESTPRM TX_MAXPWR_64_F',
        ]

        validation_dict = {}

        self.set_init_power()
        self.set_init_aclr('LTE')
        self.set_init_mod('LTE')
        self.set_input_level(26)
        self.set_tpc('ALL3')
        self.inst.query('*OPC?')

        for mod in want_mods:
            self.mod = mod[18:]
            conn_state = int(self.inst.query("CALLSTAT?").strip())
            self.count = 5
            while conn_state != cm_pmt.ANRITSU_CONNECTED:  # this is for waiting connection before change modulation if there is connection problems
                logger.info('Call drops...')
                if self.count == 0:
                    # equipment end call and start call
                    logger.info('End call and then start call')
                    self.flymode_circle()
                    time.sleep(5)
                    self.inst.write('CALLSO')
                    self.inst.query('*OPC?')
                    time.sleep(1)
                    self.inst.write('CALLSA')
                    time.sleep(10)
                    self.count = 6
                    conn_state = int(self.inst.query("CALLSTAT?").strip())

                else:
                    time.sleep(10)
                    self.count -= 1
                    logger.info('wait 10 seconds to connect')
                    logger.info(f'{6 - self.count} times to wait 10 second')
                    conn_state = int(self.inst.query("CALLSTAT?").strip())

            validation_list = []
            if mod == 'TESTPRM TX_MAXPWR_64_P':
                self.inst.write('TESTPRM TX_MAXPWR_Q_P')
                self.inst.write('ULRMC_64QAM ENABLED')
                self.inst.write('ULIMCS 21')
            elif mod == 'TESTPRM TX_MAXPWR_64_F':
                self.inst.write('TESTPRM TX_MAXPWR_Q_F')
                self.inst.write('ULRMC_64QAM ENABLED')
                self.inst.write('ULIMCS 21')
            else:
                self.inst.write('ULRMC_64QAM DISABLED')
                self.inst.write(mod)

            self.set_to_measure()
            meas_status = int(self.inst.query('MSTAT?').strip())

            while meas_status == cm_pmt.MESUREMENT_BAD:  # this is for the reference signal is not found
                logger.info('measuring status is bad(Reference signal not found)')
                logger.info('Equipment is forced to set End Call')
                self.inst.write('CALLSO')
                time.sleep(5)
                logger.info('fly on and off again')
                self.flymode_circle()
                time.sleep(10)
                self.inst.write('CALLSA')
                logger.info('waiting for 10 second to re-connect')
                logger.info('measure it again')
                self.set_to_measure()
                meas_status = int(self.inst.query('MSTAT?').strip())

            if mod == 'TESTPRM TX_MAXPWR_Q_1':  # mod[18:] -> Q_1
                logger.info(mod)
                validation_list.append(self.get_uplink_power('LTE'))
                validation_dict[mod[18:]] = validation_list
                # self.inst.query('*OPC?')
            else:  # mod[18:] -> Q_P, Q_F, 16_P, 16_F, 64_F
                logger.info(mod)
                self.pwr = self.get_uplink_power('LTE')
                validation_list.append(self.pwr)
                self.aclr = self.get_uplink_aclr('LTE')
                validation_list.append(self.aclr)
                self.evm = self.get_uplink_evm('LTE')
                validation_list.append(self.evm)
                validation_dict[mod[18:]] = validation_list
                self.inst.query('*OPC?')
        logger.debug(validation_dict)
        return validation_dict

    def get_power_aclr_evm_wcdma(self):
        """
            Only measure RB@min
            The format in dictionary is [power, aclr, evm]
            and ACLR format is [UTRA-1, URTA+1, UTRA-2, URTA+2,]
        """

        self.set_init_power()
        self.set_init_aclr('WCDMA')
        self.set_init_mod('WCDMA')
        self.set_input_level(26)
        self.set_output_level(-93)
        self.set_tpc('ALL1')
        self.inst.query('*OPC?')

        validation_list = []
        self.set_to_measure()

        self.pwr = self.get_uplink_power('WCDMA')
        validation_list.append(self.pwr)
        self.aclr = self.get_uplink_aclr('WCDMA')
        validation_list.append(self.aclr)
        self.evm = self.get_uplink_evm('WCDMA')
        validation_list.append(self.evm)
        self.inst.query('*OPC?')
        logger.debug(validation_list)
        return validation_list

    def get_uplink_power(self, standard):
        """
            Get UL power
        """
        s = standard  # WCDMA|GSM|LTE
        logger.debug("Current Format: " + s)
        if s == 'LTE':
            power = Decimal(self.inst.query('POWER? AVG').strip())
            self.inst.query('*OPC?')
            logger.info(f'POWER: {power}')
            return power
        elif s == 'WCDMA':
            power = Decimal(self.inst.query('AVG_POWER?').strip())
            self.inst.query('*OPC?')
            logger.info(f'POWER: {power}')
            return power
        elif s == 'GSM':
            pass

    def get_uplink_aclr(self, standard):
        """
            LTE:
                Get LTE ACLR
                return in [EUTRA-1, EUTRA+1, UTRA-1, URTA+1, UTRA-2, URTA+2,]
            WCDMA:
                Get LTE ACLR
                return in [LOW5, UP5, LOW10, UP10,] format

        """
        s = standard  # WCDMA|GSM|LTE
        logger.debug("Current Format: " + s)
        aclr = []
        if s == 'LTE':
            aclr.append(Decimal(self.inst.query('MODPWR? E_LOW1,AVG').strip()))
            aclr.append(Decimal(self.inst.query('MODPWR? E_UP1,AVG').strip()))
            aclr.append(Decimal(self.inst.query('MODPWR? LOW1,AVG').strip()))
            aclr.append(Decimal(self.inst.query('MODPWR? UP1,AVG').strip()))
            aclr.append(Decimal(self.inst.query('MODPWR? LOW2,AVG').strip()))
            aclr.append(Decimal(self.inst.query('MODPWR? UP2,AVG').strip()))
            self.inst.query('*OPC?')
            logger.info(f'ACLR: {aclr}')
            return aclr
        elif s == 'WCDMA':
            aclr.append(Decimal(self.inst.query('AVG_MODPWR? LOW5').strip()))
            aclr.append(Decimal(self.inst.query('AVG_MODPWR? UP5').strip()))
            aclr.append(Decimal(self.inst.query('AVG_MODPWR? LOW10').strip()))
            aclr.append(Decimal(self.inst.query('AVG_MODPWR? UP10').strip()))
            self.inst.query('*OPC?')
            logger.info(f'ACLR: {aclr}')
            return aclr
        elif s == 'GSM':
            pass

    def get_uplink_evm(self, standard):
        """
            Get Error Vector Magnitude (EVM) - PUSCH @ max power
        """
        s = standard  # WCDMA|GSM|LTE
        logger.debug("Current Format: " + s)
        if s == 'LTE':
            evm = Decimal(self.inst.query('EVM? AVG').strip())
            self.inst.query('*OPC?')
            logger.info(f'EVM: {evm}')
            return evm
        elif s == 'WCDMA':
            evm = Decimal(self.inst.query('AVG_EVM?').strip())
            self.inst.query('*OPC?')
            logger.info(f'EVM: {evm}')
            return evm
        elif s == 'GSM':
            pass

    def aclr_ch_judge(self, standard, band, dl_ch, bw=None):
        if standard == 'LTE':
            if dl_ch < cm_pmt.dl_ch_selected(self.std, band, bw)[1]:
                self.aclr_ch = 'ch01'
            elif dl_ch == cm_pmt.dl_ch_selected(self.std, band, bw)[1]:
                self.aclr_ch = 'ch02'
            elif dl_ch > cm_pmt.dl_ch_selected(self.std, band, bw)[1]:
                self.aclr_ch = 'ch03'
        elif standard == 'WCDMA':
            if dl_ch < cm_pmt.dl_ch_selected(self.std, band)[1]:
                self.aclr_ch = 'ch01'
            elif dl_ch == cm_pmt.dl_ch_selected(self.std, band)[1]:
                self.aclr_ch = 'ch02'
            elif dl_ch > cm_pmt.dl_ch_selected(self.std, band)[1]:
                self.aclr_ch = 'ch03'
        elif standard == 'GSM':
            pass

    def create_sheet_title(self, sheet):
        sh = sheet
        sh['A1'] = 'Band'
        sh['B1'] = 'DL_chan'
        sh['C1'] = 'Sensitivity'
        sh['D1'] = 'Tx_power'

    def create_excel_rx_sweep(self, standard, power_selected, bw=None):
        wb = openpyxl.Workbook()
        wb.remove(wb['Sheet'])

        sheet = None
        if power_selected == 1:
            sheet = f'Band{self.band}_Sweep_TxMax'
            wb.create_sheet(sheet)
        else:
            sheet = f'Band{self.band}_Sweep_TxMin'
            wb.create_sheet(sheet)

        # create titile for first row
        sh = wb[sheet]
        sh['A1'] = 'Band'
        sh['B1'] = 'DL_chan'
        sh['C1'] = 'Sensitivity'
        sh['D1'] = 'Tx_power'

        if standard == 'LTE':
            wb.save(f'Sweep_{bw}MHZ_LTE.xlsx')
        elif standard == 'WCDMA':
            wb.save(f'Sweep_WCDMA.xlsx')

        wb.close()

    @staticmethod
    def create_excel_rx(standard, bw=None):
        if standard == 'LTE':
            wb = openpyxl.Workbook()
            wb.remove(wb['Sheet'])
            wb.create_sheet('Sensitivity_TxMax')
            wb.create_sheet('Sensitivity_TxMin')
            wb.create_sheet('Desens')
            wb.create_sheet('PWR_TxMax')
            wb.create_sheet('PWR_TxMin')

            for sheet in wb.sheetnames:
                sh = wb[sheet]
                sh['A1'] = 'Band'
                sh['B1'] = 'ch0'
                sh['C1'] = 'ch1'
                sh['D1'] = 'ch2'

            wb.save(f'Sensitivity_{bw}MHZ_LTE.xlsx')
            wb.close()

        elif standard == 'WCDMA':
            wb = openpyxl.Workbook()
            wb.remove(wb['Sheet'])
            wb.create_sheet('Sensitivity_TxMax')
            wb.create_sheet('Sensitivity_TxMin')
            wb.create_sheet('Desens')
            wb.create_sheet('PWR_TxMax')
            wb.create_sheet('PWR_TxMin')

            for sheet in wb.sheetnames:
                sh = wb[sheet]
                sh['A1'] = 'Band'
                sh['B1'] = 'ch0'
                sh['C1'] = 'ch1'
                sh['D1'] = 'ch2'

            wb.save(f'Sensitivity_WCDMA.xlsx')
            wb.close()

    def create_excel_tx(self, standard, bw=None):
        if standard == 'LTE':
            wb = openpyxl.Workbook()
            wb.remove(wb['Sheet'])
            wb.create_sheet('PWR_Q_1')
            wb.create_sheet('PWR_Q_P')
            wb.create_sheet('PWR_Q_F')
            wb.create_sheet('PWR_16_P')
            wb.create_sheet('PWR_16_F')
            wb.create_sheet('PWR_64_P')
            wb.create_sheet('PWR_64_F')
            wb.create_sheet('ACLR_Q_P')
            wb.create_sheet('ACLR_Q_F')
            wb.create_sheet('ACLR_16_P')
            wb.create_sheet('ACLR_16_F')
            wb.create_sheet('ACLR_64_P')
            wb.create_sheet('ACLR_64_F')
            wb.create_sheet('EVM_Q_P')
            wb.create_sheet('EVM_Q_F')
            wb.create_sheet('EVM_16_P')
            wb.create_sheet('EVM_16_F')
            wb.create_sheet('EVM_64_P')
            wb.create_sheet('EVM_64_F')

            for sheet in wb.sheetnames:
                if 'ACLR' in sheet:
                    sh = wb[sheet]
                    sh['A1'] = 'Band'
                    sh['B1'] = 'Channel'
                    sh['C1'] = 'EUTRA_-1'
                    sh['D1'] = 'EUTRA_+1'
                    sh['E1'] = 'UTRA_-1'
                    sh['F1'] = 'UTRA_+1'
                    sh['G1'] = 'UTRA_-2'
                    sh['H1'] = 'UTRA_+2'

                else:
                    sh = wb[sheet]
                    sh['A1'] = 'Band'
                    sh['B1'] = 'ch0'
                    sh['C1'] = 'ch1'
                    sh['D1'] = 'ch2'

            wb.save(f'results_{bw}MHZ_LTE.xlsx')
            wb.close()

        elif standard == 'WCDMA' and self.chcoding == 'REFMEASCH':  # this is WCDMA
            wb = openpyxl.Workbook()
            wb.remove(wb['Sheet'])
            wb.create_sheet('PWR')
            wb.create_sheet('ACLR')
            wb.create_sheet('EVM')

            for sheet in wb.sheetnames:
                if 'ACLR' in sheet:
                    sh = wb[sheet]
                    sh['A1'] = 'Band'
                    sh['B1'] = 'Channel'
                    sh['C1'] = 'UTRA_-1'
                    sh['D1'] = 'UTRA_+1'
                    sh['E1'] = 'UTRA_-2'
                    sh['F1'] = 'UTRA_+2'
                else:
                    sh = wb[sheet]
                    sh['A1'] = 'Band'
                    sh['B1'] = 'ch01'
                    sh['C1'] = 'ch02'
                    sh['D1'] = 'ch03'

            wb.save(f'results_WCDMA.xlsx')
            wb.close()

        elif standard == 'WCDMA' and self.chcoding == 'EDCHTEST':  # this is HSUPA
            wb = openpyxl.Workbook()
            wb.remove(wb['Sheet'])
            wb.create_sheet('PWR')
            wb.create_sheet('ACLR')

            for sheet in wb.sheetnames:
                if 'ACLR' in sheet:
                    sh = wb[sheet]
                    sh['A1'] = 'Band'
                    sh['B1'] = 'Channel'
                    sh['C1'] = 'UTRA_-1'
                    sh['D1'] = 'UTRA_+1'
                    sh['E1'] = 'UTRA_-2'
                    sh['F1'] = 'UTRA_+2'
                    sh['G1'] = 'subtest_number'
                else:
                    sh = wb[sheet]
                    sh['A1'] = 'Band'
                    sh['B1'] = 'ch01'
                    sh['C1'] = 'ch02'
                    sh['D1'] = 'ch03'
                    sh['E1'] = 'subtest_number'

            wb.save(f'results_HSUPA.xlsx')
            wb.close()

        elif standard == 'WCDMA' and self.chcoding == 'FIXREFCH':  # this is HSDPA
            wb = openpyxl.Workbook()
            wb.remove(wb['Sheet'])
            wb.create_sheet('PWR')
            wb.create_sheet('ACLR')
            wb.create_sheet('EVM')

            for sheet in wb.sheetnames:
                if 'ACLR' in sheet:
                    sh = wb[sheet]
                    sh['A1'] = 'Band'
                    sh['B1'] = 'Channel'
                    sh['C1'] = 'UTRA_-1'
                    sh['D1'] = 'UTRA_+1'
                    sh['E1'] = 'UTRA_-2'
                    sh['F1'] = 'UTRA_+2'
                    sh['G1'] = 'subtest_number'
                else:
                    sh = wb[sheet]
                    sh['A1'] = 'Band'
                    sh['B1'] = 'ch01'
                    sh['C1'] = 'ch02'
                    sh['D1'] = 'ch03'
                    sh['E1'] = 'subtest_number'

            wb.save(f'results_HSDPA.xlsx')
            wb.close()

    @staticmethod
    def fill_desens(excel_path):

        wb = openpyxl.load_workbook(excel_path)

        ws = wb['Desens']
        ws_s_txmax = wb['Sensitivity_TxMax']
        ws_s_txmin = wb['Sensitivity_TxMin']
        max_row = max(ws_s_txmax.max_row, ws_s_txmin.max_row)
        for row in range(2, max_row + 1):
            for col in range(1, ws.max_column + 1):
                if col == 1:
                    ws.cell(row, col).value = ws_s_txmax.cell(row, col).value
                else:
                    try:
                        logger.debug(ws_s_txmax.cell(row, col).value)
                        logger.debug(ws_s_txmin.cell(row, col).value)
                        ws.cell(row, col).value = Decimal(ws_s_txmax.cell(row, col).value) - Decimal(
                            ws_s_txmin.cell(row, col).value)
                    except TypeError:
                        if ws_s_txmax.cell(row, col).value is None and ws_s_txmin.cell(row, col).value is not None:
                            logger.debug('Sensitivity_TxMax is None')
                            ws.cell(row, col).value = - Decimal(ws_s_txmin.cell(row, col).value)
                        elif ws_s_txmin.cell(row, col).value is None and ws_s_txmax.cell(row, col).value is not None:
                            logger.debug('Sensitivity_TxMin is None')
                            ws.cell(row, col).value = Decimal(ws_s_txmax.cell(row, col).value)
                        else:
                            logger.debug('Sensitivity_TxMax and Sensitivity_TxMin are None')

        wb.save(excel_path)
        wb.close()

    def fill_sensitivity_sweep(self, row, ws, band, dl_ch, data):
        # data[x]: 0 = power, 1 = sensitivity, 2 = PER
        ws.cell(row, 1).value = band
        ws.cell(row, 2).value = dl_ch
        logger.debug('sensitivity')
        ws.cell(row, 3).value = data[1]
        logger.debug('power')
        ws.cell(row, 4).value = data[0]

    def fill_sensitivity(self, standard, row, ws, band, dl_ch, data, items_selected, bw=None):
        # items_selected: 0 = power, 1 = sensitivity, 2 = PER
        if standard == 'LTE':
            ws.cell(row, 1).value = band

            if dl_ch < cm_pmt.dl_ch_selected(self.std, band, bw)[1]:
                ws.cell(row, 2).value = data[items_selected]
                if items_selected == 0:
                    logger.debug('power of L ch')
                elif items_selected == 1:
                    logger.debug('sensitivity of L ch')
            elif dl_ch == cm_pmt.dl_ch_selected(self.std, band, bw)[1]:
                ws.cell(row, 3).value = data[items_selected]
                if items_selected == 0:
                    logger.debug('power of M ch')
                elif items_selected == 1:
                    logger.debug('sensitivity of M ch')
            elif dl_ch > cm_pmt.dl_ch_selected(self.std, band, bw)[1]:
                ws.cell(row, 4).value = data[items_selected]
                if items_selected == 0:
                    logger.debug('power of H ch')
                elif items_selected == 1:
                    logger.debug('sensitivity of H ch')

        elif standard == 'WCDMA':
            ws.cell(row, 1).value = band

            if dl_ch < cm_pmt.dl_ch_selected(self.std, band)[1]:
                ws.cell(row, 2).value = data[items_selected]
                if items_selected == 0:
                    logger.debug('power of L ch')
                elif items_selected == 1:
                    logger.debug('sensitivity of L ch')
            elif dl_ch == cm_pmt.dl_ch_selected(self.std, band)[1]:
                ws.cell(row, 3).value = data[items_selected]
                if items_selected == 0:
                    logger.debug('power of M ch')
                elif items_selected == 1:
                    logger.debug('sensitivity of M ch')
            elif dl_ch > cm_pmt.dl_ch_selected(self.std, band)[1]:
                ws.cell(row, 4).value = data[items_selected]
                if items_selected == 0:
                    logger.debug('power of H ch')
                elif items_selected == 1:
                    logger.debug('sensitivity of H ch')

    def fill_power_aclr_evm_hspa(self, standard, row, ws, band, dl_ch, test_items, items_selected, subtest):
        if standard == 'WCDMA':
            if self.chcoding == 'EDCHTEST' or self.chcoding == 'FIXREFCH':  # this is for HSUPA pr HSDPA
                ws.cell(row, 1).value = band
                if items_selected == 0 or items_selected == 2:  # when select power or evm
                    ws.cell(row, 5).value = subtest
                    if dl_ch < cm_pmt.dl_ch_selected(self.std, band)[1]:
                        ws.cell(row, 2).value = test_items[items_selected]
                        if items_selected == 0:
                            logger.debug('the power of L ch')
                        elif items_selected == 2:
                            logger.debug('the evm of L ch')
                    elif dl_ch == cm_pmt.dl_ch_selected(self.std, band)[1]:
                        ws.cell(row, 3).value = test_items[items_selected]
                        if items_selected == 0:
                            logger.debug('the power of M ch')
                        elif items_selected == 2:
                            logger.debug('the evm of M ch')
                    elif dl_ch > cm_pmt.dl_ch_selected(self.std, band)[1]:
                        ws.cell(row, 4).value = test_items[items_selected]
                        if items_selected == 0:
                            logger.debug('the power of H ch')
                        elif items_selected == 2:
                            logger.debug('the evm of H ch')

                elif items_selected == 1:  # when select aclr
                    ws.cell(row, 7).value = subtest
                    if dl_ch < cm_pmt.dl_ch_selected(self.std, band)[1]:
                        ws.cell(row, 2).value = 'ch01'
                        for col, aclr_item in enumerate(test_items[items_selected]):
                            ws.cell(row, 3 + col).value = aclr_item
                        logger.debug('the ALCR of L ch')
                    elif dl_ch == cm_pmt.dl_ch_selected(self.std, band)[1]:
                        ws.cell(row, 2).value = 'ch02'
                        for col, aclr_item in enumerate(test_items[items_selected]):
                            ws.cell(row, 3 + col).value = aclr_item
                        logger.debug('the ACLR of M ch')
                    elif dl_ch > cm_pmt.dl_ch_selected(self.std, band)[1]:
                        ws.cell(row, 2).value = 'ch03'
                        for col, aclr_item in enumerate(test_items[items_selected]):
                            ws.cell(row, 3 + col).value = aclr_item
                        logger.debug('the ACLR of H ch')


            # elif self.chcoding == 'FIXREFCH':  # this is HSDPA
            #     pass

        else:
            logger.info('It might be erro to go here!!')

    def fill_power_aclr_evm(self, standard, row, ws, band, dl_ch, test_items, items_selected,
                            bw=None):  # items_selected: 0 = POWER, 1 = ACLR, 2 = EVM
        if standard == 'LTE':
            ws.cell(row, 1).value = band
            if items_selected == 0 or items_selected == 2:  # when select power or evm
                if dl_ch < cm_pmt.dl_ch_selected(self.std, band, bw)[1]:
                    ws.cell(row, 2).value = test_items[items_selected]
                    if items_selected == 0:
                        logger.debug('the power of L ch')
                    elif items_selected == 2:
                        logger.debug('the evm of L ch')
                elif dl_ch == cm_pmt.dl_ch_selected(self.std, band, bw)[1]:
                    ws.cell(row, 3).value = test_items[items_selected]
                    if items_selected == 0:
                        logger.debug('the power of M ch')
                    elif items_selected == 2:
                        logger.debug('the evm of M ch')
                elif dl_ch > cm_pmt.dl_ch_selected(self.std, band, bw)[1]:
                    ws.cell(row, 4).value = test_items[items_selected]
                    if items_selected == 0:
                        logger.debug('the power of H ch')
                    elif items_selected == 2:
                        logger.debug('the evm of H ch')

            elif items_selected == 1:  # when select aclr
                if dl_ch < cm_pmt.dl_ch_selected(self.std, band, bw)[1]:
                    ws.cell(row, 2).value = 'ch01'
                    for col, aclr_item in enumerate(test_items[items_selected]):
                        ws.cell(row, 3 + col).value = aclr_item
                    logger.debug('the ALCR of L ch')
                elif dl_ch == cm_pmt.dl_ch_selected(self.std, band, bw)[1]:
                    ws.cell(row, 2).value = 'ch02'
                    for col, aclr_item in enumerate(test_items[items_selected]):
                        ws.cell(row, 3 + col).value = aclr_item
                    logger.debug('the ACLR of M ch')
                elif dl_ch > cm_pmt.dl_ch_selected(self.std, band, bw)[1]:
                    ws.cell(row, 2).value = 'ch03'
                    for col, aclr_item in enumerate(test_items[items_selected]):
                        ws.cell(row, 3 + col).value = aclr_item
                    logger.debug('the ACLR of H ch')
        elif standard == 'WCDMA':
            if self.chcoding == 'REFMEASCH':  # this is WCDMA
                ws.cell(row, 1).value = band
                if items_selected == 0 or items_selected == 2:  # when select power or evm
                    if dl_ch < cm_pmt.dl_ch_selected(self.std, band)[1]:
                        ws.cell(row, 2).value = test_items[items_selected]
                        if items_selected == 0:
                            logger.debug('the power of L ch')
                        elif items_selected == 2:
                            logger.debug('the evm of L ch')
                    elif dl_ch == cm_pmt.dl_ch_selected(self.std, band)[1]:
                        ws.cell(row, 3).value = test_items[items_selected]
                        if items_selected == 0:
                            logger.debug('the power of M ch')
                        elif items_selected == 2:
                            logger.debug('the evm of M ch')
                    elif dl_ch > cm_pmt.dl_ch_selected(self.std, band)[1]:
                        ws.cell(row, 4).value = test_items[items_selected]
                        if items_selected == 0:
                            logger.debug('the power of H ch')
                        elif items_selected == 2:
                            logger.debug('the evm of H ch')

                elif items_selected == 1:  # when select aclr
                    if dl_ch < cm_pmt.dl_ch_selected(self.std, band)[1]:
                        ws.cell(row, 2).value = 'ch01'
                        for col, aclr_item in enumerate(test_items[items_selected]):
                            ws.cell(row, 3 + col).value = aclr_item
                        logger.debug('the ALCR of L ch')
                    elif dl_ch == cm_pmt.dl_ch_selected(self.std, band)[1]:
                        ws.cell(row, 2).value = 'ch02'
                        for col, aclr_item in enumerate(test_items[items_selected]):
                            ws.cell(row, 3 + col).value = aclr_item
                        logger.debug('the ACLR of M ch')
                    elif dl_ch > cm_pmt.dl_ch_selected(self.std, band)[1]:
                        ws.cell(row, 2).value = 'ch03'
                        for col, aclr_item in enumerate(test_items[items_selected]):
                            ws.cell(row, 3 + col).value = aclr_item
                        logger.debug('the ACLR of H ch')

    def fill_progress_rx_sweep(self, standard, ws, band, dl_ch, data,
                               power_selected):  # items_selected: 0 = power, 1 = sensitivity, 2 = PER
        if standard == 'LTE':
            if power_selected == 1:
                logger.debug(f'capture band: {band}, {self.bw}MHZ, {dl_ch}, TxMax, sensitivity')
            elif power_selected == 0:
                logger.debug(f'capture band: {band}, {self.bw}MHZ, {dl_ch}, TxMin, sensitivity')

            if ws.max_row == 1:  # only title
                self.fill_sensitivity_sweep(2, ws, band, dl_ch, data)
                logger.debug('Only title')
            else:
                for row in range(2, ws.max_row + 1):  # not only title
                    if ws.cell(row, 1).value == band and ws.cell(row, 2).value == dl_ch:  # if band is in the row
                        self.fill_sensitivity_sweep(row, ws, band, dl_ch, data)
                        logger.debug('Band and dl_ch are found')
                        break

                    elif row == ws.max_row:  # if band and dl_ch are found and final row
                        self.fill_sensitivity_sweep(row + 1, ws, band, dl_ch, data)
                        logger.debug('Band and dl_ch are not found')
                        break
                    else:
                        logger.debug('continue to search')
                        continue

        elif standard == 'WCDMA':
            if power_selected == 1:
                logger.debug(f'capture band: {band}, {dl_ch}, TxMax, sensitivity')
            elif power_selected == 0:
                logger.debug(f'capture band: {band}, {dl_ch}, TxMin, sensitivity')

            if ws.max_row == 1:  # only title
                self.fill_sensitivity_sweep(2, ws, band, dl_ch, data)
                logger.debug('Only title')
            else:
                for row in range(2, ws.max_row + 1):  # not only title
                    if ws.cell(row, 1).value == band and ws.cell(row, 2).value == dl_ch:  # if band is in the row
                        self.fill_sensitivity_sweep(row, ws, band, dl_ch, data)
                        logger.debug('Band and dl_ch are found')
                        break

                    elif row == ws.max_row:  # if band and dl_ch are found and final row
                        self.fill_sensitivity_sweep(row + 1, ws, band, dl_ch, data)
                        logger.debug('Band and dl_ch are not found')
                        break
                    else:
                        logger.debug('continue to search')
                        continue
        elif standard == 'GSM':
            pass

    def fill_progress_rx(self, standard, ws, band, dl_ch, data, items_selected, power_selected,
                         bw=None):  # items_selected: 0 = power, 1 = sensitivity, 2 = PER
        if standard == 'LTE':
            if power_selected == 1:
                logger.debug(f'capture band: {band}, {bw}MHZ, {dl_ch}, TxMax, sensitivity')
            elif power_selected == 0:
                logger.debug(f'capture band: {band}, {bw}MHZ, {dl_ch}, TxMin, sensitivity')

            if ws.max_row == 1:  # only title
                self.fill_sensitivity(standard, 2, ws, band, dl_ch, data, items_selected, bw)
                logger.debug('Only title')
            else:
                for row in range(2, ws.max_row + 1):  # not only title
                    if ws.cell(row, 1).value == band:  # if band is in the row
                        # POWER and EVM
                        self.fill_sensitivity(standard, row, ws, band, dl_ch, data, items_selected, bw)
                        logger.debug('Band is found')
                        break

                    elif ws.cell(row, 1).value != band and row == ws.max_row:  # if band is not in the row and final row
                        self.fill_sensitivity(standard, row + 1, ws, band, dl_ch, data, items_selected, bw)
                        logger.debug('Band is not found and the row is final and then to add new line')
                        break
                    else:
                        logger.debug('continue to search')
                        continue

        elif standard == 'WCDMA':
            if power_selected == 1:
                logger.debug(f'capture band: {band}, {dl_ch}, TxMax, sensitivity')
            elif power_selected == 0:
                logger.debug(f'capture band: {band}, {dl_ch}, TxMin, sensitivity')

            if ws.max_row == 1:  # only title
                self.fill_sensitivity(standard, 2, ws, band, dl_ch, data, items_selected)
                logger.debug('Only title')
            else:
                for row in range(2, ws.max_row + 1):  # not only title
                    if ws.cell(row, 1).value == band:  # if band is in the row
                        # POWER and EVM
                        self.fill_sensitivity(standard, row, ws, band, dl_ch, data, items_selected)
                        logger.debug('Band is found')
                        break

                    elif ws.cell(row, 1).value != band and row == ws.max_row:  # if band is not in the row and final row
                        self.fill_sensitivity(standard, row + 1, ws, band, dl_ch, data, items_selected)
                        logger.debug('Band is not found and the row is final and then to add new line')
                        break
                    else:
                        logger.debug('continue to search')
                        continue
        elif standard == 'GSM':
            pass

    def fill_progress_hspa_tx(self, standard, ws, band, dl_ch, test_items, test_items_selected, subtest):
        self.aclr_ch_judge(self.std, band, dl_ch)  # this is for ACLR fill in ACLR_TAB

        logger.debug(f'capture band: {band}, {self.aclr_ch}')

        if ws.max_row == 1:  # only title
            self.fill_power_aclr_evm_hspa(standard, 2, ws, band, dl_ch, test_items, test_items_selected, subtest)
            logger.debug('Only title')

        else:
            for row in range(2, ws.max_row + 1):  # not only title
                if ws.cell(row, 1).value == band and (
                        test_items_selected == 0 or test_items_selected == 2):  # if band is in the row
                    # POWER and EVM
                    if ws.cell(row, 5).value == subtest:
                        self.fill_power_aclr_evm_hspa(standard, row, ws, band, dl_ch, test_items, test_items_selected,
                                                      subtest)
                        logger.debug('Band is found')
                        break
                    elif ws.cell(row, 5).value != subtest and row == ws.max_row:
                        self.fill_power_aclr_evm_hspa(standard, row + 1, ws, band, dl_ch, test_items,
                                                      test_items_selected, subtest)
                        logger.debug('Band is the same, but subtest in not the same')
                        break

                elif ws.cell(row, 1).value != band and row == ws.max_row:  # if band is not in the row and final row
                    self.fill_power_aclr_evm_hspa(standard, row + 1, ws, band, dl_ch, test_items, test_items_selected,
                                                  subtest)
                    logger.debug('Band is not found and the row is final and then to add new line')
                    break


                elif ws.cell(row, 1).value == band and test_items_selected == 1 and ws.cell(row,
                                                                                            2).value == self.aclr_ch:

                    if ws.cell(row, 7).value == subtest:
                        self.fill_power_aclr_evm_hspa(standard, row, ws, band, dl_ch, test_items, test_items_selected,
                                                      subtest)
                        logger.debug('ch is the same for ACLR and subtest is the same')
                        break
                    elif ws.cell(row, 7).value != subtest and row == ws.max_row:
                        self.fill_power_aclr_evm_hspa(standard, row + 1, ws, band, dl_ch, test_items,
                                                      test_items_selected, subtest)
                        logger.debug('ch is the same for ACLR and subtest is not the same')
                        break

                elif ws.cell(row, 1).value == band and row == ws.max_row and test_items_selected == 1 and ws.cell(row,
                                                                                                                  2).value != self.aclr_ch:
                    if ws.cell(row, 7).value == subtest:
                        self.fill_power_aclr_evm_hspa(standard, row + 1, ws, band, dl_ch, test_items,
                                                      test_items_selected, subtest)
                        logger.debug('ch is not the same for ACLR and subtest is the same')
                        break
                    elif ws.cell(row, 7).value != subtest:
                        self.fill_power_aclr_evm_hspa(standard, row + 1, ws, band, dl_ch, test_items,
                                                      test_items_selected, subtest)
                        logger.debug('ch is not the same for ACLR and subtest is not the same')
                        break
                else:
                    logger.debug('continue to search')
                    continue

    def fill_progress_tx(self, standard, ws, band, dl_ch, test_items, test_items_selected,
                         bw=None):  # items_selected: 0 = POWER, 1 = ACLR, 2 = EVM
        self.aclr_ch_judge(self.std, band, dl_ch, bw)  # this is for ACLR fill in ACLR_TAB

        if standard == 'LTE':
            logger.debug(f'capture band: {band}, {bw}MHZ, {self.aclr_ch}')
            if ws.max_row == 1:  # only title
                self.fill_power_aclr_evm(standard, 2, ws, band, dl_ch, test_items, test_items_selected, bw)
                logger.debug('Only title')
            else:
                for row in range(2, ws.max_row + 1):  # not only title
                    if ws.cell(row, 1).value == band and (
                            test_items_selected == 0 or test_items_selected == 2):  # if band is in the row
                        # POWER and EVM
                        self.fill_power_aclr_evm(standard, row, ws, band, dl_ch, test_items, test_items_selected, bw)
                        logger.debug('Band is found')
                        break
                    elif ws.cell(row, 1).value == band and row == ws.max_row and test_items_selected == 1 and ws.cell(
                            row, 2).value != self.aclr_ch:
                        self.fill_power_aclr_evm(standard, row + 1, ws, band, dl_ch, test_items, test_items_selected,
                                                 bw)
                        logger.debug('ch is not the same for ACLR')
                        break
                    elif ws.cell(row, 1).value == band and test_items_selected == 1 and ws.cell(row,
                                                                                                2).value == self.aclr_ch:
                        self.fill_power_aclr_evm(standard, row, ws, band, dl_ch, test_items, test_items_selected, bw)
                        logger.debug('ch is the same for ACLR')
                        break
                    elif ws.cell(row, 1).value != band and row == ws.max_row:  # if band is not in the row and final row
                        self.fill_power_aclr_evm(standard, row + 1, ws, band, dl_ch, test_items, test_items_selected,
                                                 bw)
                        logger.debug('Band is not found and the row is final and then to add new line')
                        break
                    else:
                        logger.debug('continue to search')
                        continue
        elif standard == 'WCDMA':
            logger.debug(f'capture band: {band}, {self.aclr_ch}')

            if ws.max_row == 1:  # only title
                self.fill_power_aclr_evm(standard, 2, ws, band, dl_ch, test_items, test_items_selected)
                logger.debug('Only title')
            else:
                for row in range(2, ws.max_row + 1):  # not only title
                    if ws.cell(row, 1).value == band and (
                            test_items_selected == 0 or test_items_selected == 2):  # if band is in the row
                        # POWER and EVM
                        self.fill_power_aclr_evm(standard, row, ws, band, dl_ch, test_items, test_items_selected)
                        logger.debug('Band is found')
                        break
                    elif ws.cell(row, 1).value == band and row == ws.max_row and test_items_selected == 1 and ws.cell(
                            row, 2).value != self.aclr_ch:
                        self.fill_power_aclr_evm(standard, row + 1, ws, band, dl_ch, test_items, test_items_selected)
                        logger.debug('ch is not the same for ACLR')
                        break
                    elif ws.cell(row, 1).value == band and test_items_selected == 1 and ws.cell(row,
                                                                                                2).value == self.aclr_ch:
                        self.fill_power_aclr_evm(standard, row, ws, band, dl_ch, test_items, test_items_selected)
                        logger.debug('ch is the same for ACLR')
                        break
                    elif ws.cell(row, 1).value != band and row == ws.max_row:  # if band is not in the row and final row
                        self.fill_power_aclr_evm(standard, row + 1, ws, band, dl_ch, test_items, test_items_selected)
                        logger.debug('Band is not found and the row is final and then to add new line')
                        break
                    else:
                        logger.debug('continue to search')
                        continue

    def fill_values_rx_sweep(self, data, band, dl_ch, power_selected, bw=None):
        """
        data format:[Tx Power, Sensitivity, PER]
        """
        self.band = band
        self.bw = bw
        if self.std == 'LTE':
            if pathlib.Path(f'Sweep_{bw}MHZ_LTE.xlsx').exists() is False:
                self.create_excel_rx_sweep(self.std, power_selected, bw)
                logger.debug('Create Excel')

            wb = openpyxl.load_workbook(f'Sweep_{bw}MHZ_LTE.xlsx')
            logger.debug('Open Excel')

            if power_selected == 1:
                logger.debug(f'start to fill Sweep of Sensitivity and Power level')
                ws_name = f'Band{self.band}_Sweep_TxMax'
                # check sheet if it is in the workboook
                if ws_name not in wb.sheetnames:
                    wb.create_sheet(ws_name)

                ws = wb[ws_name]
                # check if it has the title at first row
                if ws.cell(1, 1).value is None:
                    self.create_sheet_title(ws)

                logger.debug('fill sensitivity and power')
                self.fill_progress_rx_sweep(self.std, ws, band, dl_ch, data,
                                            power_selected)  # progress of filling sensitivity progress


            elif power_selected == 0:
                logger.debug(f'start to fill Sweep of Sensitivity and Power level')
                ws_name = f'Band{self.band}_Sweep_TxMin'

                # check sheet if it is in the workboook
                if ws_name not in wb.sheetnames:
                    wb.create_sheet(f'Band{band}_Sweep_TxMin')

                ws = wb[ws_name]
                # check if it has the title at first row
                if ws.cell(1, 1).value is None:
                    self.create_sheet_title(ws)

                logger.debug('fill sensitivity')
                self.fill_progress_rx_sweep(self.std, ws, band, dl_ch, data,
                                            power_selected)  # progress of filling sensitivity progress

            wb.save(f'Sweep_{bw}MHZ_LTE.xlsx')
            wb.close()

            excel_path = f'Sweep_{bw}MHZ_LTE.xlsx'

            return excel_path

        elif self.std == 'WCDMA':
            if pathlib.Path(f'Sweep_WCDMA.xlsx').exists() is False:
                self.create_excel_rx_sweep(self.std, power_selected, bw)
                logger.debug('Create Excel')

            wb = openpyxl.load_workbook(f'Sweep_WCDMA.xlsx')
            logger.debug('Open Excel')

            if power_selected == 1:
                logger.debug(f'start to fill Sweep of  Sensitivity and Power level')
                ws_name = f'Band{band}_Sweep_TxMax'
                # check sheet if it is in the workboook
                if ws_name not in wb.sheetnames:
                    wb.create_sheet(ws_name)

                ws = wb[ws_name]
                # check if it has the title at first row
                if ws.cell(1, 1).value is None:
                    self.create_sheet_title(ws)

                logger.debug('fill sensitivity and power')
                self.fill_progress_rx_sweep(self.std, ws, band, dl_ch, data,
                                            power_selected)  # progress of filling sensitivity progress

            elif power_selected == 0:
                logger.debug(f'start to fill Sweep of  Sensitivity and Power level')
                ws_name = f'Band{band}_Sweep_TxMin'
                # check sheet if it is in the workboook
                if ws_name not in wb.sheetnames:
                    wb.create_sheet(ws_name)

                ws = wb[ws_name]
                # check if it has the title at first row
                if ws.cell(1, 1).value is None:
                    self.create_sheet_title(ws)

                logger.debug('fill sensitivity')
                self.fill_progress_rx_sweep(self.std, ws, band, dl_ch, data,
                                            power_selected)  # progress of filling sensitivity progress

            wb.save(f'Sweep_WCDMA.xlsx')
            wb.close()

            excel_path = f'Sweep_WCDMA.xlsx'

            return excel_path

    def fill_values_rx(self, data, band, dl_ch, power_selected, bw=None):
        """
            data format:[Tx Power, Sensitivity, PER]
        """
        if self.std == 'LTE':
            if pathlib.Path(f'Sensitivity_{bw}MHZ_LTE.xlsx').exists() is False:
                self.create_excel_rx(self.std, bw)
                logger.debug('Create Excel')

            wb = openpyxl.load_workbook(f'Sensitivity_{bw}MHZ_LTE.xlsx')
            logger.debug('Open Excel')

            if power_selected == 1:
                logger.debug(f'start to fill Sensitivity and Tx Power and Desens')
                ws = wb['Sensitivity_TxMax']
                logger.debug('fill sensitivity')
                self.fill_progress_rx(self.std, ws, band, dl_ch, data, 1, power_selected, bw)  # fill sensitivity
                ws = wb['PWR_TxMax']
                logger.debug('fill power')
                self.fill_progress_rx(self.std, ws, band, dl_ch, data, 0, power_selected, bw)  # fill power

            elif power_selected == 0:
                logger.debug(f'start to fill Sensitivity and Tx Power and Desens')
                ws = wb['Sensitivity_TxMin']
                logger.debug('fill sensitivity')
                self.fill_progress_rx(self.std, ws, band, dl_ch, data, 1, power_selected, bw)  # fill sensitivity
                ws = wb['PWR_TxMin']
                logger.debug('fill power')
                self.fill_progress_rx(self.std, ws, band, dl_ch, data, 0, power_selected, bw)  # fill power

            wb.save(f'Sensitivity_{bw}MHZ_LTE.xlsx')
            wb.close()

            excel_path = f'Sensitivity_{bw}MHZ_LTE.xlsx'

            return excel_path

        elif self.std == 'WCDMA':
            if pathlib.Path(f'Sensitivity_WCDMA.xlsx').exists() is False:
                self.create_excel_rx(self.std, bw)
                logger.debug('Create Excel')

            wb = openpyxl.load_workbook(f'Sensitivity_WCDMA.xlsx')
            logger.debug('Open Excel')

            if power_selected == 1:
                logger.debug(f'start to fill Sensitivity and Tx Power and Desens')
                ws = wb['Sensitivity_TxMax']
                logger.debug('fill sensitivity')
                self.fill_progress_rx(self.std, ws, band, dl_ch, data, 1, power_selected)  # fill sensitivity
                ws = wb['PWR_TxMax']
                logger.debug('fill power')
                self.fill_progress_rx(self.std, ws, band, dl_ch, data, 0, power_selected)  # fill power

            elif power_selected == 0:
                logger.debug(f'start to fill Sensitivity and Tx Power and Desens')
                ws = wb['Sensitivity_TxMin']
                logger.debug('fill sensitivity')
                self.fill_progress_rx(self.std, ws, band, dl_ch, data, 1, power_selected)  # fill sensitivity
                ws = wb['PWR_TxMin']
                logger.debug('fill power')
                self.fill_progress_rx(self.std, ws, band, dl_ch, data, 0, power_selected)  # fill power

            wb.save(f'Sensitivity_WCDMA.xlsx')
            wb.close()

            excel_path = f'Sensitivity_WCDMA.xlsx'

            return excel_path

    def fill_values_tx(self, data, band, dl_ch, bw=None):

        if self.std == 'LTE':
            """
                LTE format:{Q1:[power], Q_P:[power, ACLR, EVM], ...} and ACLR format is [L, M, H] 
            """
            excel_path = f'results_{bw}MHZ_LTE.xlsx'

            if pathlib.Path(excel_path).exists() is False:
                self.create_excel_tx(self.std, bw)
                logger.debug('Create Excel')

            wb = openpyxl.load_workbook(excel_path)
            logger.debug('Open Excel')
            for mod, test_items in data.items():

                ws = wb[f'PWR_{mod}']  # POWER
                logger.debug('start to fill Power')
                self.fill_progress_tx(self.std, ws, band, dl_ch, test_items, 0, bw)

                if mod != 'Q_1':
                    ws = wb[f'ACLR_{mod}']  # ACLR
                    logger.debug('start to fill ACLR')
                    self.fill_progress_tx(self.std, ws, band, dl_ch, test_items, 1, bw)

                    ws = wb[f'EVM_{mod}']  # EVM
                    logger.debug('start to fill EVM')
                    self.fill_progress_tx(self.std, ws, band, dl_ch, test_items, 2, bw)

            wb.save(excel_path)
            wb.close()

            return excel_path

        elif self.std == 'WCDMA':
            """
                WCDMA format:[power, ACLR, EVM], ...} and ACLR format is list format like [L, M, H]  
            """
            excel_path = f'results_WCDMA.xlsx'

            if self.chcoding == 'REFMEASCH':  # this is WCDMA
                if pathlib.Path(excel_path).exists() is False:
                    self.create_excel_tx(self.std)
                    logger.debug('Create Excel')

                wb = openpyxl.load_workbook(excel_path)

                logger.debug('Open Excel')

                ws = wb[f'PWR']  # POWER
                logger.debug('start to fill Power')
                self.fill_progress_tx(self.std, ws, band, dl_ch, data, 0)

                ws = wb[f'ACLR']  # ACLR
                logger.debug('start to fill ACLR')
                self.fill_progress_tx(self.std, ws, band, dl_ch, data, 1)

                ws = wb[f'EVM']  # EVM
                logger.debug('start to fill EVM')
                self.fill_progress_tx(self.std, ws, band, dl_ch, data, 2)

                wb.save(excel_path)
                wb.close()

                return excel_path

            elif self.chcoding == 'EDCHTEST':  # this is HSUPA
                """
                    HSUPA format:{subtest_number: [power, ACLR], ...} and ACLR format is list format like [L, M, H]  
                """
                excel_path = f'results_HSUPA.xlsx'

                if pathlib.Path(excel_path).exists() is False:
                    self.create_excel_tx(self.std, bw)
                    logger.debug('Create Excel')

                wb = openpyxl.load_workbook(excel_path)
                logger.debug('Open Excel')

                for subtest, test_items in data.items():
                    logger.info(f'start to fill subtest{subtest}')

                    ws = wb[f'PWR']  # POWER
                    logger.debug('start to fill Power')
                    self.fill_progress_hspa_tx(self.std, ws, band, dl_ch, test_items, 0, subtest)

                    ws = wb[f'ACLR']  # ACLR
                    logger.debug('start to fill ACLR')
                    self.fill_progress_hspa_tx(self.std, ws, band, dl_ch, test_items, 1, subtest)

                    wb.save(excel_path)
                    wb.close()

                return excel_path

            elif self.chcoding == 'FIXREFCH':  # this is HSDPA
                """
                    HSDPA format:{subtest_number: [power, ACLR], ...} and ACLR format is list format like [L, M, H]  
                    only subtest3 is for [power, ACLR, evm], evm to pickup the worst vaule from p0, p1, p2, p3
                """
                excel_path = f'results_HSDPA.xlsx'

                if pathlib.Path(excel_path).exists() is False:
                    self.create_excel_tx(self.std, bw)
                    logger.debug('Create Excel')

                wb = openpyxl.load_workbook(excel_path)
                logger.debug('Open Excel')
                for subtest, test_items in data.items():
                    logger.info(f'start to fill subtest{subtest}')

                    ws = wb[f'PWR']  # POWER
                    logger.debug('start to fill Power')
                    self.fill_progress_hspa_tx(self.std, ws, band, dl_ch, test_items, 0, subtest)

                    ws = wb[f'ACLR']  # ACLR
                    logger.debug('start to fill ACLR')
                    self.fill_progress_hspa_tx(self.std, ws, band, dl_ch, test_items, 1, subtest)

                    if subtest == 3:
                        ws = wb[f'EVM']  # EVM
                        logger.debug('start to fill EVM')
                        self.fill_progress_hspa_tx(self.std, ws, band, dl_ch, test_items, 2, subtest)

                    wb.save(excel_path)
                    wb.close()

                return excel_path

    def excel_plot_line(self, standard, excel_path):
        logger.debug('Start to plot line chart in Excel')
        if standard == 'LTE':
            try:
                wb = openpyxl.load_workbook(excel_path)
                for ws_name in wb.sheetnames:
                    ws = wb[ws_name]

                    if ws._charts != []:  # if there is charts, delete it
                        del ws._charts[0]

                    if 'PWR' in ws_name or 'EVM' in ws_name:
                        chart = LineChart()
                        chart.title = f'{ws_name[:3]}'
                        if 'PWR' in ws_name:
                            chart.y_axis.title = f'{ws_name[:3]}(dBm)'
                        elif 'EVM' in ws_name:
                            chart.y_axis.title = f'{ws_name[:3]}%'

                        chart.x_axis.title = 'Band'
                        chart.x_axis.tickLblPos = 'low'

                        chart.height = 20
                        chart.width = 32

                        y_data = Reference(ws, min_col=2, min_row=1, max_col=ws.max_column, max_row=ws.max_row)
                        x_data = Reference(ws, min_col=1, min_row=2, max_col=1, max_row=ws.max_row)
                        chart.add_data(y_data, titles_from_data=True)
                        chart.set_categories(x_data)

                        chart.series[0].graphicalProperties.line.dashStyle = 'dash'  # for L_ch
                        chart.series[1].graphicalProperties.line.width = 50000  # for M_ch
                        chart.series[2].marker.symbol = 'circle'  # for H_ch
                        chart.series[2].marker.size = 10

                        ws.add_chart(chart, "F1")

                        wb.save(excel_path)
                        wb.close()

                    elif 'Sensitivity' in ws_name or 'Desens' in ws_name:
                        chart = LineChart()
                        chart.title = f'{ws_name[:11]}'

                        chart.y_axis.title = f'Sensitivity(dBm)'

                        chart.x_axis.title = 'Band'
                        chart.x_axis.tickLblPos = 'low'

                        chart.height = 20
                        chart.width = 32

                        y_data = Reference(ws, min_col=2, min_row=1, max_col=ws.max_column, max_row=ws.max_row)
                        x_data = Reference(ws, min_col=1, min_row=2, max_col=1, max_row=ws.max_row)
                        chart.add_data(y_data, titles_from_data=True)
                        chart.set_categories(x_data)

                        chart.series[0].graphicalProperties.line.dashStyle = 'dash'  # for L_ch
                        chart.series[1].graphicalProperties.line.width = 50000  # for M_ch
                        chart.series[2].marker.symbol = 'circle'  # for H_ch
                        chart.series[2].marker.size = 10

                        ws.add_chart(chart, "F1")

                        wb.save(excel_path)
                        wb.close()

                    elif 'Sweep' in ws_name:
                        chart_sens = LineChart()
                        chart_sens.title = 'Sensitivity_Rx_Chan_Sweep'
                        chart_sens.y_axis.title = f'{ws_name[:11]}'
                        chart_sens.x_axis.title = 'Rx_chan'
                        chart_sens.x_axis.tickLblPos = 'low'
                        # chart_sens.y_axis.scaling.min = -60
                        # chart_sens.y_axis.scaling.max = -20

                        chart_sens.height = 20
                        chart_sens.width = 40

                        y_data_sens = Reference(ws, min_col=ws.max_column - 1, min_row=1, max_col=ws.max_column - 1,
                                                max_row=ws.max_row)

                        x_data = Reference(ws, min_col=1, min_row=2, max_col=2, max_row=ws.max_row)
                        chart_sens.add_data(y_data_sens, titles_from_data=True)
                        chart_sens.set_categories(x_data)

                        # chart_sens.y_axis.majorGridlines = None

                        chart_sens.series[0].marker.symbol = 'circle'  # for sensitivity
                        chart_sens.series[0].marker.size = 3

                        chart_pwr = LineChart()  # create a second chart

                        y_data_pwr = Reference(ws, min_col=ws.max_column, min_row=1, max_col=ws.max_column,
                                               max_row=ws.max_row)
                        chart_pwr.add_data(y_data_pwr, titles_from_data=True)

                        chart_pwr.series[0].graphicalProperties.line.dashStyle = 'dash'  # for power
                        chart_pwr.y_axis.title = 'Power(dBm)'
                        chart_pwr.y_axis.axId = 200
                        chart_pwr.y_axis.majorGridlines = None

                        chart_sens.y_axis.crosses = 'max'
                        chart_sens += chart_pwr

                        ws.add_chart(chart_sens, "J1")
                        # ws.add_chart(chart_sens, "J42")

                        wb.save(excel_path)
                        wb.close()


                    elif 'ACLR' in ws_name:
                        chart = LineChart()
                        chart.title = 'ACLR'
                        chart.y_axis.title = 'ACLR(dB)'
                        chart.x_axis.title = 'Band'
                        chart.x_axis.tickLblPos = 'low'
                        chart.y_axis.scaling.min = -60
                        chart.y_axis.scaling.max = -20

                        chart.height = 20
                        chart.width = 40

                        y_data = Reference(ws, min_col=3, min_row=1, max_col=ws.max_column, max_row=ws.max_row)
                        x_data = Reference(ws, min_col=1, min_row=2, max_col=2, max_row=ws.max_row)
                        chart.add_data(y_data, titles_from_data=True)
                        chart.set_categories(x_data)

                        chart.series[0].marker.symbol = 'triangle'  # for EUTRA_-1
                        chart.series[0].marker.size = 10
                        chart.series[1].marker.symbol = 'circle'  # for EUTRA_+1
                        chart.series[1].marker.size = 10
                        chart.series[2].graphicalProperties.line.width = 50000  # for UTRA_-1
                        chart.series[3].graphicalProperties.line.width = 50000  # for UTRA_+1
                        chart.series[4].graphicalProperties.line.dashStyle = 'dash'  # for UTRA_-2
                        chart.series[5].graphicalProperties.line.dashStyle = 'dash'  # for UTRA_+2

                        ws.add_chart(chart, "J1")

                        wb.save(excel_path)
                        wb.close()
            except TypeError as err:
                logger.debug(err)
                logger.info(f"This Band doesn't have this BW")

        elif standard == 'WCDMA':
            wb = openpyxl.load_workbook(excel_path)
            for ws_name in wb.sheetnames:
                ws = wb[ws_name]

                if ws._charts != []:  # if there is charts, delete it
                    del ws._charts[0]

                if 'PWR' in ws_name or 'EVM' in ws_name:
                    chart = LineChart()
                    chart.title = f'{ws_name[:3]}'
                    if 'PWR' in ws_name:
                        chart.y_axis.title = f'{ws_name[:3]}(dBm)'
                    elif 'EVM' in ws_name:
                        chart.y_axis.title = f'{ws_name[:3]}%'

                    chart.x_axis.title = 'Band'
                    chart.x_axis.tickLblPos = 'low'

                    chart.height = 20
                    chart.width = 32

                    if self.chcoding == 'REFMEASCH':  # this is WCDMA:
                        y_data = Reference(ws, min_col=2, min_row=1, max_col=ws.max_column, max_row=ws.max_row)
                        x_data = Reference(ws, min_col=1, min_row=2, max_col=1, max_row=ws.max_row)
                        chart.add_data(y_data, titles_from_data=True)
                        chart.set_categories(x_data)

                    else:  # HSUPA, HSDPA
                        y_data = Reference(ws, min_col=2, min_row=1, max_col=ws.max_column - 1, max_row=ws.max_row)
                        x_data = Reference(ws, min_col=1, min_row=2, max_col=1, max_row=ws.max_row)
                        chart.add_data(y_data, titles_from_data=True)
                        chart.set_categories(x_data)

                    chart.series[0].graphicalProperties.line.dashStyle = 'dash'  # for L_ch
                    chart.series[1].graphicalProperties.line.width = 50000  # for M_ch
                    chart.series[2].marker.symbol = 'circle'  # for H_ch
                    chart.series[2].marker.size = 10

                    ws.add_chart(chart, "F1")

                    wb.save(excel_path)
                    wb.close()

                elif 'ACLR' in ws_name:
                    chart = LineChart()
                    chart.title = 'ACLR'
                    chart.y_axis.title = 'ACLR(dB)'
                    chart.x_axis.title = 'Band'
                    chart.x_axis.tickLblPos = 'low'
                    chart.y_axis.scaling.min = -60
                    chart.y_axis.scaling.max = -20

                    chart.height = 20
                    chart.width = 40

                    if self.chcoding == 'REFMEASCH':  # this is WCDMA:
                        y_data = Reference(ws, min_col=3, min_row=1, max_col=ws.max_column, max_row=ws.max_row)
                        x_data = Reference(ws, min_col=1, min_row=2, max_col=2, max_row=ws.max_row)
                        chart.add_data(y_data, titles_from_data=True)
                        chart.set_categories(x_data)

                    else:
                        y_data = Reference(ws, min_col=3, min_row=1, max_col=ws.max_column - 1, max_row=ws.max_row)
                        x_data = Reference(ws, min_col=1, min_row=2, max_col=2, max_row=ws.max_row)
                        chart.add_data(y_data, titles_from_data=True)
                        chart.set_categories(x_data)

                    chart.series[0].graphicalProperties.line.width = 50000  # for UTRA_-1
                    chart.series[1].graphicalProperties.line.width = 50000  # for UTRA_+1
                    chart.series[2].graphicalProperties.line.dashStyle = 'dash'  # for UTRA_-2
                    chart.series[3].graphicalProperties.line.dashStyle = 'dash'  # for UTRA_+2

                    ws.add_chart(chart, "J1")

                    wb.save(excel_path)
                    wb.close()

                elif 'Sensitivity' in ws_name or 'Desens' in ws_name:
                    chart = LineChart()
                    chart.title = f'{ws_name[:11]}'

                    chart.y_axis.title = f'Sensitivity(dBm)'

                    chart.x_axis.title = 'Band'
                    chart.x_axis.tickLblPos = 'low'

                    chart.height = 20
                    chart.width = 32

                    y_data = Reference(ws, min_col=2, min_row=1, max_col=ws.max_column, max_row=ws.max_row)
                    x_data = Reference(ws, min_col=1, min_row=2, max_col=1, max_row=ws.max_row)
                    chart.add_data(y_data, titles_from_data=True)
                    chart.set_categories(x_data)

                    chart.series[0].graphicalProperties.line.dashStyle = 'dash'  # for L_ch
                    chart.series[1].graphicalProperties.line.width = 50000  # for M_ch
                    chart.series[2].marker.symbol = 'circle'  # for H_ch
                    chart.series[2].marker.size = 10

                    ws.add_chart(chart, "F1")

                    wb.save(excel_path)
                    wb.close()

        elif standard == 'GSM':
            pass

    def tx_core(self, standard, band, dl_ch, bw=None):
        conn_state = int(self.inst.query("CALLSTAT?").strip())
        self.dl_ch = dl_ch

        # calling process
        if standard == 'LTE':
            if conn_state != cm_pmt.ANRITSU_CONNECTED:
                self.set_init_before_calling(standard, dl_ch, bw)
                self.set_registration_calling(standard)
        elif standard == 'WCDMA' and self.chcoding == 'REFMEASCH':  # this is WCDMA
            if conn_state != cm_pmt.ANRITSU_LOOP_MODE_1:
                self.set_init_before_calling(standard, dl_ch, bw)
                self.set_registration_calling(standard)
        elif standard == 'WCDMA' and self.chcoding == 'EDCHTEST':  # this is HSUPA
            if conn_state != cm_pmt.ANRITSU_LOOP_MODE_1:
                self.set_init_before_calling(standard, dl_ch, bw)
                self.set_init_hspa()
                self.set_registration_calling(standard)
        elif standard == 'WCDMA' and self.chcoding == 'FIXREFCH':  # this is HSDPA
            if conn_state != cm_pmt.ANRITSU_LOOP_MODE_1:
                self.set_init_before_calling(standard, dl_ch, bw)
                self.set_init_hspa()
                self.set_registration_calling(standard)

        if standard == 'LTE':
            logger.info(f'Start to measure B{band}, bandwidth: {bw} MHz, downlink_chan: {dl_ch}')
        elif standard == 'WCDMA' and self.chcoding == 'REFMEASCH':  # this is WCDMA
            logger.info(f'Start WCDMA to measure B{band}, downlink_chan: {dl_ch}')
        elif standard == 'WCDMA' and self.chcoding == 'EDCHTEST':  # this is HSUPA
            logger.info(f'Start HSUPA to measure B{band}, downlink_chan: {dl_ch}')
        elif standard == 'WCDMA' and self.chcoding == 'FIXREFCH':  # this is HSDPA
            logger.info(f'Start HSDPA to measure B{band}, downlink_chan: {dl_ch}')

        self.set_handover(standard, dl_ch, bw)

        # HSUPA and HSDPA need to setting some parameters
        if standard == 'WCDMA' and (self.chcoding == 'EDCHTEST' or self.chcoding == 'FIXREFCH'):
            self.set_registration_after_calling_hspa()

        data = self.get_validation(standard)
        self.excel_path = self.fill_values_tx(data, band, dl_ch, bw)

    def rx_core(self, standard, band, dl_ch, bw=None):
        conn_state = int(self.inst.query("CALLSTAT?").strip())
        if standard == 'LTE':
            if conn_state != cm_pmt.ANRITSU_CONNECTED:
                self.set_init_before_calling(standard, dl_ch, bw)
                self.set_registration_calling(standard)
        elif standard == 'WCDMA':
            if conn_state != cm_pmt.ANRITSU_LOOP_MODE_1:
                self.set_init_before_calling(standard, dl_ch, bw)
                self.set_registration_calling(standard)

        if standard == 'LTE':
            logger.info(f'Start to sensitivity B{band}, bandwidth: {bw} MHz, downlink_chan: {dl_ch}')
        elif standard == 'WCDMA':
            logger.info(f'Start to sensitivity B{band}, downlink_chan: {dl_ch}')

        self.set_init_rx(standard)

        for power_selected in wt.tx_max_pwr_sensitivity:
            if power_selected == 1:
                self.set_tpc('ALL1')
                self.set_input_level(30)
                sens_list = self.get_sensitivity(standard, band, dl_ch, bw)
                logger.debug(f'Sensitivity list:{sens_list}')
                self.excel_path = self.fill_values_rx(sens_list, band, dl_ch, power_selected, bw)
                self.set_output_level(-70)
            elif power_selected == 0:
                if standard == 'LTE':
                    self.set_tpc('AUTO')
                elif standard == 'WCDMA':
                    self.set_tpc('ILPC')
                self.set_input_level(-10)
                sens_list = self.get_sensitivity(standard, band, dl_ch, bw)
                logger.debug(f'Sensitivity list:{sens_list}')
                self.excel_path = self.fill_values_rx(sens_list, band, dl_ch, power_selected, bw)
                self.set_output_level(-70)

    def rx_sweep_core(self, standard, band, dl_ch, bw=None):
        conn_state = int(self.inst.query("CALLSTAT?").strip())
        if standard == 'LTE':
            if conn_state != cm_pmt.ANRITSU_CONNECTED:
                self.set_init_before_calling(standard, dl_ch, bw)
                self.set_registration_calling(standard)
        elif standard == 'WCDMA':
            if conn_state != cm_pmt.ANRITSU_LOOP_MODE_1:
                self.set_init_before_calling(standard, dl_ch, bw)
                self.set_registration_calling(standard)

        if standard == 'LTE':
            logger.info(f'Start to sweep B{band}, bandwidth: {bw} MHz, downlink_chan: {dl_ch}')
        elif standard == 'WCDMA':
            logger.info(f'Start to sweep B{band}, downlink_chan: {dl_ch}')

        self.set_init_rx(standard)

        for power_selected in wt.tx_max_pwr_sensitivity:
            if power_selected == 1:
                self.set_tpc('ALL1')
                self.set_input_level(30)
                sens_list = self.get_sensitivity(standard, band, dl_ch, bw)
                logger.debug(f'Sensitivity list:{sens_list}')
                self.excel_path = self.fill_values_rx_sweep(sens_list, band, dl_ch, power_selected,
                                                            bw)  # this is different
                self.set_output_level(-70)
            elif power_selected == 0:
                if standard == 'LTE':
                    self.set_tpc('AUTO')
                elif standard == 'WCDMA':
                    self.set_tpc('ILPC')
                self.set_input_level(-10)
                sens_list = self.get_sensitivity(standard, band, dl_ch, bw)
                logger.debug(f'Sensitivity list:{sens_list}')
                self.excel_path = self.fill_values_rx_sweep(sens_list, band, dl_ch, power_selected,
                                                            bw)  # this is different
                self.set_output_level(-70)

    def run_tx(self):
        for tech in wt.tech:
            if tech == 'LTE' and wt.lte_bands != []:
                standard = self.switch_to_lte()
                self.inst.query('*OPC?')
                logger.info(standard)
                for bw in wt.lte_bandwidths:
                    for band in wt.lte_bands:
                        if bw in cm_pmt.bandwidths_selected(band):
                            self.set_test_parameter_normal()
                            ch_list = []
                            for wt_ch in wt.channel:
                                if wt_ch == 'L':
                                    ch_list.append(cm_pmt.dl_ch_selected(standard, band, bw)[0])
                                elif wt_ch == 'M':
                                    ch_list.append(cm_pmt.dl_ch_selected(standard, band, bw)[1])
                                elif wt_ch == 'H':
                                    ch_list.append(cm_pmt.dl_ch_selected(standard, band, bw)[2])
                            logger.debug(f'Test Channel List: {band}, {bw}MHZ, downlink channel list:{ch_list}')
                            for dl_ch in ch_list:
                                self.tx_core(standard, band, dl_ch, bw)
                        else:
                            logger.info(f'B{band} do not have BW {bw}MHZ')
                    self.excel_plot_line(standard, self.excel_path)
            elif (tech == 'WCDMA' or tech == 'HSUPA' or tech == 'HSDPA') and (
                    wt.wcdma_bands != [] or wt.hsupa_bands != [] or wt.hsdpa_bands != []):
                standard = self.switch_to_wcdma()
                self.inst.query('*OPC?')
                self.inst.write('CALLSO')

                if tech == 'WCDMA':
                    self.inst.write('CHCODING REFMEASCH')
                    logger.info('Set to WCDMA')
                elif tech == 'HSUPA':
                    self.inst.write('CHCODING EDCHTEST')
                    logger.info('Set to HSUPA')
                elif tech == 'HSDPA':
                    self.inst.write('CHCODING FIXREFCH')
                    logger.info('Set to HSDPA')

                self.chcoding = self.inst.query('CHCODING?').strip()
                logger.info(f'CHCODING: {self.chcoding}')

                if self.chcoding == 'REFMEASCH':  # this is WCDMA
                    for band in wt.wcdma_bands:
                        ch_list = []
                        for wt_ch in wt.channel:
                            if wt_ch == 'L':
                                ch_list.append(cm_pmt.dl_ch_selected(standard, band)[0])
                            elif wt_ch == 'M':
                                ch_list.append(cm_pmt.dl_ch_selected(standard, band)[1])
                            elif wt_ch == 'H':
                                ch_list.append(cm_pmt.dl_ch_selected(standard, band)[2])
                        logger.debug(f'Test Channel List: {band}, downlink channel list:{ch_list}')
                        for dl_ch in ch_list:
                            self.tx_core(standard, band, dl_ch)

                elif self.chcoding == 'EDCHTEST':  # this is HSUPA
                    for band in wt.hsupa_bands:
                        ch_list = []
                        for wt_ch in wt.channel:
                            if wt_ch == 'L':
                                ch_list.append(cm_pmt.dl_ch_selected(standard, band)[0])
                            elif wt_ch == 'M':
                                ch_list.append(cm_pmt.dl_ch_selected(standard, band)[1])
                            elif wt_ch == 'H':
                                ch_list.append(cm_pmt.dl_ch_selected(standard, band)[2])
                        logger.debug(f'Test Channel List: {band}, downlink channel list:{ch_list}')
                        for dl_ch in ch_list:
                            self.tx_core(standard, band, dl_ch)
                elif self.chcoding == 'FIXREFCH':  # this is HSDPA
                    for band in wt.hsdpa_bands:
                        ch_list = []
                        for wt_ch in wt.channel:
                            if wt_ch == 'L':
                                ch_list.append(cm_pmt.dl_ch_selected(standard, band)[0])
                            elif wt_ch == 'M':
                                ch_list.append(cm_pmt.dl_ch_selected(standard, band)[1])
                            elif wt_ch == 'H':
                                ch_list.append(cm_pmt.dl_ch_selected(standard, band)[2])
                        logger.debug(f'Test Channel List: {band}, downlink channel list:{ch_list}')
                        for dl_ch in ch_list:
                            self.tx_core(standard, band, dl_ch)

                self.excel_plot_line(standard, self.excel_path)
            elif tech == 'GSM' and wt.gsm_bands != []:
                pass

            else:
                logger.info(f'Finished')

    def run_rx(self):
        for tech in wt.tech:
            if tech == 'LTE' and wt.lte_bands != []:
                standard = self.switch_to_lte()
                logger.info(standard)
                for bw in wt.lte_bandwidths:
                    for band in wt.lte_bands:
                        if bw in cm_pmt.bandwidths_selected(band):
                            self.set_test_parameter_normal()
                            ch_list = []
                            for wt_ch in wt.channel:
                                if wt_ch == 'L':
                                    ch_list.append(cm_pmt.dl_ch_selected(standard, band, bw)[0])
                                elif wt_ch == 'M':
                                    ch_list.append(cm_pmt.dl_ch_selected(standard, band, bw)[1])
                                elif wt_ch == 'H':
                                    ch_list.append(cm_pmt.dl_ch_selected(standard, band, bw)[2])
                            logger.debug(f'Test Channel List: {band}, {bw}MHZ, downlink channel list:{ch_list}')
                            for dl_ch in ch_list:
                                self.rx_core(standard, band, dl_ch, bw)
                    self.fill_desens(self.excel_path)
                    self.excel_plot_line(standard, self.excel_path)
            elif tech == 'WCDMA' and wt.wcdma_bands != []:
                standard = self.switch_to_wcdma()
                for band in wt.wcdma_bands:
                    ch_list = []
                    for wt_ch in wt.channel:
                        if wt_ch == 'L':
                            ch_list.append(cm_pmt.dl_ch_selected(standard, band)[0])
                        elif wt_ch == 'M':
                            ch_list.append(cm_pmt.dl_ch_selected(standard, band)[1])
                        elif wt_ch == 'H':
                            ch_list.append(cm_pmt.dl_ch_selected(standard, band)[2])
                    logger.debug(f'Test Channel List: {band}, downlink channel list:{ch_list}')
                    for dl_ch in ch_list:
                        self.rx_core(standard, band, dl_ch)
                self.fill_desens(self.excel_path)
                self.excel_plot_line(standard, self.excel_path)
            elif tech == wt.gsm_bands:
                pass
            else:
                logger.info(f'Finished')

    def run_rx_sweep_ch(self):
        for tech in wt.tech:
            if tech == 'LTE' and wt.lte_bands != []:
                standard = self.switch_to_lte()
                logger.info(standard)
                for bw in wt.lte_bandwidths:
                    for band in wt.lte_bands:
                        if bw in cm_pmt.bandwidths_selected(band):
                            logger.info(f'Sweep Channel List: {band}, {bw}MHZ')
                            self.set_test_parameter_normal()
                            lch = cm_pmt.dl_ch_selected(standard, band, bw)[0]
                            hch = cm_pmt.dl_ch_selected(standard, band, bw)[2]
                            step = cm_pmt.SWEEP_STEP
                            if cm_pmt.CHAN_LIST:
                                ch_list = cm_pmt.CHAN_LIST
                            else:
                                ch_list = range(lch, hch + 1, step)
                            for dl_ch in ch_list:
                                self.rx_sweep_core(standard, band, dl_ch, bw)
                        self.excel_plot_line(standard, self.excel_path)
            elif tech == 'WCDMA' and wt.wcdma_bands != []:
                standard = self.switch_to_wcdma()
                for band in wt.wcdma_bands:
                    logger.info(f'Sweep Channel List: {band}')
                    lch = cm_pmt.dl_ch_selected(standard, band)[0]
                    hch = cm_pmt.dl_ch_selected(standard, band)[2]
                    step = cm_pmt.SWEEP_STEP
                    if cm_pmt.CHAN_LIST:
                        ch_list = cm_pmt.CHAN_LIST
                    else:
                        ch_list = range(lch, hch + 1, step)
                    for dl_ch in ch_list:
                        self.rx_sweep_core(standard, band, dl_ch)
                    self.excel_plot_line(standard, self.excel_path)
            elif tech == wt.gsm_bands:
                pass
            else:
                logger.info(f'Finished')


def main():
    start = datetime.datetime.now()

    anritsu = Anritsu8820()
    anritsu.run_tx()  # run_tx() | run_rx() | run_rx_sweep_ch()
    # anritsu.excel_plot_line('WCDMA', 'results_HSDPA.xlsx')
    stop = datetime.datetime.now()

    logger.info(f'Timer: {stop - start}')


if __name__ == '__main__':
    main()
