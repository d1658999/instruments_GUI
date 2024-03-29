import pathlib

import pyvisa
import time
import datetime
import logging
from logging.config import fileConfig
import openpyxl
from decimal import Decimal
from openpyxl.chart import LineChart, Reference, ScatterChart

from loss_table import loss_table
import common_parameters as cm_pmt
import want_test_band as wt
from fly_mode import Flymode, get_comport_wanted
from anritsu8820 import Anritsu8820

fileConfig('logging.ini')
logger = logging.getLogger()


class Anritsu8821(Anritsu8820):
    def __init__(self, psu_object=None):
        super().__init__(psu_object)

    def set_init_before_calling_lte(self, dl_ch, bw):
        """
            preset before start to calling for LTE
        """
        self.preset()
        self.set_band_cal()
        self.set_screen('OFF')
        self.set_display_remain()
        self.preset_extarb()
        self.set_lvl_status('OFF')
        self.set_test_mode('OFF')
        self.set_integrity('LTE', 'SNOW3G')
        self.set_scenario()
        self.set_pdn_type()
        self.set_mcc_mnc()
        self.set_ant_config()
        self.set_imsi()
        self.set_authentication()
        self.set_all_measurement_items_off()
        self.set_init_miscs('LTE')
        self.set_path_loss('LTE')
        self.set_init_level('LTE')
        self.set_handover('LTE', dl_ch, bw)
        self.inst.write('ULRB_POS MIN')
        self.inst.query('*OPC?')

    def set_path_loss(self, standard):
        logger.info('Set LOSS')
        self.inst.write('DELLOSSTBL')  # delete the unknown loss table first

        loss_title = 'LOSSTBLVAL2'
        freq = sorted(loss_table.keys())
        for keys in freq:
            loss = f'{loss_title} {str(keys)}MHZ, {str(loss_table[keys])}, {str(loss_table[keys])},,, {str(loss_table[keys])},,,'
            logger.info(loss)
            self.inst.write(loss)
        self.inst.write('DELLOSSTBL_P2')
        s = standard  # WCDMA|GSM|LTE
        logger.debug("Current Format: " + s)
        if s == 'LTE':
            self.inst.write("EXTLOSSW COMMON")
            self.inst.query('*OPC?')
        elif s == 'WCDMA':
            self.inst.write("DLEXTLOSSW COMMON")  # Set DL external loss to COMMON
            self.inst.write("ULEXTLOSSW COMMON")  # Set UL external loss to COMMON
            self.inst.write("AUEXTLOSSW COMMON")  # Set AUX external loss to COMMON
            self.inst.query('*OPC?')
        elif s == "GSM":
            self.inst.write("EXTLOSSW COMMON")
            self.inst.query('*OPC?')

    def set_init_miscs(self, standard):
        s = standard
        if s == 'LTE':
            self.inst.write('SEM_ADDREQ REL11')
            self.inst.write('RRCUPDATE PAGING')
            self.inst.write('PT_TRGSRC FRAME')
            self.inst.write('MODIFPERIOD N2')
            self.inst.write('PCYCLE 32')
            self.inst.write('RRCRELEASE OFF')
            self.inst.write('FREQERRRNG NARROW')  # NORMAL | NARROW
            self.inst.write('ROBUSTCON OFF')  # ON | OFF
            self.inst.write('TESTMODE OFF')
            self.inst.write('UECAT CAT3')
            self.inst.write('SIB2_NS NS_01')
            self.inst.write('ULTPSEL 1')
            self.inst.write('DLTPSEL 1')
            self.inst.write('TXOUT 1,MAIN')

            logger.debug(self.inst.query('UE_CAP? REL'))

    def set_registration_calling_lte(self, times=30):
        """
            ANRITSU_IDLE = 1	        #Idle state
            ANRITSU_REGIST = 3			# Under location registration
            ANRITSU_CONNECTED = 6	    # Under communication or connected
        """
        self.inst.write('CALLTHLD 1')
        self.set_lvl_status('ON')
        self.set_test_mode()
        conn_state = int(self.inst.query("CALLSTAT?").strip())
        while conn_state != cm_pmt.ANRITSU_CONNECTED:  # this is for waiting connection
            self.inst.write('CALLRFR')
            while conn_state == cm_pmt.ANRITSU_IDLE:
                self.inst.write('CALLSO')
                logger.info('IDLE')
                logger.info('Start to ON and OFF')
                self.flymode_circle()
                time.sleep(3)
                self.inst.write('CALLSA')
                while conn_state != cm_pmt.ANRITSU_CONNECTED:
                    logger.info('Waiting for connection...')
                    time.sleep(1)
                    conn_state = int(self.inst.query("CALLSTAT?").strip())

                # logger.info('Start calling')
                # conn_state = int(self.inst.query("CALLSTAT?").strip())
            # logger.info('START CALL')
            # self.inst.write('CALLSA')
            logger.info('Connected')
            time.sleep(1)

    def get_power_aclr_evm_lte(self):
        """
            Only measure RB@min
            The format in dictionary is {Q_1: [power, aclr, evm], Q_P: [power, aclr, evm], ...}
            and ACLR format is [EUTRA-1, EUTRA+1, UTRA-1, URTA+1, UTRA-2, URTA+2,]
        """
        want_mods = [
            'TESTPRM TX_MAXPWR_Q_1',
            'TESTPRM TX_MAXPWR_Q_P',
            'TESTPRM TX_MAXPWR_Q_F',
            'TESTPRM TX_MAXPWR_16_P',
            'TESTPRM TX_MAXPWR_16_F',
            'TESTPRM TX_MAXPWR_64_P',
            'TESTPRM TX_MAXPWR_64_F',
            'TESTPRM TX_MAXPWR_256_F',
        ]

        validation_dict = {}

        self.set_init_power()
        self.set_init_aclr('LTE')
        self.set_init_mod('LTE')
        self.set_input_level(26)
        self.set_tpc('ALL3')
        self.inst.query('*OPC?')

        for mod in want_mods:
            self.mod = mod[18:]
            conn_state = int(self.inst.query("CALLSTAT?").strip())
            self.count = 5
            while conn_state != cm_pmt.ANRITSU_CONNECTED:  # this is for waiting connection before change modulation if there is connection problems
                logger.info('Call drops...')
                if self.count == 0:
                    # equipment end call and start call
                    logger.info('End call and then start call')
                    self.flymode_circle()
                    time.sleep(5)
                    self.inst.write('CALLSO')
                    self.inst.query('*OPC?')
                    time.sleep(1)
                    self.inst.write('CALLSA')
                    time.sleep(10)
                    self.count = 6
                    conn_state = int(self.inst.query("CALLSTAT?").strip())

                else:
                    time.sleep(10)
                    self.count -= 1
                    logger.info('wait 10 seconds to connect')
                    logger.info(f'{6 - self.count} times to wait 10 second')
                    conn_state = int(self.inst.query("CALLSTAT?").strip())

            validation_list = []
            if wt.fdd_tdd_cross_test == 1:
                if mod in ['TESTPRM TX_MAXPWR_Q_1', 'TESTPRM TX_MAXPWR_Q_P', 'TESTPRM TX_MAXPWR_Q_F']:
                    self.inst.write(mod)
                    self.inst.write('ULRMC_64QAM ENABLED')
                    self.inst.write('ULRMC_256QAM ENABLED')
                    self.inst.write('ULIMCS 5')

                elif mod in ['TESTPRM TX_MAXPWR_16_P', 'TESTPRM TX_MAXPWR_16_F']:
                    self.inst.write(mod)
                    self.inst.write('ULRMC_64QAM ENABLED')
                    self.inst.write('ULRMC_256QAM ENABLED')
                    self.inst.write('ULIMCS 13')

                elif mod in ['TESTPRM TX_MAXPWR_64_P', 'TESTPRM TX_MAXPWR_64_F']:
                    self.inst.write(mod)
                    self.inst.write('ULRMC_64QAM ENABLED')
                    self.inst.write('ULRMC_256QAM ENABLED')
                    self.inst.write('ULIMCS 22')

                elif mod in ['TESTPRM TX_MAXPWR_256_F']:
                    self.inst.write(mod)
                    self.inst.write('ULRMC_64QAM ENABLED')
                    self.inst.write('ULRMC_256QAM ENABLED')
                    self.inst.write('ULIMCS 28')
            else:
                self.inst.write(mod)
            self.set_to_measure()
            self.inst.query('*OPC?')
            meas_status = int(self.inst.query('MSTAT?').strip())

            while meas_status == cm_pmt.MESUREMENT_BAD:  # this is for the reference signal is not found
                logger.info('measuring status is bad(Reference signal not found)')
                logger.info('Equipment is forced to set End Call')
                self.inst.write('CALLSO')
                time.sleep(5)
                logger.info('fly on and off again')
                self.flymode_circle()
                time.sleep(10)
                self.inst.write('CALLSA')
                logger.info('waiting for 10 second to re-connect')
                logger.info('measure it again')
                self.set_to_measure()
                meas_status = int(self.inst.query('MSTAT?').strip())

            # self.inst.query('*OPC?')

            if mod == 'TESTPRM TX_MAXPWR_Q_1':  # mod[18:] -> Q_1
                logger.info(mod)
                validation_list.append(self.get_uplink_power('LTE'))
                validation_dict[mod[18:]] = validation_list
                # self.inst.query('*OPC?')
            else:  # mod[18:] -> Q_P, Q_F, 16_P, 16_F, 64_F, 256_F
                logger.info(mod)
                self.pwr = self.get_uplink_power('LTE')
                validation_list.append(self.pwr)
                self.aclr = self.get_uplink_aclr('LTE')
                validation_list.append(self.aclr)
                self.evm = self.get_uplink_evm('LTE')
                validation_list.append(self.evm)
                validation_dict[mod[18:]] = validation_list
                self.inst.query('*OPC?')
        logger.debug(validation_dict)
        return validation_dict

    @staticmethod
    def create_excel_tx(standard, bw=None):
        if standard == 'LTE':
            wb = openpyxl.Workbook()
            wb.remove(wb['Sheet'])
            wb.create_sheet('PWR_Q_1')
            wb.create_sheet('PWR_Q_P')
            wb.create_sheet('PWR_Q_F')
            wb.create_sheet('PWR_16_P')
            wb.create_sheet('PWR_16_F')
            wb.create_sheet('PWR_64_P')
            wb.create_sheet('PWR_64_F')
            wb.create_sheet('PWR_256_F')
            wb.create_sheet('ACLR_Q_P')
            wb.create_sheet('ACLR_Q_F')
            wb.create_sheet('ACLR_16_P')
            wb.create_sheet('ACLR_16_F')
            wb.create_sheet('ACLR_64_P')
            wb.create_sheet('ACLR_64_F')
            wb.create_sheet('ACLR_256_F')
            wb.create_sheet('EVM_Q_P')
            wb.create_sheet('EVM_Q_F')
            wb.create_sheet('EVM_16_P')
            wb.create_sheet('EVM_16_F')
            wb.create_sheet('EVM_64_P')
            wb.create_sheet('EVM_64_F')
            wb.create_sheet('EVM_256_F')

            for sheet in wb.sheetnames:
                if 'ACLR' in sheet:
                    sh = wb[sheet]
                    sh['A1'] = 'Band'
                    sh['B1'] = 'Channel'
                    sh['C1'] = 'EUTRA_-1'
                    sh['D1'] = 'EUTRA_+1'
                    sh['E1'] = 'UTRA_-1'
                    sh['F1'] = 'UTRA_+1'
                    sh['G1'] = 'UTRA_-2'
                    sh['H1'] = 'UTRA_+2'

                else:
                    sh = wb[sheet]
                    sh['A1'] = 'Band'
                    sh['B1'] = 'ch0'
                    sh['C1'] = 'ch1'
                    sh['D1'] = 'ch2'
            excel_path = f'results_{bw}MHZ_LTE.xlsx' if wt.condition is None else f'results_{bw}MHZ_LTE_{wt.condition}.xlsx'
            wb.save(excel_path)
            wb.close()

        elif standard == 'WCDMA':
            wb = openpyxl.Workbook()
            wb.remove(wb['Sheet'])
            wb.create_sheet('PWR')
            wb.create_sheet('ACLR')
            wb.create_sheet('EVM')

            for sheet in wb.sheetnames:
                if 'ACLR' in sheet:
                    sh = wb[sheet]
                    sh['A1'] = 'Band'
                    sh['B1'] = 'Channel'
                    sh['C1'] = 'UTRA_-1'
                    sh['D1'] = 'UTRA_+1'
                    sh['E1'] = 'UTRA_-2'
                    sh['F1'] = 'UTRA_+2'
                else:
                    sh = wb[sheet]
                    sh['A1'] = 'Band'
                    sh['B1'] = 'ch01'
                    sh['C1'] = 'ch02'
                    sh['D1'] = 'ch03'
            excel_path = f'results_WCDMA.xlsx' if wt.condition is None else f'results_WCDMA_{wt.condition}.xlsx'
            wb.save(excel_path)
            wb.close()

    def run_tx(self):
        for tech in wt.tech:
            if tech == 'LTE' and wt.lte_bands != []:
                standard = self.switch_to_lte()
                logger.info(standard)
                for bw in wt.lte_bandwidths:
                    for band in wt.lte_bands:
                        if bw in cm_pmt.bandwidths_selected(band):
                            if band == 28:
                                self.band_segment = wt.band_segment
                            self.set_test_parameter_normal()
                            ch_list = []
                            for wt_ch in wt.channel:
                                if wt_ch == 'L':
                                    ch_list.append(cm_pmt.dl_ch_selected(standard, band, bw)[0])
                                elif wt_ch == 'M':
                                    ch_list.append(cm_pmt.dl_ch_selected(standard, band, bw)[1])
                                elif wt_ch == 'H':
                                    ch_list.append(cm_pmt.dl_ch_selected(standard, band, bw)[2])
                            logger.debug(f'Test Channel List: {band}, {bw}MHZ, downlink channel list:{ch_list}')
                            for dl_ch in ch_list:
                                self.band = band

                                conn_state = int(self.inst.query("CALLSTAT?").strip())
                                if conn_state != cm_pmt.ANRITSU_CONNECTED:
                                    self.set_init_before_calling(standard, dl_ch, bw)
                                    self.set_registration_calling(standard)
                                logger.info(f'Start to measure B{band}, bandwidth: {bw} MHz, downlink_chan: {dl_ch}')
                                self.set_handover(standard, dl_ch, bw)
                                time.sleep(2)
                                data = self.get_validation(standard)
                                self.excel_path = self.fill_values_tx(data, band, dl_ch, bw)
                                self.set_test_parameter_normal()  # this is for 8821 to be more stable
                        else:
                            logger.info(f'B{band} do not have BW {bw}MHZ')
                    self.excel_plot_line(standard, self.excel_path)
            else:
                logger.info(f'Finished')


def main():
    start = datetime.datetime.now()

    anritsu = Anritsu8821()
    # anritsu.create_excel_rx('LTE', 5)
    anritsu.run_rx()

    stop = datetime.datetime.now()
    logger.info(f'Timer: {stop - start}')


if __name__ == '__main__':
    main()
