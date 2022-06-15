import serial
import time
import datetime
import serial.tools.list_ports
import logging
import pyvisa
from logging.config import fileConfig
import openpyxl
from openpyxl.chart import LineChart, Reference, ScatterChart
import pathlib
import scripts
from varname import nameof
import math

import common_parameters_ftm as cm_pmt_ftm
import want_test_band as wt
from loss_table import loss_table

fileConfig('logging.ini')
logger = logging.getLogger()

class CMW100:
    def __init__(self):
        self.begin_serial()
        self.get_resource()
        self.parameters_init()

    def parameters_init(self):
        self.asw_on_off = 0  # 1: AS ON, 0: AS OFF
        self.asw_tech_dict = {
            'GSM': 0,
            'WCDMA': 1,
            'LTE': 2,
            'FR1': 6,
        }
        self.asw_tech = None
        self.asw_path = 0
        self.rx_level = -70
        self.tx_level = None
        self.port = None
        self.chan = None
        self.sync_path = 0  # 0: Main, 1: 4RX, 2: 6RX
        self.sync_mode = 0  # 0: Main, 1: CA#1, 2: CA#2, 3: CA#3
        self.sa_nsa_mode = None  # SA: 0, NSA: 1
        self.filename = None
        self.tech = None
        self.band_lte = None
        self.band_fr1 = None
        self.bw_lte = None
        self.bw_fr1 = None
        self.tx_freq_lte = None
        self.rx_freq_lte = None
        self.tx_freq_fr1 = None
        self.rx_freq_fr1 = None
        self.mcs_lte = None
        self.mcs_fr1 = None
        self.tx_path = None
        self.scs = None
        self.type_fr1 = None
        self.rb_state = None
        self.rb_size_lte = None
        self.rb_size_fr1 = None
        self.rb_start_lte = None
        self.rb_start_fr1 = None
        self.loss_tx = None
        self.loss_rx = None
        self.tx_path_dict = {
            'TX1': 0,
            'TX2': 1,
        }
        self.bw_lte_dict = {
            1.4: 0,
            3: 1,
            5: 2,
            10: 3,
            15: 4,
            20: 5,
        }
        self.bw_fr1_dict = {
            5: 0,
            10: 1,
            15: 2,
            20: 3,
            25: 4,
            30: 5,
            40: 6,
            50: 7,
            60: 8,
            80: 9,
            90: 10,
            100: 11,
            70: 12,
        }
        self.mcs_lte_dict = {
            'QPSK': 0,
            'Q16': 11,
            'Q64': 25,
            'Q256': 27,
        }
        self.mcs_fr1_dict = {
            'BPSK': 1,
            'QPSK': 2,
            'Q16': 4,
            'Q64': 6,
            'Q256': 8,
        }
        self.rb_select_lte_dict = {
            'PRB': 0,
            'FRB': 1,
        }
        self.type_dict = {
            'DFTS': 0,
            'CP': 1,
        }
        self.rb_alloc_fr1_dict = {
            'EDGE_FULL_LEFT': 0,
            'EDGE_FULL_RIGHT': 1,
            'EDGE_1RB_LEFT': 2,
            'EDGE_1RB_RIGHT': 3,
            'OUTER_FULL': 4,
            'INNER_FULL': 5,
            'INNER_1RB_LEFT': 6,
            'INNER_1RB_RIGHT': 7,
        }
        self.scs_dict = {
            15: 0,
            30: 1,
            60: 2,
        }
    def begin_serial(self):
        self.ser = serial.Serial()
        self.ser.baudrate = 230400
        self.ser.timeout = 0.15
        self.ser.port = self.get_comport_wanted()
        self.com_open()

    def com_open(self):
        try:
            self.ser.open()
            logger.info('open modem comport')
        except:
            logger.info('check the comport is locked or drop comport')

    def com_close(self):
        self.ser.close()

    def get_resource(self):
        self.cmw100 = pyvisa.ResourceManager().open_resource('TCPIP0::127.0.0.1::INSTR')
        self.cmw100.timeout = 5000
        logger.info('Connect to CMW100')
        logger.info('TCPIP0::127.0.0.1::INSTR')

    def preset_instrument(self):
        logger.info('----------Preset CMW100----------')
        self.command_cmw100_write('CONFigure:FDCorrection:DEACtivate RFAC')
        self.command_cmw100_write('CONFigure:BASE:FDCorrection:CTABle:DELete:ALL')
        self.command_cmw100_query('*OPC?')
        self.command_cmw100_write('*RST')
        self.command_cmw100_query('*OPC?')

    @staticmethod
    def get_comport_wanted():
        comports = serial.tools.list_ports.comports()
        comport_waned = None
        for comport in comports:
            if 'Modem' in comport.description:
                comport_waned = comport.name
                logger.info(f'Modem comport is: {comport_waned}')
        return comport_waned

    @staticmethod
    def get_loss(freq):
        want_loss = None
        for f in loss_table.keys():
            if freq > f*1000:
                want_loss = loss_table[f]
            elif f*1000 > freq:
                break
        return want_loss

    @staticmethod
    def chan_judge_lte(band_lte, bw_lte, tx_freq_lte):
        rx_freq_lte = cm_pmt_ftm.transfer_freq_tx2rx_lte(band_lte, tx_freq_lte)
        if rx_freq_lte < cm_pmt_ftm.dl_freq_selected('LTE', band_lte, bw_lte)[1]:
            return 'L'
        elif rx_freq_lte == cm_pmt_ftm.dl_freq_selected('LTE', band_lte, bw_lte)[1]:
            return 'M'
        elif rx_freq_lte > cm_pmt_ftm.dl_freq_selected('LTE', band_lte, bw_lte)[1]:
            return 'H'

    @staticmethod
    def chan_judge_fr1(band_fr1, bw_fr1, tx_freq_fr1):
        rx_freq_fr1 = cm_pmt_ftm.transfer_freq_tx2rx_fr1(band_fr1, tx_freq_fr1)
        if rx_freq_fr1 < cm_pmt_ftm.dl_freq_selected('FR1', band_fr1, bw_fr1)[1]:
            return 'L'
        elif rx_freq_fr1 == cm_pmt_ftm.dl_freq_selected('FR1', band_fr1, bw_fr1)[1]:
            return 'M'
        elif rx_freq_fr1 > cm_pmt_ftm.dl_freq_selected('FR1', band_fr1, bw_fr1)[1]:
            return 'H'

    def set_test_mode_lte(self):
        logger.info('----------Set Test Mode----------')
        self.command(f'AT+LRFFINALSTART=1,{self.band_lte}')
        self.command(f'AT+LMODETEST')

    def set_test_mode_fr1(self):  # SA: 0, NSA: 1
        """
        SA: 0, NSA: 1
        """
        logger.info('----------Set Test Mode----------')
        self.command(f'AT+NRFFINALSTART={self.band_fr1},{self.sa_nsa_mode}')

    def set_test_end_lte(self):
        logger.info('----------Set End----------')
        self.command(f'AT+LRFFINALFINISH')

    def set_test_end_fr1(self):
        logger.info('----------Set End----------')
        self.command(f'AT+NRFFINALFINISH')

    def sig_gen_lte(self):
        logger.info('----------Sig Gen----------')
        self.command_cmw100_query('SYSTem:BASE:OPTion:VERSion?  "CMW_NRSub6G_Meas"')
        self.command_cmw100_write('ROUT:GPRF:GEN:SCEN:SAL R118, TX1')
        self.command_cmw100_query('*OPC?')
        self.command_cmw100_write('CONFigure:GPRF:GEN:CMWS:USAGe:TX:ALL R118, ON,ON,ON,ON,ON,ON,ON,ON')
        self.command_cmw100_query('*OPC?')
        self.command_cmw100_write('SOUR:GPRF:GEN1:LIST OFF')
        self.command_cmw100_query('*OPC?')
        self.command_cmw100_write(f'SOUR:GPRF:GEN1:RFS:EATT {self.loss_rx}')
        self.command_cmw100_query('*OPC?')
        self.command_cmw100_write('SOUR:GPRF:GEN1:BBM ARB')
        self.command_cmw100_query('*OPC?')
        if self.band_lte in [34, 38, 39, 40, 41, 42, 48]:
            self.command_cmw100_write(f"SOUR:GPRF:GEN1:ARB:FILE 'C:\CMW100_WV\SMU_Channel_CC0_RxAnt0_RF_Verification_10M_SIMO_01.wv'")
        else:
            self.command_cmw100_write(f"SOUR:GPRF:GEN1:ARB:FILE 'C:\CMW100_WV\SMU_NodeB_Ant0_FRC_10MHz.wv'")
        self.command_cmw100_query('*OPC?')
        self.command_cmw100_query('SOUR:GPRF:GEN1:ARB:FILE?')
        self.command_cmw100_write(f'SOUR:GPRF:GEN1:RFS:FREQ {self.rx_freq_lte}KHz')
        self.command_cmw100_write(f'SOUR:GPRF:GEN1:RFS:LEV {self.rx_level}.000000')
        self.command_cmw100_query('SOUR:GPRF:GEN1:STAT?')
        self.command_cmw100_write(f'SOUR:GPRF:GEN1:STAT ON')
        self.command_cmw100_query('SOUR:GPRF:GEN1:STAT?')

    def sig_gen_fr1(self):
        """
        scs: FDD is forced to 15KHz and TDD is to be 30KHz
        """
        logger.info('----------Sig Gen----------')
        scs = 1 if self.band_fr1 in [34, 38, 39, 40, 41, 42, 48, 77, 78, 79] else 0  # for now FDD is forced to 15KHz and TDD is to be 30KHz
        scs = 15 * (2 ** scs)
        mcs_fr1_wv = 4
        self.scs = scs
        self.command_cmw100_query('SYSTem:BASE:OPTion:VERSion?  "CMW_NRSub6G_Meas"')
        self.command_cmw100_write('ROUT:GPRF:GEN:SCEN:SAL R118, TX1')
        self.command_cmw100_query('*OPC?')
        self.command_cmw100_write('CONFigure:GPRF:GEN:CMWS:USAGe:TX:ALL R118, ON,ON,ON,ON,ON,ON,ON,ON')
        self.command_cmw100_query('*OPC?')
        self.command_cmw100_write('SOUR:GPRF:GEN1:LIST OFF')
        self.command_cmw100_query('*OPC?')
        self.command_cmw100_write(f'SOUR:GPRF:GEN1:RFS:EATT {self.loss_rx}')
        self.command_cmw100_query('*OPC?')
        self.command_cmw100_write('SOUR:GPRF:GEN1:BBM ARB')
        self.command_cmw100_query('*OPC?')
        self.command_cmw100_write('CONFigure:NRSub:MEAS:ULDL:PERiodicity MS10')
        self.command_cmw100_query('*OPC?')
        if self.band_fr1 in [34, 38, 39, 40, 41, 42, 48, 77, 78, 79]:
            self.command_cmw100_write(f"SOUR:GPRF:GEN1:ARB:FILE 'C:\CMW100_WV\SMU_NodeB_NR_Ant0_NR_{self.bw_fr1}MHz_SCS{scs}_TDD_Sens_MCS{mcs_fr1_wv}_rescale.wv'")
        else:
            self.command_cmw100_write(f"SOUR:GPRF:GEN1:ARB:FILE 'C:\CMW100_WV\SMU_NodeB_NR_Ant0_LTE_NR_{self.bw_fr1}MHz_SCS{scs}_FDD_Sens_MCS{mcs_fr1_wv}.wv'")
        self.command_cmw100_query('*OPC?')
        self.command_cmw100_query('SOUR:GPRF:GEN1:ARB:FILE?')
        self.command_cmw100_write(f'SOUR:GPRF:GEN1:RFS:FREQ {self.rx_freq_fr1}KHz')
        self.command_cmw100_write(f'SOUR:GPRF:GEN1:RFS:LEV {self.rx_level}.000000')
        self.command_cmw100_query('SOUR:GPRF:GEN1:STAT?')
        self.command_cmw100_write('SOUR:GPRF:GEN1:STAT ON')
        self.command_cmw100_query('*OPC?')
        self.command_cmw100_query('SOUR:GPRF:GEN1:STAT?')

    def sync_lte(self):
        logger.info('---------Sync----------')
        response = self.command(f'AT+LSYNC={self.sync_path},{self.sync_mode},{self.rx_freq_lte}', delay=1.2)
        while b'+LSYNC:1\r\n' not in response:
            logger.info('**********Sync repeat**********')
            time.sleep(1)
            response = self.command(f'AT+LSYNC={self.sync_path},{self.sync_mode},{self.rx_freq_lte}', delay=2)

    def sync_fr1(self):
        logger.info('---------Sync----------')
        scs = 1 if self.band_fr1 in [34, 38, 39, 40, 41, 42, 48, 75, 76, 77, 78, 79] else 0
        response = self.command(f'AT+NRFSYNC={self.sync_path},{self.sync_mode},{scs},{self.bw_fr1_dict[self.bw_fr1]},0,{self.rx_freq_fr1}', delay=1.2)
        while b'+NRFSYNC:1\r\n' not in response:
            logger.info('**********Sync repeat**********')
            time.sleep(1)
            response = self.command(f'AT+NRFSYNC={self.sync_path},{self.sync_mode},{scs},{self.bw_fr1_dict[self.bw_fr1]},0,{self.rx_freq_fr1}', delay=2)

    def tx_set_lte(self):
        """
        :param tx_path: TX1: 0 (main path)| TX2: 1 (sub path)
        :param bw_lte: 1.4: 0 | 3: 1 | 5: 2 | 10: 3 | 15: 4 | 20: 5
        :param tx_freq_lte:
        :param rb_num:
        :param rb_start:
        :param mcs: "QPSK": 0 | "Q16": 11 | "Q64": 25 | "Q256": 27
        :param pwr:
        :return:
        """
        logger.info('---------Tx Set----------')
        self.command(f'AT+LTXSENDREQ={self.tx_path_dict[self.tx_path]},{self.bw_lte_dict[self.bw_lte]},{self.tx_freq_lte},{self.rb_size_lte},{self.rb_start_lte},{self.mcs_lte_dict[self.mcs_lte]},2,1,{self.tx_level}')
        logger.info(f'TX_PATH: {self.tx_path}, BW: {self.bw_lte}, TX_FREQ: {self.tx_freq_lte}, RB_SIZE: {self.rb_size_lte}, RB_OFFSET: {self.rb_start_lte}, MCS: {self.mcs_lte}, TX_LEVEL: {self.tx_level}')

    def tx_set_fr1(self):
        logger.info('---------Tx Set----------')
        self.command(f'AT+NTXSENDREQ={self.tx_path_dict[self.tx_path]},{self.tx_freq_fr1},{self.bw_fr1_dict[self.bw_fr1]},{self.scs_dict[self.scs]},{self.rb_size_fr1},{self.rb_start_fr1},{self.mcs_fr1_dict[self.mcs_fr1]},{self.type_dict[self.type_fr1]},{self.tx_level}')
        logger.info(f'TX_PATH: {self.tx_path}, BW: {self.bw_fr1}, TX_FREQ: {self.tx_freq_fr1}, RB_SIZE: {self.rb_size_fr1}, RB_OFFSET: {self.rb_start_fr1}, MCS: {self.mcs_fr1}, TX_LEVEL: {self.tx_level}')

    def antenna_switch(self):  # 1: AS ON, 0: AS OFF
        logger.info('---------Antenna Switch----------')
        self.command(f'AT+LTXASTUNESET={self.asw_on_off}')
        if self.asw_on_off == 0:
            logger.info('Antenna Switch OFF')
        elif self.asw_on_off == 1:
            logger.info('Antenna Switch ON')

    def antenna_switch_v2(self):
        """
        this is to place on the first to activate
        AT+ANTSWSEL=P0,P1	//Set Tx DPDT switch
        P0: RAT (0=GSM, 1=WCDMA, 2=LTE, 4=CDMA, 6=NR)
        P1: ANT path (0=default, 1=switched, 4=dynamic mode),
        P1 (P0=NR): ANT Path (0=Tx-Ant1, 1=Tx-Ant2, 2=Tx-Ant3, 3=Tx-Ant4, 4=dynamic switch mode)
        :param tech:
        :param ant_path:
        :return:
        """
        self.asw_tech = self.tech
        logger.info('---------Antenna Switch----------')
        self.command(f'AT+ANTSWSEL={self.asw_tech_dict[self.asw_tech]},{self.asw_path}')
        logger.info(f'RAT: {self.asw_tech}, ANT_PATH: {self.asw_path}')

    def tx_measure_lte(self):
        logger.info('---------Tx Measure----------')
        mode = 'TDD' if self.band_lte in [38, 39, 40, 41, 42, 48] else 'FDD'
        self.command_cmw100_write(f'CONF:LTE:MEAS:DMODe {mode}')
        self.command_cmw100_write(f'CONF:LTE:MEAS:BAND OB{self.band_lte}')
        self.command_cmw100_write(f'CONF:LTE:MEAS:RFS:FREQ {self.tx_freq_lte}KHz')
        self.command_cmw100_query('*OPC?')
        rb = f'0{self.bw_lte*10}' if self.bw_lte < 10 else f'{self.bw_lte*10}'
        self.command_cmw100_write(f'CONF:LTE:MEAS:MEV:CBAN B{rb}')
        self.command_cmw100_write(f'CONF:LTE:MEAS:MEV:MOD:MSCH {self.mcs_lte}')
        self.command_cmw100_write(f'CONF:LTE:MEAS:MEV:RBAL:NRB {self.rb_size_lte}')
        self.command_cmw100_write(f'CONF:LTE:MEAS:MEV:RBAL:ORB {self.rb_start_lte}')
        self.command_cmw100_write(f'CONF:LTE:MEAS:MEV:CPR NORM')
        self.command_cmw100_write(f'CONF:LTE:MEAS:MEV:PLC 0')
        self.command_cmw100_write(f'CONF:LTE:MEAS:MEV:DSSP 0')
        self.command_cmw100_write(f'CONF:LTE:MEAS:MEV:RBAL:AUTO OFF')
        self.command_cmw100_write(f'CONF:LTE:MEAS:MEV:MOEX ON')
        self.command_cmw100_write(f'CONF:LTE:MEAS:MEV:LIM:SEM:LIM1:CBAN100 ON,0MHz,1MHz,-18,K030')
        self.command_cmw100_write(f'CONF:LTE:MEAS:MEV:LIM:SEM:LIM2:CBAN100 ON,1MHz,2.5MHz,-10,M1')
        self.command_cmw100_write(f'CONF:LTE:MEAS:MEV:LIM:SEM:LIM3:CBAN100 ON,2.5MHz,2.8MHz,-10,M1')
        self.command_cmw100_write(f'CONF:LTE:MEAS:MEV:LIM:SEM:LIM4:CBAN100 ON,2.8MHz,5MHz,-10,M1')
        self.command_cmw100_write(f'CONF:LTE:MEAS:MEV:LIM:SEM:LIM5:CBAN100 ON,5MHz,6MHz,-13,M1')
        self.command_cmw100_write(f'CONF:LTE:MEAS:MEV:LIM:SEM:LIM6:CBAN100 ON,6MHz,10MHz,-13,M1')
        self.command_cmw100_write(f'CONF:LTE:MEAS:MEV:LIM:SEM:LIM7:CBAN100 ON,10MHz,15MHz,-25,M1')
        self.command_cmw100_write(f'CONF:LTE:MEAS:MEV:LIM:SEM:LIM8:CBAN100 OFF,15MHz,15MHz,-25,M1')
        self.command_cmw100_write(f'CONF:LTE:MEAS:MEV:LIM:SEM:LIM9:CBAN100 OFF,15MHz,15MHz,-25,M1')
        self.command_cmw100_query('SYST:ERR:ALL?')
        self.command_cmw100_write(f'CONFigure:LTE:MEAS:MEValuation:MSLot ALL')
        self.command_cmw100_write(f'CONF:LTE:MEAS:RFS:UMAR 10.000000')
        self.command_cmw100_write(f'CONF:LTE:MEAS:RFS:ENP {self.tx_level+5}.00')
        self.command_cmw100_write(f'ROUT:LTE:MEAS:SCEN:SAL R11, RX1')
        self.command_cmw100_query('*OPC?')
        self.command_cmw100_write(f'CONF:LTE:MEAS:RFS:UMAR 10.000000')
        self.command_cmw100_write(f'CONF:LTE:MEAS:MEV:RBAL:AUTO ON')
        self.command_cmw100_write(f'CONF:LTE:MEAS:MEV:SCO:MOD 5')
        self.command_cmw100_query('*OPC?')
        self.command_cmw100_write(f'CONF:LTE:MEAS:MEV:SCO:SPEC:ACLR 5')
        self.command_cmw100_query('*OPC?')
        self.command_cmw100_write(f'CONF:LTE:MEAS:MEV:SCO:SPEC:SEM 5')
        self.command_cmw100_query('*OPC?')
        self.command_cmw100_write(f"TRIG:LTE:MEAS:MEV:SOUR 'GPRF Gen1: Restart Marker'")
        self.command_cmw100_write(f'TRIG:LTE:MEAS:MEV:THR -20.0')
        self.command_cmw100_write(f'CONF:LTE:MEAS:MEV:REP SING')
        self.command_cmw100_write(f'CONF:LTE:MEAS:MEV:RES:ALL ON, ON, ON, ON, ON, ON, ON, ON, ON, ON')
        self.command_cmw100_query('*OPC?')
        self.command_cmw100_write(f'CONF:LTE:MEAS:MEV:MSUB 2, 10, 0')
        self.command_cmw100_write(f'CONF:LTE:MEAS:SCEN:ACT SAL')
        self.command_cmw100_query('SYST:ERR:ALL?')
        self.command_cmw100_write(f'ROUT:GPRF:MEAS:SCEN:SAL R1{self.port}, RX1')
        self.command_cmw100_query('*OPC?')
        self.command_cmw100_write(f'ROUT:LTE:MEAS:SCEN:SAL R1{self.port}, RX1')
        self.command_cmw100_query('*OPC?')
        self.command_cmw100_write(f'CONF:LTE:MEAS:RFS:EATT {self.loss_tx}')
        self.command_cmw100_query('*OPC?')
        time.sleep(0.2)
        mod_results = self.command_cmw100_query('READ:LTE:MEAS:MEV:MOD:AVER?')  # P3 is EVM, P15 is Ferr, P14 is IQ Offset
        mod_results = mod_results.split(',')
        mod_results = [mod_results[3], mod_results[15], mod_results[14]]
        mod_results = [eval(m) for m in mod_results]
        logger.info(f'EVM: {mod_results[0]:.2f}, FREQ_ERR: {mod_results[1]:.2f}, IQ_OFFSET: {mod_results[2]:.2f}')
        self.command_cmw100_write(f'INIT:LTE:MEAS:MEV')
        self.command_cmw100_query('*OPC?')
        f_state = self.command_cmw100_query('FETC:LTE:MEAS:MEV:STAT?')
        while f_state != 'RDY':
            time.sleep(0.2)
            f_state = self.command_cmw100_query('FETC:LTE:MEAS:MEV:STAT?')
        aclr_results = self.command_cmw100_query('FETC:LTE:MEAS:MEV:ACLR:AVER?')
        aclr_results = aclr_results.split(',')[1:]
        aclr_results = [eval(aclr)*-1 if eval(aclr) > 30 else eval(aclr) for aclr in aclr_results]  # U_-2, U_-1, E_-1, Pwr, E_+1, U_+1, U_+2
        logger.info(f'Power: {aclr_results[3]:.2f}, E-UTRA: [{aclr_results[2]:.2f}, {aclr_results[4]:.2f}], UTRA_1: [{aclr_results[1]:.2f}, {aclr_results[5]:.2f}], UTRA_2: [{aclr_results[0]:.2f}, {aclr_results[6]:.2f}]')
        iem_results = self.command_cmw100_query('FETC:LTE:MEAS:MEV:IEM:MARG?')
        iem_results = iem_results.split(',')
        logger.info(f'InBandEmissions Margin: {eval(iem_results[2]):.2f}dB')
        # logger.info(f'IEM_MARG results: {iem_results}')
        esfl_results = self.command_cmw100_query(f'FETC:LTE:MEAS:MEV:ESFL:EXTR?')
        esfl_results = esfl_results.split(',')
        ripple1 = round(eval(esfl_results[2]),2) if esfl_results[2] != 'NCAP' else esfl_results[2]
        ripple2 = round(eval(esfl_results[3]),2) if esfl_results[3] != 'NCAP' else esfl_results[3]
        logger.info(f'Equalize Spectrum Flatness: Ripple1:{ripple1} dBpp, Ripple2:{ripple2} dBpp')
        time.sleep(0.2)
        # logger.info(f'ESFL results: {esfl_results}')
        sem_results = self.command_cmw100_query(f'FETC:LTE:MEAS:MEV:SEM:MARG?')
        logger.info(f'SEM_MARG results: {sem_results}')
        sem_avg_results = self.command_cmw100_query(f'FETC:LTE:MEAS:MEV:SEM:AVER?')
        sem_avg_results = sem_avg_results.split(',')
        logger.info(f'OBW: {eval(sem_avg_results[2])/1000000:.3f} MHz, Total TX Power: {eval(sem_avg_results[3]):.2f} dBm')
        # logger.info(f'SEM_AVER results: {sem_avg_results}')
        self.command_cmw100_write(f'STOP:LTE:MEAS:MEV')
        self.command_cmw100_query('*OPC?')

        logger.debug(aclr_results + mod_results)
        return aclr_results + mod_results  # U_-2, U_-1, E_-1, Pwr, E_+1, U_+1, U_+2, EVM, Freq_Err, IQ_OFFSET

    def tx_measure_fr1(self):
        scs = 1 if self.band_fr1 in [34, 38, 39, 40, 41, 42, 48, 75, 76, 77, 78, 79] else 0  # for now FDD is forced to 15KHz and TDD is to be 30KHz
        scs = 15 * (2 ** scs)  # for now TDD only use 30KHz, FDD only use 15KHz
        logger.info('---------Tx Measure----------')
        mode = "TDD" if self.band_fr1 in [34, 38, 39, 40, 41, 42, 48, 75, 76, 77, 78, 79] else "FDD"
        self.command_cmw100_query(f'SYSTem:BASE:OPTion:VERSion?  "CMW_NRSub6G_Meas"')
        self.command_cmw100_write(f'CONF:NRS:MEAS:MEV:DMODe {mode}')
        self.command_cmw100_write(f'CONF:NRS:MEAS:BAND OB{self.band_fr1}')
        self.command_cmw100_write(f'CONF:NRS:MEAS:RFS:FREQ {self.tx_freq_fr1}KHz')
        self.command_cmw100_query(f'*OPC?')
        self.command_cmw100_write(f'CONF:NRS:MEAS:MEV:PLC 0')
        self.command_cmw100_write(f'CONF:NRS:MEAS:MEV:MOEX ON')
        self.command_cmw100_write(f'CONF:NRS:MEAS:MEV:BWC S{scs}K, B{self.bw_fr1}')
        self.command_cmw100_write(f'CONF:NRS:MEAS:MEV:LIM:SEM:AREA1:CBAN{self.bw_fr1}   ON, 0.015MHz, 0.0985MHz, {round(-13.5-10*math.log10(self.bw_fr1/5), 1)},K030')
        self.command_cmw100_write(f'CONF:NRS:MEAS:MEV:LIM:SEM:AREA2:CBAN{self.bw_fr1}   ON,   1.5MHz,    4.5MHz,  -8.5,  M1')
        self.command_cmw100_write(f'CONF:NRS:MEAS:MEV:LIM:SEM:AREA3:CBAN{self.bw_fr1}   ON,   5.5MHz,   {round(-0.5+self.bw_fr1, 1)}MHz, -11.5,  M1')
        self.command_cmw100_write(f'CONF:NRS:MEAS:MEV:LIM:SEM:AREA4:CBAN{self.bw_fr1}   ON, 0 {round(0.5+self.bw_fr1, 1)}MHz,  {round(4.5+self.bw_fr1, 1)}MHz, -23.5,  M1')
        _256Q_flag = 2 if self.mcs_fr1 == 'Q256' else 0
        self.command_cmw100_write(f'CONFigure:NRSub:MEASurement:MEValuation:PUSChconfig {self.mcs_fr1},A,OFF,{self.rb_size_fr1},{self.rb_start_fr1},14,0,T1,SING,{_256Q_flag},2')
        type = 'ON' if self.type_fr1 == 'DFTS' else 'OFF'  # DFTS: ON, CP: OFF
        self.command_cmw100_write(f'CONFigure:NRSub:MEASurement:MEValuation:DFTPrecoding {type}')
        self.command_cmw100_write(f'CONFigure:NRSub:MEASurement:MEValuation:PCOMp OFF, 6000E+6')
        self.command_cmw100_query(f'*OPC?')
        self.command_cmw100_write(f'CONFigure:NRSub:MEASurement:MEValuation:REPetition SING')
        self.command_cmw100_write(f'CONFigure:NRSub:MEASurement:MEValuation:PLCid 0')
        self.command_cmw100_write(f'CONFigure:NRSub:MEASurement:MEValuation:CTYPe PUSC')
        self.command_cmw100_write(f'CONF:NRS:MEAS:ULDL:PER MS25')
        self.command_cmw100_write(f'CONF:NRS:MEAS:ULDL:PATT S{scs}K, 3,0,1,14')
        self.command_cmw100_write(f'CONF:NRS:MEAS:RFS:UMAR 10.000000')
        self.command_cmw100_write(f'CONF:NRS:MEAS:RFS:ENP {self.tx_level+5}.00')
        self.command_cmw100_write(f'ROUT:NRS:MEAS:SCEN:SAL R1{self.port}, RX1')
        self.command_cmw100_write(f'CONF:NRS:MEAS:RFS:UMAR 10.000000')
        self.command_cmw100_write(f'CONF:NRS:MEAS:MEV:SCO:MOD 5')
        self.command_cmw100_write(f'CONF:NRS:MEAS:MEV:SCO:SPEC:ACLR 5')
        self.command_cmw100_write(f'CONF:NRS:MEAS:MEV:SCO:SPEC:SEM 5')
        self.command_cmw100_write(f"TRIG:NRS:MEAS:MEV:SOUR 'GPRF GEN1: Restart Marker'")
        self.command_cmw100_write(f'TRIG:NRS:MEAS:MEV:THR -20.0')
        self.command_cmw100_write(f'CONF:NRS:MEAS:MEV:REP SING')
        self.command_cmw100_write(f'CONF:NRS:MEAS:MEV:RES:ALL ON, ON, ON, ON, ON, ON, ON, ON, ON, ON')
        self.command_cmw100_write(f'CONF:NRS:MEAS:MEV:NSUB 10')
        self.command_cmw100_write(f'CONFigure:NRSub:MEASurement:MEValuation:MSLot ALL')
        self.command_cmw100_write(f'CONF:NRS:MEAS:SCEN:ACT SAL')
        self.command_cmw100_write(f'CONF:NRS:MEAS:RFS:EATT {self.loss_tx}')
        self.command_cmw100_query(f'*OPC?')
        self.command_cmw100_write(f'ROUT:GPRF:MEAS:SCEN:SAL R1{self.port}, RX1')
        self.command_cmw100_query(f'*OPC?')
        self.command_cmw100_write(f'ROUT:NRS:MEAS:SCEN:SAL R1{self.port}, RX1')
        self.command_cmw100_query(f'*OPC?')
        self.command_cmw100_write(f'INIT:NRS:MEAS:MEV')
        self.command_cmw100_write(f'*OPC?')
        f_state = self.command_cmw100_query(f'FETC:NRS:MEAS:MEV:STAT?')
        while f_state != 'RDY':
            f_state = self.command_cmw100_query(f'FETC:NRS:MEAS:MEV:STAT?')
            time.sleep(0.3)
        mod_results = self.command_cmw100_query('FETC:NRS:MEAS:MEV:MOD:AVER?')  # P3 is EVM, P15 is Ferr, P14 is IQ Offset
        mod_results = mod_results.split(',')
        mod_results = [mod_results[3], mod_results[15], mod_results[14]]
        mod_results = [eval(m) for m in mod_results]
        logger.info(f'EVM: {mod_results[0]:.2f}, FREQ_ERR: {mod_results[1]:.2f}, IQ_OFFSET: {mod_results[2]:.2f}')
        aclr_results = self.command_cmw100_query('FETC:NRS:MEAS:MEV:ACLR:AVER?')
        aclr_results = aclr_results.split(',')[1:]
        aclr_results = [eval(aclr) * -1 if eval(aclr) > 30 else eval(aclr) for aclr in
                        aclr_results]  # UTRA2(-), UTRA1(-), NR(-), TxP, NR(+), UTRA1(+), UTRA2(+)
        logger.info(f'Power: {aclr_results[3]:.2f}, E-UTRA: [{aclr_results[2]:.2f}, {aclr_results[4]:.2f}], UTRA_1: [{aclr_results[1]:.2f}, {aclr_results[5]:.2f}], UTRA_2: [{aclr_results[0]:.2f}, {aclr_results[6]:.2f}]')
        iem_results = self.command_cmw100_query('FETC:NRS:MEAS:MEV:IEM:MARG:AVER?')
        iem_results = iem_results.split(',')
        iem = f'{eval(iem_results[2]):.2f}' if iem_results[2] != 'INV' else 'INV'
        logger.info(f'InBandEmissions Margin: {iem}dB')
        # logger.info(f'IEM_MARG results: {iem_results}')
        esfl_results = self.command_cmw100_query(f'FETC:NRS:MEAS:MEV:ESFL:EXTR?')
        esfl_results = esfl_results.split(',')
        ripple1 = round(eval(esfl_results[2]), 2) if esfl_results[2] != 'NCAP' else esfl_results[2]
        ripple2 = round(eval(esfl_results[3]), 2) if esfl_results[3] != 'NCAP' else esfl_results[3]
        logger.info(f'Equalize Spectrum Flatness: Ripple1:{ripple1} dBpp, Ripple2:{ripple2} dBpp')
        time.sleep(0.2)
        # logger.info(f'ESFL results: {esfl_results}')
        sem_results = self.command_cmw100_query(f'FETC:NRS:MEAS:MEV:SEM:MARG:ALL?')
        logger.info(f'SEM_MARG results: {sem_results}')
        sem_avg_results = self.command_cmw100_query(f'FETC:NRS:MEAS:MEV:SEM:AVERage?')
        sem_avg_results = sem_avg_results.split(',')
        logger.info(
            f'OBW: {eval(sem_avg_results[2]) / 1000000:.3f} MHz, Total TX Power: {eval(sem_avg_results[3]):.2f} dBm')
        # logger.info(f'SEM_AVER results: {sem_avg_results}')
        self.command_cmw100_write(f'STOP:NRS:MEAS:MEV')
        self.command_cmw100_query('*OPC?')

        logger.debug(aclr_results + mod_results)
        return aclr_results + mod_results  # U_-2, U_-1, E_-1, Pwr, E_+1, U_+1, U_+2, EVM, Freq_Err, IQ_OFFSET

    def tx_power_relative_test_export_excel(self, data, band, bw, tx_freq_level, mode=0):  # mode general: 0,  mode LMH: 1
        """
        data is dict like:
        tx_level: [ U_-2, U_-1, E_-1, Pwr, E_+1, U_+1, U_+2, EVM, Freq_Err, IQ_OFFSET]
        """
        logger.info('----------save to excel----------')
        filename = None
        if self.script == 'GENERAL':
            if mode == 1:
                if tx_freq_level >= 100:
                    filename = f'relative power_{bw}MHZ_{self.tech}_LMH.xlsx'
                elif tx_freq_level <= 100:
                    filename = f'TxP_ACLR_EVM_{bw}MHZ_{self.tech}_LMH.xlsx'


                if pathlib.Path(filename).exists() is False:
                    logger.info('----------file does not exist----------')
                    wb = openpyxl.Workbook()
                    wb.remove(wb['Sheet'])
                    # to create sheet
                    if self.tech == 'LTE':
                        for mcs in ['QPSK', 'Q16', 'Q64', 'Q256']:  # some cmw10 might not have licesnse of Q256
                            wb.create_sheet(f'Raw_Data_{mcs}_PRB')
                            wb.create_sheet(f'Raw_Data_{mcs}_FRB')

                            for sheetname in wb.sheetnames:
                                ws = wb[sheetname]
                                ws['A1'] = 'Band'
                                ws['B1'] = 'Chan'
                                ws['C1'] = 'Tx_Freq'
                                ws['D1'] = 'Tx_level'
                                ws['E1'] = 'Measured_Power'
                                ws['F1'] = 'E_-1'
                                ws['G1'] = 'E_+1'
                                ws['H1'] = 'U_-1'
                                ws['I1'] = 'U_+1'
                                ws['J1'] = 'U_-2'
                                ws['K1'] = 'U_+2'
                                ws['L1'] = 'EVM'
                                ws['M1'] = 'Freq_Err'
                                ws['N1'] = 'IQ_OFFSET'
                                ws['O1'] = 'RB_num'
                                ws['P1'] = 'RB_start'
                                ws['Q1'] = 'MCS'
                                ws['R1'] = 'Tx_Path'

                    elif self.tech == 'FR1':
                        for mcs in ['QPSK', 'Q16', 'Q64', 'Q256', 'BPSK']:  # some cmw10 might not have licesnse of Q256
                            wb.create_sheet(f'Raw_Data_{mcs}')

                            for sheetname in wb.sheetnames:
                                ws = wb[sheetname]
                                ws['A1'] = 'Band'
                                ws['B1'] = 'Chan'
                                ws['C1'] = 'Tx_Freq'
                                ws['D1'] = 'Tx_level'
                                ws['E1'] = 'Measured_Power'
                                ws['F1'] = 'E_-1'
                                ws['G1'] = 'E_+1'
                                ws['H1'] = 'U_-1'
                                ws['I1'] = 'U_+1'
                                ws['J1'] = 'U_-2'
                                ws['K1'] = 'U_+2'
                                ws['L1'] = 'EVM'
                                ws['M1'] = 'Freq_Err'
                                ws['N1'] = 'IQ_OFFSET'
                                ws['O1'] = 'RB_num'
                                ws['P1'] = 'RB_start'
                                ws['Q1'] = 'Type'
                                ws['R1'] = 'Tx_Path'
                                ws['S1'] = 'SCS(KHz)'
                                ws['T1'] = 'RB_STATE'

                    wb.save(filename)
                    wb.close()

                logger.info('----------file exist----------')
                wb = openpyxl.load_workbook(filename)
                ws = None
                if self.tech == 'LTE':
                    ws = wb[f'Raw_Data_{self.mcs_lte}_{self.rb_state}']
                elif self.tech == 'FR1':
                    ws = wb[f'Raw_Data_{self.mcs_fr1}']

                if self.tech == 'LTE':
                    max_row = ws.max_row
                    row = max_row + 1
                    if tx_freq_level >= 100:
                        for tx_level, measured_data in data.items():
                            chan = self.chan_judge_lte(band, bw, tx_freq_level)
                            ws.cell(row, 1).value = band
                            ws.cell(row, 2).value = chan  # LMH
                            ws.cell(row, 3).value = tx_freq_level  # this tx_freq_lte
                            ws.cell(row, 4).value = tx_level
                            ws.cell(row, 5).value = measured_data[3]
                            ws.cell(row, 6).value = measured_data[2]
                            ws.cell(row, 7).value = measured_data[4]
                            ws.cell(row, 8).value = measured_data[1]
                            ws.cell(row, 9).value = measured_data[5]
                            ws.cell(row, 10).value = measured_data[0]
                            ws.cell(row, 11).value = measured_data[6]
                            ws.cell(row, 12).value = measured_data[7]
                            ws.cell(row, 13).value = measured_data[8]
                            ws.cell(row, 14).value = measured_data[9]
                            ws.cell(row, 15).value = self.rb_size_lte
                            ws.cell(row, 16).value = self.rb_start_lte
                            ws.cell(row, 17).value = self.mcs_lte
                            ws.cell(row, 18).value = self.tx_path

                            row += 1

                    elif tx_freq_level <= 100:
                        for tx_freq, measured_data in data.items():
                            chan = self.chan_judge_lte(band, bw, tx_freq)
                            ws.cell(row, 1).value = band
                            ws.cell(row, 2).value = chan  # # LMH
                            ws.cell(row, 3).value = tx_freq
                            ws.cell(row, 4).value = tx_freq_level  # this tx_level
                            ws.cell(row, 5).value = measured_data[3]
                            ws.cell(row, 6).value = measured_data[2]
                            ws.cell(row, 7).value = measured_data[4]
                            ws.cell(row, 8).value = measured_data[1]
                            ws.cell(row, 9).value = measured_data[5]
                            ws.cell(row, 10).value = measured_data[0]
                            ws.cell(row, 11).value = measured_data[6]
                            ws.cell(row, 12).value = measured_data[7]
                            ws.cell(row, 13).value = measured_data[8]
                            ws.cell(row, 14).value = measured_data[9]
                            ws.cell(row, 15).value = self.rb_size_lte
                            ws.cell(row, 16).value = self.rb_start_lte
                            ws.cell(row, 17).value = self.mcs_lte
                            ws.cell(row, 18).value = self.tx_path

                            row += 1
                elif self.tech == 'FR1':
                    max_row = ws.max_row
                    row = max_row + 1
                    if tx_freq_level >= 100:
                        for tx_level, measured_data in data.items():
                            chan = self.chan_judge_fr1(band, bw, tx_freq_level)
                            ws.cell(row, 1).value = band
                            ws.cell(row, 2).value = chan  # LMH
                            ws.cell(row, 3).value = tx_freq_level  # this tx_freq_lte
                            ws.cell(row, 4).value = tx_level
                            ws.cell(row, 5).value = measured_data[3]
                            ws.cell(row, 6).value = measured_data[2]
                            ws.cell(row, 7).value = measured_data[4]
                            ws.cell(row, 8).value = measured_data[1]
                            ws.cell(row, 9).value = measured_data[5]
                            ws.cell(row, 10).value = measured_data[0]
                            ws.cell(row, 11).value = measured_data[6]
                            ws.cell(row, 12).value = measured_data[7]
                            ws.cell(row, 13).value = measured_data[8]
                            ws.cell(row, 14).value = measured_data[9]
                            ws.cell(row, 15).value = self.rb_size_fr1
                            ws.cell(row, 16).value = self.rb_start_fr1
                            ws.cell(row, 17).value = self.type_fr1
                            ws.cell(row, 18).value = self.tx_path
                            ws.cell(row, 19).value = self.scs
                            ws.cell(row, 20).value = self.rb_state

                            row += 1

                    elif tx_freq_level <= 100:
                        for tx_freq, measured_data in data.items():
                            chan = self.chan_judge_fr1(band, bw, tx_freq)
                            ws.cell(row, 1).value = band
                            ws.cell(row, 2).value = chan  # LMH
                            ws.cell(row, 3).value = tx_freq
                            ws.cell(row, 4).value = tx_freq_level  # this tx_level
                            ws.cell(row, 5).value = measured_data[3]
                            ws.cell(row, 6).value = measured_data[2]
                            ws.cell(row, 7).value = measured_data[4]
                            ws.cell(row, 8).value = measured_data[1]
                            ws.cell(row, 9).value = measured_data[5]
                            ws.cell(row, 10).value = measured_data[0]
                            ws.cell(row, 11).value = measured_data[6]
                            ws.cell(row, 12).value = measured_data[7]
                            ws.cell(row, 13).value = measured_data[8]
                            ws.cell(row, 14).value = measured_data[9]
                            ws.cell(row, 15).value = self.rb_size_fr1
                            ws.cell(row, 16).value = self.rb_start_fr1
                            ws.cell(row, 17).value = self.type_fr1
                            ws.cell(row, 18).value = self.tx_path
                            ws.cell(row, 19).value = self.scs
                            ws.cell(row, 20).value = self.rb_state

                            row += 1

                wb.save(filename)
                wb.close()

                return filename
            else:  # mode = 0
                if tx_freq_level >= 100:
                    filename = f'Tx_level_sweep_{bw}MHZ_{self.tech}.xlsx'
                elif tx_freq_level <= 100:
                    filename = f'Freq_sweep_{bw}MHZ_{self.tech}.xlsx'

                if pathlib.Path(filename).exists() is False:
                    logger.info('----------file does not exist----------')
                    wb = openpyxl.Workbook()
                    wb.remove(wb['Sheet'])
                    # to create sheet
                    if self.tech == 'LTE':
                        for mcs in ['QPSK', 'Q16', 'Q64', 'Q256']:  # some cmw10 might not have licesnse of Q256
                            wb.create_sheet(f'Raw_Data_{mcs}_PRB')
                            wb.create_sheet(f'Raw_Data_{mcs}_FRB')

                        for sheetname in wb.sheetnames:
                            ws = wb[sheetname]
                            ws['A1'] = 'Band'
                            ws['B1'] = 'Tx_Freq'
                            ws['C1'] = 'Tx_level'
                            ws['D1'] = 'Measured_Power'
                            ws['E1'] = 'E_-1'
                            ws['F1'] = 'E_+1'
                            ws['G1'] = 'U_-1'
                            ws['H1'] = 'U_+1'
                            ws['I1'] = 'U_-2'
                            ws['J1'] = 'U_+2'
                            ws['K1'] = 'EVM'
                            ws['L1'] = 'Freq_Err'
                            ws['M1'] = 'IQ_OFFSET'
                            ws['N1'] = 'RB_num'
                            ws['O1'] = 'RB_start'
                            ws['P1'] = 'MCS'
                            ws['Q1'] = 'Tx_Path'
                    elif self.tech == 'FR1':
                        for mcs in ['QPSK', 'Q16', 'Q64', 'Q256', 'BPSK']:  # some cmw10 might not have licesnse of Q256
                            wb.create_sheet(f'Raw_Data_{mcs}')

                            for sheetname in wb.sheetnames:
                                ws = wb[sheetname]
                                ws['A1'] = 'Band'
                                ws['B1'] = 'Tx_Freq'
                                ws['C1'] = 'Tx_level'
                                ws['D1'] = 'Measured_Power'
                                ws['E1'] = 'E_-1'
                                ws['F1'] = 'E_+1'
                                ws['G1'] = 'U_-1'
                                ws['H1'] = 'U_+1'
                                ws['I1'] = 'U_-2'
                                ws['J1'] = 'U_+2'
                                ws['K1'] = 'EVM'
                                ws['L1'] = 'Freq_Err'
                                ws['M1'] = 'IQ_OFFSET'
                                ws['N1'] = 'RB_num'
                                ws['O1'] = 'RB_start'
                                ws['P1'] = 'Type'
                                ws['Q1'] = 'Tx_Path'
                                ws['R1'] = 'SCS(KHz)'
                                ws['S1'] = 'RB_STATE'

                    wb.save(filename)
                    wb.close()

                logger.info('----------file exist----------')
                wb = openpyxl.load_workbook(filename)
                ws = None
                if self.tech == 'LTE':
                    ws = wb[f'Raw_Data_{self.mcs_lte}_{self.rb_state}']
                elif self.tech == 'FR1':
                    ws = wb[f'Raw_Data_{self.mcs_fr1}']

                if self.tech == 'LTE':
                    max_row = ws.max_row
                    row = max_row + 1
                    if tx_freq_level >= 100:
                        for tx_level, measured_data in data.items():
                            ws.cell(row, 1).value = band
                            ws.cell(row, 2).value = tx_freq_level  # this tx_freq
                            ws.cell(row, 3).value = tx_level
                            ws.cell(row, 4).value = measured_data[3]
                            ws.cell(row, 5).value = measured_data[2]
                            ws.cell(row, 6).value = measured_data[4]
                            ws.cell(row, 7).value = measured_data[1]
                            ws.cell(row, 8).value = measured_data[5]
                            ws.cell(row, 9).value = measured_data[0]
                            ws.cell(row, 10).value = measured_data[6]
                            ws.cell(row, 11).value = measured_data[7]
                            ws.cell(row, 12).value = measured_data[8]
                            ws.cell(row, 13).value = measured_data[9]
                            ws.cell(row, 14).value = self.rb_size_lte
                            ws.cell(row, 15).value = self.rb_start_lte
                            ws.cell(row, 16).value = self.mcs_lte
                            ws.cell(row, 17).value = self.tx_path

                            row += 1

                    elif tx_freq_level <= 100:
                        for tx_freq, measured_data in data.items():
                            ws.cell(row, 1).value = band
                            ws.cell(row, 2).value = tx_freq
                            ws.cell(row, 3).value = tx_freq_level  # this tx_level
                            ws.cell(row, 4).value = measured_data[3]
                            ws.cell(row, 5).value = measured_data[2]
                            ws.cell(row, 6).value = measured_data[4]
                            ws.cell(row, 7).value = measured_data[1]
                            ws.cell(row, 8).value = measured_data[5]
                            ws.cell(row, 9).value = measured_data[0]
                            ws.cell(row, 10).value = measured_data[6]
                            ws.cell(row, 11).value = measured_data[7]
                            ws.cell(row, 12).value = measured_data[8]
                            ws.cell(row, 13).value = measured_data[9]
                            ws.cell(row, 14).value = self.rb_size_lte
                            ws.cell(row, 15).value = self.rb_start_lte
                            ws.cell(row, 16).value = self.mcs_lte
                            ws.cell(row, 17).value = self.tx_path

                            row += 1
                elif self.tech == 'FR1':
                    max_row = ws.max_row
                    row = max_row + 1
                    if tx_freq_level >= 100:
                        for tx_level, measured_data in data.items():
                            ws.cell(row, 1).value = band
                            ws.cell(row, 2).value = tx_freq_level  # this tx_freq
                            ws.cell(row, 3).value = tx_level
                            ws.cell(row, 4).value = measured_data[3]
                            ws.cell(row, 5).value = measured_data[2]
                            ws.cell(row, 6).value = measured_data[4]
                            ws.cell(row, 7).value = measured_data[1]
                            ws.cell(row, 8).value = measured_data[5]
                            ws.cell(row, 9).value = measured_data[0]
                            ws.cell(row, 10).value = measured_data[6]
                            ws.cell(row, 11).value = measured_data[7]
                            ws.cell(row, 12).value = measured_data[8]
                            ws.cell(row, 13).value = measured_data[9]
                            ws.cell(row, 14).value = self.rb_size_fr1
                            ws.cell(row, 15).value = self.rb_start_fr1
                            ws.cell(row, 16).value = self.type_fr1
                            ws.cell(row, 17).value = self.tx_path
                            ws.cell(row, 18).value = self.scs
                            ws.cell(row, 19).value = self.rb_state

                            row += 1

                    elif tx_freq_level <= 100:
                        for tx_freq, measured_data in data.items():
                            ws.cell(row, 1).value = band
                            ws.cell(row, 2).value = tx_freq
                            ws.cell(row, 3).value = tx_freq_level  # this tx_level
                            ws.cell(row, 4).value = measured_data[3]
                            ws.cell(row, 5).value = measured_data[2]
                            ws.cell(row, 6).value = measured_data[4]
                            ws.cell(row, 7).value = measured_data[1]
                            ws.cell(row, 8).value = measured_data[5]
                            ws.cell(row, 9).value = measured_data[0]
                            ws.cell(row, 10).value = measured_data[6]
                            ws.cell(row, 11).value = measured_data[7]
                            ws.cell(row, 12).value = measured_data[8]
                            ws.cell(row, 13).value = measured_data[9]
                            ws.cell(row, 14).value = self.rb_size_fr1
                            ws.cell(row, 15).value = self.rb_start_fr1
                            ws.cell(row, 16).value = self.type_fr1
                            ws.cell(row, 17).value = self.tx_path
                            ws.cell(row, 18).value = self.scs
                            ws.cell(row, 19).value = self.rb_state

                            row += 1

                wb.save(filename)
                wb.close()

                return filename

    def tx_power_relative_test_initial_lte(self):
        logger.info('----------Relatvie test initial----------')
        mode = 'TDD' if self.band_lte in [38, 39, 40, 41, 42, 48] else 'FDD'
        self.command_cmw100_write(f'CONF:LTE:MEAS:DMODe {mode}')
        self.command_cmw100_write(f'CONF:LTE:MEAS:BAND OB{self.band_lte}')
        self.command_cmw100_write(f'CONF:LTE:MEAS:RFS:FREQ {self.tx_freq_lte}KHz')
        self.command_cmw100_query(f'*OPC?')
        rb = f'0{self.bw_lte * 10}' if self.bw_lte < 10 else f'{self.bw_lte * 10}'
        self.command_cmw100_write(f'CONF:LTE:MEAS:MEV:CBAN B{rb}')
        self.command_cmw100_write(f'CONF:LTE:MEAS:MEV:MOD:MSCH {self.mcs_lte}')
        self.command_cmw100_write(f'CONF:LTE:MEAS:MEV:RBAL:NRB {self.rb_size_lte}')
        self.command_cmw100_write(f'CONF:LTE:MEAS:MEV:RBAL:ORB {self.rb_start_lte}')
        self.command_cmw100_write(f'CONF:LTE:MEAS:MEV:CPR NORM')
        self.command_cmw100_write(f'CONF:LTE:MEAS:MEV:PLC 0')
        self.command_cmw100_write(f'CONF:LTE:MEAS:MEV:DSSP 0')
        self.command_cmw100_write(f'CONF:LTE:MEAS:MEV:RBAL:AUTO OFF')
        self.command_cmw100_write(f'CONF:LTE:MEAS:MEV:MOEX ON')
        self.command_cmw100_write(f'CONF:LTE:MEAS:MEV:LIM:SEM:LIM1:CBAN100 ON,0MHz,1MHz,-18,K030')
        self.command_cmw100_write(f'CONF:LTE:MEAS:MEV:LIM:SEM:LIM2:CBAN100 ON,1MHz,2.5MHz,-10,M1')
        self.command_cmw100_write(f'CONF:LTE:MEAS:MEV:LIM:SEM:LIM3:CBAN100 ON,2.5MHz,2.8MHz,-10,M1')
        self.command_cmw100_write(f'CONF:LTE:MEAS:MEV:LIM:SEM:LIM4:CBAN100 ON,2.8MHz,5MHz,-10,M1')
        self.command_cmw100_write(f'CONF:LTE:MEAS:MEV:LIM:SEM:LIM5:CBAN100 ON,5MHz,6MHz,-13,M1')
        self.command_cmw100_write(f'CONF:LTE:MEAS:MEV:LIM:SEM:LIM6:CBAN100 ON,6MHz,10MHz,-13,M1')
        self.command_cmw100_write(f'CONF:LTE:MEAS:MEV:LIM:SEM:LIM7:CBAN100 ON,10MHz,15MHz,-25,M1')
        self.command_cmw100_write(f'CONF:LTE:MEAS:MEV:LIM:SEM:LIM8:CBAN100 OFF,15MHz,15MHz,-25,M1')
        self.command_cmw100_write(f'CONF:LTE:MEAS:MEV:LIM:SEM:LIM9:CBAN100 OFF,15MHz,15MHz,-25,M1')
        self.command_cmw100_query(f'SYST:ERR:ALL?')
        self.command_cmw100_write(f'CONFigure:LTE:MEAS:MEValuation:MSLot ALL')
        self.command_cmw100_write(f'CONF:LTE:MEAS:RFS:UMAR 10.000000')
        self.command_cmw100_write(f'CONF:LTE:MEAS:RFS:ENP {self.tx_level+5}')
        self.command_cmw100_write(f'ROUT:LTE:MEAS:SCEN:SAL R1{self.port}, RX1')
        self.command_cmw100_query(f'*OPC?')
        self.command_cmw100_write(f'CONF:LTE:MEAS:RFS:UMAR 10.000000')
        self.command_cmw100_write(f'CONF:LTE:MEAS:MEV:RBAL:AUTO ON')
        self.command_cmw100_write(f'CONF:LTE:MEAS:MEV:SCO:MOD 5')
        self.command_cmw100_query(f'*OPC?')
        self.command_cmw100_write(f'CONF:LTE:MEAS:MEV:SCO:SPEC:ACLR 5')
        self.command_cmw100_query(f'*OPC?')
        self.command_cmw100_write(f'CONF:LTE:MEAS:MEV:SCO:SPEC:SEM 5')
        self.command_cmw100_query(f'*OPC?')
        self.command_cmw100_write(f"TRIG:LTE:MEAS:MEV:SOUR 'GPRF Gen1: Restart Marker'")
        self.command_cmw100_write(f'TRIG:LTE:MEAS:MEV:THR -20.0')
        self.command_cmw100_write(f'CONF:LTE:MEAS:MEV:REP SING')
        self.command_cmw100_write(f'CONF:LTE:MEAS:MEV:RES:ALL ON, ON, ON, ON, ON, ON, ON, ON, ON, ON')
        self.command_cmw100_query(f'*OPC?')
        self.command_cmw100_write(f'CONF:LTE:MEAS:MEV:MSUB 2, 10, 0')
        self.command_cmw100_write(f'CONF:LTE:MEAS:SCEN:ACT SAL')
        self.command_cmw100_query(f'SYST:ERR:ALL?')
        self.command_cmw100_write(f'CONF:LTE:MEAS:RFS:EATT {self.loss_tx}')
        self.command_cmw100_query(f'*OPC?')
        self.command_cmw100_write(f'ROUT:GPRF:MEAS:SCEN:SAL R1{self.port}, RX1')
        self.command_cmw100_write(f'CONF:GPRF:MEAS:POW:SCO 2')
        self.command_cmw100_write(f'CONF:GPRF:MEAS:POW:REP SING')
        self.command_cmw100_write(f'CONF:GPRF:MEAS:POW:LIST OFF')
        self.command_cmw100_write(f"TRIGger:GPRF:MEAS:POWer:SOURce 'Free Run'")
        self.command_cmw100_write(f'CONF:GPRF:MEAS:POW:TRIG:SLOP REDG')
        self.command_cmw100_write(f'CONF:GPRF:MEAS:POW:SLEN 5.0e-3')
        self.command_cmw100_write(f'CONF:GPRF:MEAS:POW:MLEN 8.0e-4')
        self.command_cmw100_write(f'TRIGger:GPRF:MEAS:POWer:OFFSet 2.1E-3')
        self.command_cmw100_write(f'TRIG:GPRF:MEAS:POW:MODE ONCE')
        self.command_cmw100_write(f'CONF:GPRF:MEAS:RFS:UMAR 10.000000')
        self.command_cmw100_write(f'CONF:GPRF:MEAS:RFS:ENP {self.tx_level+5}.00')

    def tx_power_relative_test_initial_fr1(self):
        scs = 1 if self.band_fr1 in [34, 38, 39, 40, 41, 42, 48, 75, 76, 77, 78, 79] else 0  # for now FDD is forced to 15KHz and TDD is to be 30KHz
        scs = 15 * (2 ** scs)  # for now TDD only use 30KHz, FDD only use 15KHz
        logger.info('----------Relatvie test initial----------')
        mode = 'TDD' if band_lte in [34, 38, 39, 40, 41, 42, 48, 75, 76, 77, 78, 79] else 'FDD'
        self.command_cmw100_write(f'CONF:NRS:MEAS:MEV:DMODe {mode}')
        self.command_cmw100_write(f'CONF:NRS:MEAS:BAND OB{self.band_fr1}')
        self.command_cmw100_write(f'CONF:NRS:MEAS:RFS:FREQ {self.tx_freq_fr1}KHz')
        self.command_cmw100_query(f'*OPC?')
        self.command_cmw100_write(f'CONF:NRS:MEAS:MEV:PLC 0')
        self.command_cmw100_write(f'CONF:NRS:MEAS:MEV:MOEX ON')
        self.command_cmw100_write(f'CONF:NRS:MEAS:MEV:BWC S{scs}K, B{self.bw_fr1}')
        self.command_cmw100_write(f'CONF:NRS:MEAS:MEV:LIM:SEM:AREA1:CBAN{self.bw_fr1}   ON, 0.015MHz, 0.0985MHz, {round(-13.5 - 10 * math.log10(self.bw_fr1 / 5), 1)},K030')
        self.command_cmw100_write(f'CONF:NRS:MEAS:MEV:LIM:SEM:AREA2:CBAN{self.bw_fr1}   ON,   1.5MHz,    4.5MHz,  -8.5,  M1')
        self.command_cmw100_write(f'CONF:NRS:MEAS:MEV:LIM:SEM:AREA3:CBAN{self.bw_fr1}   ON,   5.5MHz,   {round(-0.5 + self.bw_fr1, 1)}MHz, -11.5,  M1')
        self.command_cmw100_write(f'CONF:NRS:MEAS:MEV:LIM:SEM:AREA4:CBAN{self.bw_fr1}   ON, 0 {round(0.5 + self.bw_fr1, 1)}MHz,  {round(4.5 + self.bw_fr1, 1)}MHz, -23.5,  M1')
        _256Q_flag = 2 if self.mcs_fr1 == 'Q256' else 0
        self.command_cmw100_write(f'CONFigure:NRSub:MEASurement:MEValuation:PUSChconfig {self.mcs_fr1},A,OFF,{self.rb_size_fr1},{self.rb_start_fr1},14,0,T1,SING,{_256Q_flag},2')
        type = 'ON' if self.type_fr1 == 'DFTS' else 'OFF'  # DFTS: ON, CP: OFF
        self.command_cmw100_write(f'CONFigure:NRSub:MEASurement:MEValuation:DFTPrecoding {type}')
        self.command_cmw100_write(f'CONFigure:NRSub:MEASurement:MEValuation:PCOMp OFF, 6000E+6')
        self.command_cmw100_query(f'*OPC?')
        self.command_cmw100_write(f'CONFigure:NRSub:MEASurement:MEValuation:REPetition SING')
        self.command_cmw100_write(f'CONFigure:NRSub:MEASurement:MEValuation:PLCid 0')
        self.command_cmw100_write(f'CONFigure:NRSub:MEASurement:MEValuation:CTYPe PUSC')
        self.command_cmw100_write(f'CONF:NRS:MEAS:ULDL:PER MS25')
        self.command_cmw100_write(f'CONF:NRS:MEAS:ULDL:PATT S{scs}K, 3,0,1,14')
        self.command_cmw100_write(f'CONFigure:NRSub:MEASurement:MEValuation:MSLot ALL')
        self.command_cmw100_write(f'ROUT:NRS:MEAS:SCEN:SAL R1{self.port}, RX1')
        self.command_cmw100_write(f'CONF:NRS:MEAS:RFS:UMAR 10.000000')
        self.command_cmw100_write(f'CONF:NRS:MEAS:MEV:SCO:MOD 5')
        self.command_cmw100_write(f'CONF:NRS:MEAS:MEV:SCO:SPEC:ACLR 5')
        self.command_cmw100_write(f'CONF:NRS:MEAS:MEV:SCO:SPEC:SEM 5')
        self.command_cmw100_write(f"TRIG:NRS:MEAS:MEV:SOUR 'GPRF GEN1: Restart Marker'")
        self.command_cmw100_write(f'TRIG:NRS:MEAS:MEV:THR -20.0')
        self.command_cmw100_write(f'CONF:NRS:MEAS:MEV:REP SING')
        self.command_cmw100_write(f'CONF:NRS:MEAS:MEV:RES:ALL ON, ON, ON, ON, ON, ON, ON, ON, ON, ON')
        self.command_cmw100_write(f'CONF:NRS:MEAS:MEV:NSUB 10')
        self.command_cmw100_write(f'CONFigure:NRSub:MEASurement:MEValuation:MSLot ALL')
        self.command_cmw100_write(f'CONF:NRS:MEAS:SCEN:ACT SAL')
        self.command_cmw100_write(f'CONF:NRS:MEAS:RFS:EATT {self.loss_tx}')

    def tx_power_aclr_evm_lmh_pipeline_lte(self):
        self.tx_level = wt.tx_level
        self.port = wt.port_lte
        self.chan = wt.channel
        for tech in wt.tech:
            if tech == 'LTE' and wt.lte_bands != []:
                self.tech = 'LTE'
                for tx_path in wt.tx_path:
                    self.tx_path = tx_path
                    for bw in wt.lte_bandwidths:
                        self.bw_lte = bw
                        try:
                            for band in wt.lte_bands:
                                self.band_lte = band
                                if bw in cm_pmt_ftm.bandwidths_selected_lte(self.band_lte):
                                    self.tx_power_aclr_evm_lmh_lte(plot=False)
                                else:
                                    logger.info(f'B{self.band_lte} does not have BW {self.bw_lte}MHZ')
                            self.txp_aclr_evm_plot(self.filename, mode=1)  # mode=1: LMH mode
                        except TypeError:
                            logger.info(f'there is no data to plot because the band does not have this BW ')

    def tx_power_aclr_evm_lmh_pipeline_fr1(self):
        self.tx_level = wt.tx_level
        self.port = wt.port_lte
        self.chan = wt.channel
        self.sa_nsa_mode = wt.sa_nas
        for tech in wt.tech:
            if tech == 'FR1' and wt.fr1_bands != []:
                self.tech = 'FR1'
                for tx_path in wt.tx_path:
                    self.tx_path = tx_path
                    for bw in wt.fr1_bandwidths:
                        self.bw_fr1 = bw
                        try:
                            for band in wt.fr1_bands:
                                self.band_fr1 = band
                                if bw in cm_pmt_ftm.bandwidths_selected_fr1(self.band_fr1):
                                    for type in wt.type_fr1:
                                        self.type_fr1 = type
                                        self.tx_power_aclr_evm_lmh_fr1(plot=False)
                                else:
                                    logger.info(f'B{self.band_fr1} does not have BW {self.bw_fr1}MHZ')
                            self.txp_aclr_evm_plot(self.filename, mode=1)  # mode=1: LMH mode
                        except TypeError:
                            logger.info(f'there is no data to plot because the band does not have this BW ')

    def tx_freq_sweep_pipline_lte(self):
        self.tx_level = wt.tx_level
        self.port = wt.port_lte
        self.chan = wt.channel
        for tech in wt.tech:
            if tech == 'LTE' and wt.lte_bands != []:
                self.tech = tech
                for tx_path in wt.tx_path:
                    self.tx_path = tx_path
                    for bw in wt.lte_bandwidths:
                        self.bw_lte = bw
                        try:
                            for band in wt.lte_bands:
                                self.band_lte = band
                                if bw in cm_pmt_ftm.bandwidths_selected_lte(self.band_lte):
                                    self.tx_freq_sweep_progress(plot=False)
                                else:
                                    logger.info(f'B{band} does not have BW {bw}MHZ')
                            self.txp_aclr_evm_plot(self.filename, mode=0)
                        except TypeError:
                            logger.info(f'there is no data to plot because the band does not have this BW ')

    def tx_level_sweep_pipeline_lte(self):
        self.tx_level = wt.tx_level
        self.port = wt.port_lte
        self.chan = wt.channel
        for tech in wt.tech:
            if tech == 'LTE' and wt.lte_bands != []:
                self.tech = tech
                for tx_path in wt.tx_path:
                    self.tx_path = tx_path
                    for bw in wt.lte_bandwidths:
                        self.bw_lte = bw
                        try:
                            for band in wt.lte_bands:
                                self.band_lte = band
                                if bw in cm_pmt_ftm.bandwidths_selected_lte(self.band_lte):
                                    self.tx_level_sweep_progress_lte(plot=False)
                                else:
                                    logger.info(f'B{band} does not have BW {bw}MHZ')
                            self.txp_aclr_evm_plot(self.filename, mode=0)
                        except TypeError:
                            logger.info(f'there is no data to plot because the band does not have this BW ')

    def tx_measure_single(self):  # this is incompleted
        band_lte = wt.band_lte
        bw_lte = wt.bw_lte
        tx_freq_lte = wt.tx_freq_lte
        rb_num = wt.rb_num
        rb_start = wt.rb_start
        freq_select = wt.channel
        mcs = wt.mcs_lte
        tx_level = wt.tx_level
        rx_freq_list = cm_pmt_ftm.dl_freq_selected('LTE', band_lte, bw_lte)  # [L_rx_freq, M_rx_ferq, H_rx_freq]
        # tx_freq_lte_mch = cm_pmt_ftm.transfer_freq_rx2tx_lte(band_lte, rx_freq_list[1])  # M_tx_freq
        rx_loss = self.get_loss(rx_freq_list[1])
        tx_loss = self.get_loss(cm_pmt_ftm.transfer_freq_rx2tx_lte(band_lte, rx_freq_list[1]))
        logger.info('----------Test LMH progress---------')
        self.preset_instrument()
        self.set_test_end_lte()
        self.antenna_switch_v2('LTE')
        self.set_test_mode_lte(band_lte)
        self.sig_gen_lte(band_lte, rx_freq_list[1], bw_lte, rx_loss)
        self.sync_lte(rx_freq_list[1])

        freq_list = []
        for freq in freq_select:
            if freq == 'L':
                freq_list.append(cm_pmt_ftm.transfer_freq_rx2tx_lte(band_lte, rx_freq_list[0]))
            elif freq == 'M':
                freq_list.append(cm_pmt_ftm.transfer_freq_rx2tx_lte(band_lte, rx_freq_list[1]))
            elif freq == 'H':
                freq_list.append(cm_pmt_ftm.transfer_freq_rx2tx_lte(band_lte, rx_freq_list[2]))

        self.tx_set_lte(tx_path, bw_lte, tx_freq_lte, rb_num, rb_start, mcs, tx_level)
        self.tx_measure_lte(band_lte, bw_lte, tx_freq_lte, rb_num, rb_start, mcs, tx_level, rf_port, tx_loss)

    def tx_power_aclr_evm_lmh_lte(self, plot=True):
        """
        order: tx_path > bw > band > mcs > rb > chan
        :param band_lte:
        :param bw_lte:
        :param tx_level:
        :param rf_port:
        :param freq_select: 'LMH'
        :param tx_path:
        data: {tx_level: [ U_-2, U_-1, E_-1, Pwr, E_+1, U_+1, U_+2, EVM, Freq_Err, IQ_OFFSET], ...}
        """
        rx_freq_list = cm_pmt_ftm.dl_freq_selected('LTE', self.band_lte, self.bw_lte)  # [L_rx_freq, M_rx_ferq, H_rx_freq]
        self.rx_freq_lte = rx_freq_list[1]
        self.loss_rx = self.get_loss(rx_freq_list[1])
        self.loss_tx = self.get_loss(cm_pmt_ftm.transfer_freq_rx2tx_lte(self.band_lte, rx_freq_list[1]))
        logger.info('----------Test LMH progress---------')
        self.preset_instrument()
        self.set_test_end_lte()
        self.antenna_switch_v2()
        self.set_test_mode_lte()
        self.sig_gen_lte()
        self.sync_lte()

        tx_freq_select_list = []
        for chan in self.chan:
            if chan == 'L':
                tx_freq_select_list.append(cm_pmt_ftm.transfer_freq_rx2tx_lte(self.band_lte, rx_freq_list[0]))
            elif chan == 'M':
                tx_freq_select_list.append(cm_pmt_ftm.transfer_freq_rx2tx_lte(self.band_lte, rx_freq_list[1]))
            elif chan == 'H':
                tx_freq_select_list.append(cm_pmt_ftm.transfer_freq_rx2tx_lte(self.band_lte, rx_freq_list[2]))

        for mcs in wt.mcs_lte:
            self.mcs_lte = mcs
            for script in wt.scripts:
                if script == 'GENERAL':
                    self.script = script
                    for rb_ftm in wt.rb_ftm_lte:  # PRB, FRB
                        self.rb_size_lte, self.rb_start_lte = scripts.GENERAL_LTE[self.bw_lte][self.rb_select_lte_dict[rb_ftm]]  # PRB: 0, # FRB: 1
                        self.rb_state = rb_ftm  # PRB, FRB
                        data_freq = {}
                        for tx_freq_lte in tx_freq_select_list:
                            self.tx_freq_lte = tx_freq_lte
                            self.tx_set_lte()
                            aclr_mod_results = self.tx_measure_lte()
                            logger.debug(aclr_mod_results)
                            data_freq[self.tx_freq_lte] = aclr_mod_results
                        logger.debug(data_freq)
                        # ready to export to excel
                        self.filename = self.tx_power_relative_test_export_excel(data_freq, self.band_lte, self.bw_lte, self.tx_level, mode=1)  # mode=1: LMH mode
        self.set_test_end_lte()
        if plot == True:
            self.txp_aclr_evm_plot(self.filename, mode=1)  # mode=1: LMH mode
        else:
            pass

    def tx_power_aclr_evm_lmh_fr1(self, plot=True):
        """
        order: tx_path > bw > band > mcs > rb > chan
        :param band_lte:
        :param bw_lte:
        :param tx_level:
        :param rf_port:
        :param freq_select: 'LMH'
        :param tx_path:
        data: {tx_level: [ U_-2, U_-1, E_-1, Pwr, E_+1, U_+1, U_+2, EVM, Freq_Err, IQ_OFFSET], ...}
        """
        rx_freq_list = cm_pmt_ftm.dl_freq_selected('FR1', self.band_fr1, self.bw_fr1)  # [L_rx_freq, M_rx_ferq, H_rx_freq]
        self.rx_freq_fr1 = rx_freq_list[1]
        self.loss_rx = self.get_loss(rx_freq_list[1])
        self.loss_tx = self.get_loss(cm_pmt_ftm.transfer_freq_rx2tx_fr1(self.band_fr1, rx_freq_list[1]))
        logger.info('----------Test LMH progress---------')
        self.preset_instrument()
        self.set_test_end_fr1()
        self.antenna_switch_v2()
        self.set_test_mode_fr1()
        self.sig_gen_fr1()
        self.sync_fr1()

        tx_freq_select_list = []
        for chan in self.chan:
            if chan == 'L':
                tx_freq_select_list.append(cm_pmt_ftm.transfer_freq_rx2tx_fr1(self.band_fr1, rx_freq_list[0]))
            elif chan == 'M':
                tx_freq_select_list.append(cm_pmt_ftm.transfer_freq_rx2tx_fr1(self.band_fr1, rx_freq_list[1]))
            elif chan == 'H':
                tx_freq_select_list.append(cm_pmt_ftm.transfer_freq_rx2tx_fr1(self.band_fr1, rx_freq_list[2]))

        for mcs in wt.mcs_fr1:
            self.mcs_fr1 = mcs
            for script in wt.scripts:
                if script == 'GENERAL':
                    self.script = script
                    for rb_ftm in wt.rb_ftm_fr1:  # INNER_FULL, OUTER_FULL
                        self.rb_size_fr1, self.rb_start_fr1 = scripts.GENERAL_FR1[self.bw_fr1][self.scs][self.type_fr1][self.rb_alloc_fr1_dict[rb_ftm]]
                        self.rb_state = rb_ftm  # INNER_FULL, OUTER_FULL
                        data_freq = {}
                        for tx_freq_fr1 in tx_freq_select_list:
                            self.tx_freq_fr1 = tx_freq_fr1
                            self.tx_set_fr1()
                            aclr_mod_results = self.tx_measure_fr1()
                            logger.debug(aclr_mod_results)
                            data_freq[self.tx_freq_fr1] = aclr_mod_results
                        logger.debug(data_freq)
                        # ready to export to excel
                        self.filename = self.tx_power_relative_test_export_excel(data_freq, self.band_fr1, self.bw_fr1, self.tx_level, mode=1)  # mode=1: LMH mode
        self.set_test_end_fr1()
        if plot == True:
            self.txp_aclr_evm_plot(self.filename, mode=1)  # mode=1: LMH mode
        else:
            pass

    def tx_freq_sweep_progress(self, plot=True):
        """
        :param band_lte:
        :param bw_lte:
        :param tx_freq_lte:
        :param rb_num:
        :param rb_start:
        :param mcs:
        :param tx_level:
        :param rf_port:
        :param freq_range_list: [freq_level_1, freq_level_2, freq_step]
        :param tx_path:
        data: {tx_level: [ U_-2, U_-1, E_-1, Pwr, E_+1, U_+1, U_+2, EVM, Freq_Err, IQ_OFFSET], ...}
        """
        logger.info('----------Freq Sweep progress ---------')
        rx_freq_list = cm_pmt_ftm.dl_freq_selected('LTE', self.band_lte, self.bw_lte)
        tx_freq_list = [cm_pmt_ftm.transfer_freq_rx2tx_lte(self.band_lte, rx_freq) for rx_freq in rx_freq_list]
        self.rx_freq_lte = rx_freq_list[1]
        self.loss_rx = self.get_loss(rx_freq_list[1])
        self.loss_tx = self.get_loss(cm_pmt_ftm.transfer_freq_tx2rx_lte(self.band_lte, tx_freq_list[1]))
        self.preset_instrument()
        self.set_test_end_lte()
        self.antenna_switch_v2()
        self.set_test_mode_lte()
        self.sig_gen_lte()
        self.sync_lte()

        tx_freq_select_list = []
        for chan in self.chan:
            if chan == 'L':
                tx_freq_select_list.append(cm_pmt_ftm.transfer_freq_rx2tx_lte(self.band_lte, rx_freq_list[0]))
            elif chan == 'M':
                tx_freq_select_list.append(cm_pmt_ftm.transfer_freq_rx2tx_lte(self.band_lte, rx_freq_list[1]))
            elif chan == 'H':
                tx_freq_select_list.append(cm_pmt_ftm.transfer_freq_rx2tx_lte(self.band_lte, rx_freq_list[2]))

        freq_range_list = [tx_freq_list[0], tx_freq_list[2], 1000]
        step = freq_range_list[2]

        for mcs in wt.mcs_lte:
            self.mcs_lte = mcs
            for script in wt.scripts:
                if script == 'GENERAL':
                    self.script = script
                    for rb_ftm in wt.rb_ftm_lte:  # PRB, FRB
                        self.rb_size_lte, self.rb_start_lte = scripts.GENERAL_LTE[self.bw_lte][self.rb_select_lte_dict[rb_ftm]]  # PRB: 0, # FRB: 1
                        self.rb_state = rb_ftm  # PRB, FRB
                        data = {}
                        for tx_freq_lte in range(freq_range_list[0], freq_range_list[1] + step, step):
                            self.tx_freq_lte = tx_freq_lte
                            self.tx_set_lte()
                            aclr_mod_results = self.tx_measure_lte()
                            logger.debug(aclr_mod_results)
                            data[self.tx_freq_lte] = aclr_mod_results
                        logger.debug(data)
                        self.filename = self.tx_power_relative_test_export_excel(data, self.band_lte, self.bw_lte, self.tx_level, mode=0)
        self.set_test_end_lte()
        if plot == True:
            self.txp_aclr_evm_plot(self.filename, mode=0)
        else:
            pass

    def tx_level_sweep_progress_lte(self, plot=True):
        """
        :param band_lte:
        :param bw_lte:
        :param tx_freq_lte:
        :param rb_num:
        :param rb_start:
        :param mcs:
        :param pwr:
        :param rf_port:
        :param loss:
        :param tx_path:
        data {tx_level: [ U_-2, U_-1, E_-1, Pwr, E_+1, U_+1, U_+2, EVM, Freq_Err, IQ_OFFSET], ...}
        """
        rx_freq_list = cm_pmt_ftm.dl_freq_selected('LTE', self.band_lte, self.bw_lte)
        tx_freq_list = [cm_pmt_ftm.transfer_freq_rx2tx_lte(self.band_lte, rx_freq) for rx_freq in rx_freq_list]
        self.rx_freq_lte = rx_freq_list[1]
        self.tx_freq_lte = tx_freq_list[1]
        self.loss_rx = self.get_loss(rx_freq_list[1])
        self.loss_tx = self.get_loss(cm_pmt_ftm.transfer_freq_tx2rx_lte(self.band_lte, tx_freq_list[1]))
        self.preset_instrument()
        self.set_test_end_lte()
        self.antenna_switch_v2()
        self.set_test_mode_lte()
        self.sig_gen_lte()
        self.sync_lte()

        for mcs in wt.mcs_lte:
            self.mcs_lte = mcs
            for script in wt.scripts:
                if script == 'GENERAL':
                    self.script = script
                    for rb_ftm in wt.rb_ftm_lte:  # PRB, FRB
                        self.rb_size_lte, self.rb_start_lte = scripts.GENERAL_LTE[self.bw_lte][self.rb_select_lte_dict[rb_ftm]]  # PRB: 0, # FRB: 1
                        self.rb_state = rb_ftm  # PRB, FRB

                        self.tx_set_lte()
                        tx_freq_select_list = []
                        for chan in self.chan:
                            if chan == 'L':
                                tx_freq_select_list.append(cm_pmt_ftm.transfer_freq_rx2tx_lte(self.band_lte, rx_freq_list[0]))
                            elif chan == 'M':
                                tx_freq_select_list.append(cm_pmt_ftm.transfer_freq_rx2tx_lte(self.band_lte, rx_freq_list[1]))
                            elif chan == 'H':
                                tx_freq_select_list.append(cm_pmt_ftm.transfer_freq_rx2tx_lte(self.band_lte, rx_freq_list[2]))

                        #  initial all before tx level prgress
                        for tx_freq_select in tx_freq_select_list:
                            self.tx_freq_lte = tx_freq_select
                            self.tx_power_relative_test_initial_lte()

                        tx_range_list = wt.tx_level_range_list  # [tx_level_1, tx_level_2]

                        logger.info('----------TX Level Sweep progress---------')
                        logger.info(f'----------from {tx_range_list[0]} dBm to {tx_range_list[1]} dBm----------')

                        step = -1 if tx_range_list[0] > tx_range_list[1] else 1

                        #  following is real change tx level prgress


                        data = {}
                        for tx_level in range(tx_range_list[0], tx_range_list[1]+step, step):
                            self.tx_level = tx_level
                            logger.info(f'========Now Tx level = {self.tx_level} dBm========')
                            self.command(f'AT+LTXPWRLVLSET={self.tx_level}')
                            self.command(f'AT+LTXCHNSDREQ')
                            self.command_cmw100_write('CONF:LTE:MEAS:RFS:UMAR 10.000000')
                            self.command_cmw100_write(f'CONF:LTE:MEAS:RFS:ENP {self.tx_level + 5}.00')
                            mod_results = self.command_cmw100_query(f'READ:LTE:MEAS:MEV:MOD:AVER?')
                            mod_results = mod_results.split(',')
                            mod_results = [mod_results[3], mod_results[15], mod_results[14]]
                            mod_results = [eval(m) for m in mod_results]
                            logger.info(f'mod_results = {mod_results}')
                            self.command_cmw100_write('INIT:LTE:MEAS:MEV')
                            self.command_cmw100_query('*OPC?')
                            f_state = self.command_cmw100_query('FETC:LTE:MEAS:MEV:STAT?')
                            while f_state != 'RDY':
                                time.sleep(0.2)
                                f_state = self.command_cmw100_query('FETC:LTE:MEAS:MEV:STAT?')
                            aclr_results = self.command_cmw100_query('FETC:LTE:MEAS:MEV:ACLR:AVER?')
                            aclr_results = aclr_results.split(',')[1:]
                            aclr_results = [eval(aclr) * -1 if eval(aclr) > 30 else eval(aclr) for aclr in
                                            aclr_results]  # U_-2, U_-1, E_-1, Pwr, E_+1, U_+1, U_+2
                            logger.info(
                                f'Power: {aclr_results[3]:.2f}, E-UTRA: [{aclr_results[2]:.2f}, {aclr_results[4]:.2f}], UTRA_1: [{aclr_results[1]:.2f}, {aclr_results[5]:.2f}], UTRA_2: [{aclr_results[0]:.2f}, {aclr_results[6]:.2f}]')
                            iem_results = self.command_cmw100_query('FETC:LTE:MEAS:MEV:IEM:MARG?')
                            iem_results = iem_results.split(',')
                            logger.info(f'InBandEmissions Margin: {eval(iem_results[2]):.2f}dB')
                            # logger.info(f'IEM_MARG results: {iem_results}')
                            esfl_results = self.command_cmw100_query(f'FETC:LTE:MEAS:MEV:ESFL:EXTR?')
                            esfl_results = esfl_results.split(',')
                            ripple1 = f'{eval(esfl_results[2]):.2f}' if esfl_results[2] != 'NCAP' else 'NCAP'
                            ripple2 = f'{eval(esfl_results[3]):.2f}' if esfl_results[3] != 'NCAP' else 'NCAP'
                            logger.info(
                                f'Equalize Spectrum Flatness: Ripple1:{ripple1} dBpp, Ripple2:{ripple2} dBpp')
                            time.sleep(0.2)
                            # logger.info(f'ESFL results: {esfl_results}')
                            sem_results = self.command_cmw100_query(f'FETC:LTE:MEAS:MEV:SEM:MARG?')
                            logger.info(f'SEM_MARG results: {sem_results}')
                            sem_avg_results = self.command_cmw100_query(f'FETC:LTE:MEAS:MEV:SEM:AVER?')
                            sem_avg_results = sem_avg_results.split(',')
                            logger.info(
                                f'OBW: {eval(sem_avg_results[2]) / 1000000:.3f} MHz, Total TX Power: {eval(sem_avg_results[3]):.2f} dBm')
                            # logger.info(f'SEM_AVER results: {sem_avg_results}')
                            self.command_cmw100_write(f'STOP:LTE:MEAS:MEV')
                            self.command_cmw100_query('*OPC?')

                            logger.debug(aclr_results + mod_results)
                            data[tx_level] = aclr_results + mod_results
                        logger.debug(data)
                        self.filename = self.tx_power_relative_test_export_excel(data, self.band_lte, self.bw_lte, self.tx_freq_lte)
        self.set_test_end_lte()
        if plot == True:
            self.txp_aclr_evm_plot(self.filename, mode=0)
        else:
            pass

    def tx_level_sweep_progress_fr1(self, band_fr1, bw_fr1, channel, rb_num, rb_start, mcs, type_fr1, tx_level, rf_port, tx_path='TX1', plot=True):
        """
        :param band_fr1:
        :param bw_fr1:
        :param tx_freq_fr1:
        :param rb_num:
        :param rb_start:
        :param mcs:
        :param pwr:
        :param rf_port:
        :param loss:
        :param tx_path:
        data {tx_level: [ U_-2, U_-1, E_-1, Pwr, E_+1, U_+1, U_+2, EVM, Freq_Err, IQ_OFFSET], ...}
        """
        rx_freq_list = cm_pmt_ftm.dl_freq_selected('LTE', band_fr1, bw_fr1)  # [L_rx_freq, M_rx_ferq, H_rx_freq]
        tx_freq_list = [cm_pmt_ftm.transfer_freq_rx2tx_fr1(band_fr1, rx_freq) for rx_freq in rx_freq_list]

        tx_freq_fr1 = None
        if channel == 'L':
            tx_freq_fr1 = tx_freq_list[0]
        elif channel == 'M':
            tx_freq_fr1 = tx_freq_list[1]
        elif channel == 'H':
            tx_freq_fr1 = tx_freq_list[2]

        tx_loss = self.get_loss(tx_freq_fr1)
        rx_loss = self.get_loss(cm_pmt_ftm.transfer_freq_tx2rx_fr1(band_fr1, tx_freq_fr1))
        self.preset_instrument()
        self.set_test_end_fr1()
        self.antenna_switch_v2('LTE')
        self.set_test_mode_fr1(band_fr1, 0)
        self.sig_gen_fr1(band_fr1, cm_pmt_ftm.transfer_freq_tx2rx_fr1(band_fr1, tx_freq_fr1), bw_fr1, self.scs,  self.mcs, rx_loss)
        self.sync_fr1(bw_fr1, cm_pmt_ftm.transfer_freq_tx2rx_fr1(band_fr1,tx_freq_fr1), self.scs)
        self.tx_set_fr1(tx_path, bw_fr1, tx_freq_fr1, self.scs, rb_num, rb_start, mcs, type, tx_level)

        #  initial all before tx level prgress
        self.tx_power_relative_test_initial_fr1(band_fr1, bw_fr1, type_fr1, tx_freq_fr1, self.scs, rb_num, rb_start, mcs, tx_level, rf_port, tx_loss)

        tx_range_list = wt.tx_level_range_list  # [tx_level_1, tx_level_2]

        logger.info('----------TX Level Sweep progress---------')
        logger.info(f'----------from {tx_range_list[0]} dBm to {tx_range_list[1]} dBm----------')

        step = -1 if tx_range_list[0] > tx_range_list[1] else 1

        #  following is real change tx level prgress
        data = {}
        for tx_level in range(tx_range_list[0], tx_range_list[1]+step, step):
            logger.info(f'========Now Tx level = {tx_level} dBm========')
            self.command(f'AT+NTXPWRLVLSET={tx_level}')
            self.command_cmw100_write('CONF:NRS:MEAS:RFS:UMAR 10.000000')
            self.command_cmw100_write(f'CONF:NRS:MEAS:RFS:ENP {tx_level + 5}.00')
            self.command_cmw100_write('INIT:NRS:MEAS:MEV')
            self.command_cmw100_query('*OPC?')
            f_state = self.command_cmw100_query('FETC:LTE:MEAS:MEV:STAT?')
            while f_state != 'RDY':
                time.sleep(0.2)
                f_state = self.command_cmw100_query('FETC:LTE:MEAS:MEV:STAT?')
            mod_results = self.command_cmw100_query(f'FETC:NRS:MEAS:MEV:MOD:AVER?')
            mod_results = mod_results.split(',')
            mod_results = [mod_results[3], mod_results[15], mod_results[14]]
            mod_results = [eval(m) for m in mod_results]
            logger.info(f'mod_results = {mod_results}')
            aclr_results = self.command_cmw100_query('FETC:NRS:MEAS:MEV:ACLR:AVER?')
            aclr_results = aclr_results.split(',')[1:]
            aclr_results = [eval(aclr) * -1 if eval(aclr) > 30 else eval(aclr) for aclr in
                            aclr_results]  # U_-2, U_-1, E_-1, Pwr, E_+1, U_+1, U_+2
            logger.info(
                f'Power: {aclr_results[3]:.2f}, E-UTRA: [{aclr_results[2]:.2f}, {aclr_results[4]:.2f}], UTRA_1: [{aclr_results[1]:.2f}, {aclr_results[5]:.2f}], UTRA_2: [{aclr_results[0]:.2f}, {aclr_results[6]:.2f}]')
            iem_results = self.command_cmw100_query('FETC:NRS:MEAS:MEV:IEM:MARG:AVER?')
            iem_results = iem_results.split(',')
            logger.info(f'InBandEmissions Margin: {eval(iem_results[2]):.2f}dB')
            # logger.info(f'IEM_MARG results: {iem_results}')
            esfl_results = self.command_cmw100_query(f'FETC:NRS:MEAS:MEV:ESFL:EXTR?')
            esfl_results = esfl_results.split(',')
            ripple1 = f'{eval(esfl_results[2]):.2f}' if esfl_results[2] != 'NCAP' else 'NCAP'
            ripple2 = f'{eval(esfl_results[3]):.2f}' if esfl_results[3] != 'NCAP' else 'NCAP'
            logger.info(
                f'Equalize Spectrum Flatness: Ripple1:{ripple1} dBpp, Ripple2:{ripple2} dBpp')
            time.sleep(0.2)
            # logger.info(f'ESFL results: {esfl_results}')
            sem_results = self.command_cmw100_query(f'FETC:NRS:MEAS:MEV:SEM:MARG:ALL?')
            logger.info(f'SEM_MARG results: {sem_results}')
            sem_avg_results = self.command_cmw100_query(f'FETC:NRS:MEAS:MEV:SEM:AVERage?')
            sem_avg_results = sem_avg_results.split(',')
            logger.info(
                f'OBW: {eval(sem_avg_results[2]) / 1000000:.3f} MHz, Total TX Power: {eval(sem_avg_results[3]):.2f} dBm')
            # logger.info(f'SEM_AVER results: {sem_avg_results}')
            self.command_cmw100_write(f'STOP:LTE:MEAS:MEV')
            self.command_cmw100_query('*OPC?')

            logger.debug(aclr_results + mod_results)
            data[tx_level] = aclr_results + mod_results

        self.set_test_end_fr1()
        logger.debug(data)

        self.filename = self.tx_power_relative_test_export_excel(data, band_fr1, bw_fr1, tx_freq_fr1)
        if plot == True:
            self.txp_aclr_evm_plot(self.filename)
        else:
            pass

    def txp_aclr_evm_plot(self, filename, mode=0):  # mode 0: general, mode 1: LMH
        logger.info('----------Plot Chart---------')
        wb = openpyxl.load_workbook(filename)
        if self.script == 'GENERAL':
            if self.tech == 'LTE':
                if mode == 1:
                    for ws_name in wb.sheetnames:
                        ws = wb[ws_name]

                        if ws._charts != []:  # if there is charts, delete it
                            ws._charts.clear()

                        logger.info('----------Power---------')
                        chart = LineChart()
                        chart.title = 'Power'
                        chart.y_axis.title = 'Power(dBm)'
                        chart.x_axis.title = 'Band'
                        chart.x_axis.tickLblPos = 'low'

                        chart.height = 20
                        chart.width = 32

                        y_data = Reference(ws, min_col=5, min_row=1, max_col=5, max_row=ws.max_row)
                        x_data = Reference(ws, min_col=1, min_row=2, max_col=2, max_row=ws.max_row)
                        chart.add_data(y_data, titles_from_data=True)
                        chart.set_categories(x_data)

                        chart.series[0].marker.symbol = 'circle'  # for EUTRA_+1
                        chart.series[0].marker.size = 10

                        ws.add_chart(chart, "U1")

                        logger.info('----------ACLR---------')
                        chart = LineChart()
                        chart.title = 'ACLR'
                        chart.y_axis.title = 'ACLR(dB)'
                        chart.x_axis.title = 'Band'
                        chart.x_axis.tickLblPos = 'low'
                        chart.y_axis.scaling.min = -60
                        chart.y_axis.scaling.max = -20

                        chart.height = 20
                        chart.width = 32

                        y_data = Reference(ws, min_col=6, min_row=1, max_col=11, max_row=ws.max_row)
                        x_data = Reference(ws, min_col=1, min_row=2, max_col=2, max_row=ws.max_row)
                        chart.add_data(y_data, titles_from_data=True)
                        chart.set_categories(x_data)

                        chart.series[0].marker.symbol = 'triangle'  # for EUTRA_-1
                        chart.series[0].marker.size = 10
                        chart.series[1].marker.symbol = 'circle'  # for EUTRA_+1
                        chart.series[1].marker.size = 10
                        chart.series[2].graphicalProperties.line.width = 50000  # for UTRA_-1
                        chart.series[3].graphicalProperties.line.width = 50000  # for UTRA_+1
                        chart.series[4].graphicalProperties.line.dashStyle = 'dash'  # for UTRA_-2
                        chart.series[5].graphicalProperties.line.dashStyle = 'dash'  # for UTRA_+2

                        ws.add_chart(chart, "U39")

                        logger.info('----------EVM---------')
                        chart = LineChart()
                        chart.title = 'EVM'
                        chart.y_axis.title = 'EVM(%)'
                        chart.x_axis.title = 'Band'
                        chart.x_axis.tickLblPos = 'low'

                        chart.height = 20
                        chart.width = 32

                        y_data = Reference(ws, min_col=12, min_row=1, max_col=12, max_row=ws.max_row)
                        x_data = Reference(ws, min_col=1, min_row=2, max_col=2, max_row=ws.max_row)
                        chart.add_data(y_data, titles_from_data=True)
                        chart.set_categories(x_data)

                        chart.series[0].marker.symbol = 'circle'  # for EUTRA_+1
                        chart.series[0].marker.size = 10

                        ws.add_chart(chart, "U77")

                    wb.save(filename)
                    wb.close()
                else:
                    for ws_name in wb.sheetnames:
                        ws = wb[ws_name]

                        if ws._charts != []:  # if there is charts, delete it
                            ws._charts.clear()


                        logger.info('----------Power---------')
                        chart = LineChart()
                        chart.title = 'Power'
                        chart.y_axis.title = 'Power(dBm)'
                        chart.x_axis.title = 'Band'
                        chart.x_axis.tickLblPos = 'low'

                        chart.height = 20
                        chart.width = 32

                        y_data = Reference(ws, min_col=4, min_row=1, max_col=4, max_row=ws.max_row)
                        x_data = Reference(ws, min_col=1, min_row=2, max_col=2, max_row=ws.max_row)
                        chart.add_data(y_data, titles_from_data=True)
                        chart.set_categories(x_data)

                        chart.series[0].marker.symbol = 'circle'  # for EUTRA_+1
                        chart.series[0].marker.size = 10


                        ws.add_chart(chart, "R1")


                        logger.info('----------ACLR---------')
                        chart = LineChart()
                        chart.title = 'ACLR'
                        chart.y_axis.title = 'ACLR(dB)'
                        chart.x_axis.title = 'Band'
                        chart.x_axis.tickLblPos = 'low'
                        chart.y_axis.scaling.min = -60
                        chart.y_axis.scaling.max = -20

                        chart.height = 20
                        chart.width = 32

                        y_data = Reference(ws, min_col=5, min_row=1, max_col=10, max_row=ws.max_row)
                        x_data = Reference(ws, min_col=1, min_row=2, max_col=2, max_row=ws.max_row)
                        chart.add_data(y_data, titles_from_data=True)
                        chart.set_categories(x_data)

                        chart.series[0].marker.symbol = 'triangle'  # for EUTRA_-1
                        chart.series[0].marker.size = 10
                        chart.series[1].marker.symbol = 'circle'  # for EUTRA_+1
                        chart.series[1].marker.size = 10
                        chart.series[2].graphicalProperties.line.width = 50000  # for UTRA_-1
                        chart.series[3].graphicalProperties.line.width = 50000  # for UTRA_+1
                        chart.series[4].graphicalProperties.line.dashStyle = 'dash'  # for UTRA_-2
                        chart.series[5].graphicalProperties.line.dashStyle = 'dash'  # for UTRA_+2

                        ws.add_chart(chart, "R39")

                        logger.info('----------EVM---------')
                        chart = LineChart()
                        chart.title = 'EVM'
                        chart.y_axis.title = 'EVM(%)'
                        chart.x_axis.title = 'Band'
                        chart.x_axis.tickLblPos = 'low'

                        chart.height = 20
                        chart.width = 32

                        y_data = Reference(ws, min_col=11, min_row=1, max_col=11, max_row=ws.max_row)
                        x_data = Reference(ws, min_col=1, min_row=2, max_col=2, max_row=ws.max_row)
                        chart.add_data(y_data, titles_from_data=True)
                        chart.set_categories(x_data)

                        chart.series[0].marker.symbol = 'circle'  # for EUTRA_+1
                        chart.series[0].marker.size = 10

                        ws.add_chart(chart, "R77")

                    wb.save(filename)
                    wb.close()
            elif self.tech == 'FR1':
                if mode == 1:
                    for ws_name in wb.sheetnames:
                        ws = wb[ws_name]

                        if ws._charts != []:  # if there is charts, delete it
                            ws._charts.clear()

                        logger.info('----------Power---------')
                        chart = LineChart()
                        chart.title = 'Power'
                        chart.y_axis.title = 'Power(dBm)'
                        chart.x_axis.title = 'Band'
                        chart.x_axis.tickLblPos = 'low'

                        chart.height = 20
                        chart.width = 32

                        y_data = Reference(ws, min_col=5, min_row=1, max_col=5, max_row=ws.max_row)
                        x_data = Reference(ws, min_col=1, min_row=2, max_col=2, max_row=ws.max_row)
                        chart.add_data(y_data, titles_from_data=True)
                        chart.set_categories(x_data)

                        chart.series[0].marker.symbol = 'circle'  # for EUTRA_+1
                        chart.series[0].marker.size = 10

                        ws.add_chart(chart, "U1")

                        logger.info('----------ACLR---------')
                        chart = LineChart()
                        chart.title = 'ACLR'
                        chart.y_axis.title = 'ACLR(dB)'
                        chart.x_axis.title = 'Band'
                        chart.x_axis.tickLblPos = 'low'
                        chart.y_axis.scaling.min = -60
                        chart.y_axis.scaling.max = -20

                        chart.height = 20
                        chart.width = 32

                        y_data = Reference(ws, min_col=6, min_row=1, max_col=11, max_row=ws.max_row)
                        x_data = Reference(ws, min_col=1, min_row=2, max_col=2, max_row=ws.max_row)
                        chart.add_data(y_data, titles_from_data=True)
                        chart.set_categories(x_data)

                        chart.series[0].marker.symbol = 'triangle'  # for EUTRA_-1
                        chart.series[0].marker.size = 10
                        chart.series[1].marker.symbol = 'circle'  # for EUTRA_+1
                        chart.series[1].marker.size = 10
                        chart.series[2].graphicalProperties.line.width = 50000  # for UTRA_-1
                        chart.series[3].graphicalProperties.line.width = 50000  # for UTRA_+1
                        chart.series[4].graphicalProperties.line.dashStyle = 'dash'  # for UTRA_-2
                        chart.series[5].graphicalProperties.line.dashStyle = 'dash'  # for UTRA_+2

                        ws.add_chart(chart, "U39")

                        logger.info('----------EVM---------')
                        chart = LineChart()
                        chart.title = 'EVM'
                        chart.y_axis.title = 'EVM(%)'
                        chart.x_axis.title = 'Band'
                        chart.x_axis.tickLblPos = 'low'

                        chart.height = 20
                        chart.width = 32

                        y_data = Reference(ws, min_col=12, min_row=1, max_col=12, max_row=ws.max_row)
                        x_data = Reference(ws, min_col=1, min_row=2, max_col=2, max_row=ws.max_row)
                        chart.add_data(y_data, titles_from_data=True)
                        chart.set_categories(x_data)

                        chart.series[0].marker.symbol = 'circle'  # for EUTRA_+1
                        chart.series[0].marker.size = 10

                        ws.add_chart(chart, "U77")

                    wb.save(filename)
                    wb.close()
                else:
                    for ws_name in wb.sheetnames:
                        ws = wb[ws_name]

                        if ws._charts != []:  # if there is charts, delete it
                            ws._charts.clear()

                        logger.info('----------Power---------')
                        chart = LineChart()
                        chart.title = 'Power'
                        chart.y_axis.title = 'Power(dBm)'
                        chart.x_axis.title = 'Band'
                        chart.x_axis.tickLblPos = 'low'

                        chart.height = 20
                        chart.width = 32

                        y_data = Reference(ws, min_col=4, min_row=1, max_col=4, max_row=ws.max_row)
                        x_data = Reference(ws, min_col=1, min_row=2, max_col=2, max_row=ws.max_row)
                        chart.add_data(y_data, titles_from_data=True)
                        chart.set_categories(x_data)

                        chart.series[0].marker.symbol = 'circle'  # for EUTRA_+1
                        chart.series[0].marker.size = 10

                        ws.add_chart(chart, "R1")

                        logger.info('----------ACLR---------')
                        chart = LineChart()
                        chart.title = 'ACLR'
                        chart.y_axis.title = 'ACLR(dB)'
                        chart.x_axis.title = 'Band'
                        chart.x_axis.tickLblPos = 'low'
                        chart.y_axis.scaling.min = -60
                        chart.y_axis.scaling.max = -20

                        chart.height = 20
                        chart.width = 32

                        y_data = Reference(ws, min_col=5, min_row=1, max_col=10, max_row=ws.max_row)
                        x_data = Reference(ws, min_col=1, min_row=2, max_col=2, max_row=ws.max_row)
                        chart.add_data(y_data, titles_from_data=True)
                        chart.set_categories(x_data)

                        chart.series[0].marker.symbol = 'triangle'  # for EUTRA_-1
                        chart.series[0].marker.size = 10
                        chart.series[1].marker.symbol = 'circle'  # for EUTRA_+1
                        chart.series[1].marker.size = 10
                        chart.series[2].graphicalProperties.line.width = 50000  # for UTRA_-1
                        chart.series[3].graphicalProperties.line.width = 50000  # for UTRA_+1
                        chart.series[4].graphicalProperties.line.dashStyle = 'dash'  # for UTRA_-2
                        chart.series[5].graphicalProperties.line.dashStyle = 'dash'  # for UTRA_+2

                        ws.add_chart(chart, "R39")

                        logger.info('----------EVM---------')
                        chart = LineChart()
                        chart.title = 'EVM'
                        chart.y_axis.title = 'EVM(%)'
                        chart.x_axis.title = 'Band'
                        chart.x_axis.tickLblPos = 'low'

                        chart.height = 20
                        chart.width = 32

                        y_data = Reference(ws, min_col=11, min_row=1, max_col=11, max_row=ws.max_row)
                        x_data = Reference(ws, min_col=1, min_row=2, max_col=2, max_row=ws.max_row)
                        chart.add_data(y_data, titles_from_data=True)
                        chart.set_categories(x_data)

                        chart.series[0].marker.symbol = 'circle'  # for EUTRA_+1
                        chart.series[0].marker.size = 10

                        ws.add_chart(chart, "R77")

                    wb.save(filename)
                    wb.close()

    def command_cmw100_query(self, tcpip_command):
        tcpip_response = self.cmw100.query(tcpip_command).strip()
        logger.info(f'TCPIP::<<{tcpip_command}')
        logger.info(f'TCPIP::>>{tcpip_response}')
        return tcpip_response

    def command_cmw100_write(self, tcpip_command):
        self.cmw100.write(tcpip_command)
        logger.info(f'TCPIP::<<{tcpip_command}')

    def command(self, command='at', delay=0.1):
        logger.info(f'MTM: <<{command}')
        command = command + '\r'
        self.ser.write(command.encode())
        time.sleep(delay)
        response = self.ser.readlines()
        for res in response:
            r = res.decode().split()
            if len(r) > 1:  # with more than one response
                for rr in r:
                    logger.info(f'MTM: >>{rr}')
            else:
                if r == []:  # sometimes there is not \r\n in the middle response
                    continue
                else:  # only one response
                    logger.info(f'MTM: >>{r[0]}')


        while b'OK\r\n' not in response:
            logger.info('OK is not at the end, so repeat again')
            logger.info(f'==========repeat to response again=========')
            response = self.ser.readlines()
            time.sleep(1)
            for res in response:
                r = res.decode().split()
                if len(r) > 1:  # with more than one response
                    for rr in r:
                        logger.info(f'MTM: >>{rr}')
                else:
                    if r == []:  # sometimes there is not \r\n in the middle response
                        continue
                    else:  # only one response
                        logger.info(f'MTM: >>{r[0]}')

        return response


def main():
    start = datetime.datetime.now()
    cmw100 = CMW100()

    # cmw100.tx_power_aclr_evm_lmh_pipeline_lte()
    # cmw100.tx_freq_sweep_pipline_lte()
    # cmw100.tx_level_sweep_pipeline_lte()

    cmw100.tx_power_aclr_evm_lmh_pipeline_fr1()

    stop = datetime.datetime.now()


    logger.info(f'Timer: {stop - start}')


if __name__ == '__main__':
    main()